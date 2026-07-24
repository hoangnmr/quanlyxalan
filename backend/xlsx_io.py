from __future__ import annotations

import io
import math
import re
import unicodedata
import zipfile
from copy import copy
from datetime import date, datetime
from html import escape
from pathlib import Path
from typing import Any
from xml.etree import ElementTree as ET

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
MAX_XLSX_ARCHIVE_ENTRIES = 256
MAX_XLSX_COMPRESSED_BYTES = 12 * 1024 * 1024
MAX_XLSX_UNCOMPRESSED_BYTES = 48 * 1024 * 1024
MAX_XLSX_COMPRESSION_RATIO = 100
MAX_XLSX_XML_PART_BYTES = 8 * 1024 * 1024
MAX_XLSX_SHARED_STRINGS = 50_000
MAX_XLSX_CELLS = 100_000
SAFE_IGNORED_EXTERNAL_RELATIONSHIPS = {"hyperlink", "externalLinkPath"}

VESSEL_IMPORT_TEXT_FIELDS = {
    "name", "registration_no", "registry_or_imo", "vessel_type",
    "vessel_class", "shell_material", "safety_certificate_no", "notes",
    "tracking_master_name", "tracking_master_phone",
}


def import_match_key(value: Any) -> str:
    """Accent/spacing-insensitive key used only for controlled matching."""
    text = unicodedata.normalize("NFKD", str(value or "")).replace("Đ", "D").replace("đ", "d")
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", "", text.upper())


# vessel_type ghi lại nguyên văn Công dụng theo GCN đăng ký/đăng kiểm (ví dụ
# "Chở hàng khô hoặc container") — trường tự do, KHÔNG map vào một danh mục cố
# định. Bộ này chỉ chuẩn hóa chính tả cho các cách viết "container" hay gặp
# trong chứng từ/workbook ("công te nơ", "côngtenơ"…), giữ nguyên quy ước viết
# hoa chung của import (normalize_import_text .upper()) và ngữ nghĩa "Chở..."
# của chứng từ gốc — không dịch sang danh mục "Tàu hàng khô/Tàu container".
VESSEL_TYPE_CANONICAL = {
    import_match_key("SÀ LAN"): "SÀ LAN",
    import_match_key("SALAN"): "SÀ LAN",
    import_match_key("CHỞ HÀNG KHÔ"): "CHỞ HÀNG KHÔ",
    import_match_key("CONTAINER"): "CONTAINER",
    import_match_key("CÔNG TE NƠ"): "CONTAINER",
    import_match_key("CÔNG-TÊ-NƠ"): "CONTAINER",
    import_match_key("CÔNGTENƠ"): "CONTAINER",
    import_match_key("CHỞ HÀNG KHÔ HOẶC CONTAINER"): "CHỞ HÀNG KHÔ HOẶC CONTAINER",
    import_match_key("CHỞ HÀNG KHÔ HOẶC CÔNG TE NƠ"): "CHỞ HÀNG KHÔ HOẶC CONTAINER",
    import_match_key("CHỞ HÀNG KHÔ HOẶC CÔNGTENƠ"): "CHỞ HÀNG KHÔ HOẶC CONTAINER",
}


def normalize_import_text(value: Any, *, field: str = "") -> tuple[str, str | None]:
    """Normalize imported text without guessing arbitrary Vietnamese names.

    All imported text is NFC-normalized, whitespace-collapsed and uppercased.
    Field-specific aliases correct known operational vocabulary despite missing
    accents, extra spaces or words typed together. Free-form names are never
    dictionary-guessed because an incorrect vessel name is worse than a visible
    preview warning.
    """
    source = unicodedata.normalize("NFC", str(value or ""))
    collapsed = re.sub(r"\s+", " ", source.replace("\u00a0", " ")).strip()
    normalized = collapsed.upper()
    if field == "registration_no":
        normalized = re.sub(r"\s*-\s*", "-", normalized)
    elif field == "name":
        normalized = re.sub(r"(?<=\D)(?=\d)|(?<=\d)(?=\D)", " ", normalized)
        normalized = re.sub(r"\s+", " ", normalized).strip()
    elif field == "vessel_type":
        normalized = VESSEL_TYPE_CANONICAL.get(import_match_key(normalized), normalized)

    if normalized == source:
        return normalized, None
    if normalized == collapsed.upper() and collapsed == source.strip():
        return normalized, None  # case-only normalization is expected, not noisy
    return normalized, f"Chuẩn hóa '{source}' → '{normalized}'"


def _safe_xml(content: bytes, label: str) -> ET.Element:
    if len(content) > MAX_XLSX_XML_PART_BYTES:
        raise ValueError(f"Phần XML {label} vượt quá giới hạn kích thước.")
    if b"<!DOCTYPE" in content.upper() or b"<!ENTITY" in content.upper():
        raise ValueError(f"Phần XML {label} không được chứa DTD hoặc external entity.")
    try:
        return ET.fromstring(content)
    except ET.ParseError as exc:
        raise ValueError(f"XML {label} không hợp lệ.") from exc


def _validate_archive(archive: zipfile.ZipFile) -> None:
    entries = archive.infolist()
    if not entries or len(entries) > MAX_XLSX_ARCHIVE_ENTRIES:
        raise ValueError("Workbook có số lượng entry ZIP không hợp lệ.")
    compressed = sum(info.compress_size for info in entries)
    uncompressed = sum(info.file_size for info in entries)
    if compressed > MAX_XLSX_COMPRESSED_BYTES or uncompressed > MAX_XLSX_UNCOMPRESSED_BYTES:
        raise ValueError("Workbook vượt quá giới hạn kích thước.")
    for info in entries:
        name = info.filename.replace("\\", "/")
        if name.startswith("/") or ".." in name.split("/") or info.flag_bits & 0x1:
            raise ValueError("Workbook có ZIP entry không an toàn hoặc được mã hóa.")
        if info.compress_size and info.file_size / info.compress_size > MAX_XLSX_COMPRESSION_RATIO:
            raise ValueError("Workbook có tỷ lệ nén không an toàn.")
        if not info.compress_size and info.file_size > MAX_XLSX_XML_PART_BYTES:
            raise ValueError("Workbook có ZIP entry không nén vượt giới hạn.")
        if name.endswith(".rels"):
            rels = archive.read(info)
            if b'TargetMode="External"' in rels or b"TargetMode='External'" in rels:
                root = _safe_xml(rels, name)
                for relationship in root:
                    if relationship.attrib.get("TargetMode") != "External":
                        continue
                    relation_type = relationship.attrib.get("Type", "").rsplit("/", 1)[-1]
                    if relation_type not in SAFE_IGNORED_EXTERNAL_RELATIONSHIPS:
                        raise ValueError("Workbook chứa external relationship không được hỗ trợ.")


def _col_index(ref: str) -> int:
    letters = re.match(r"[A-Z]+", ref.upper())
    value = 0
    for char in letters.group(0) if letters else "A":
        value = value * 26 + ord(char) - 64
    return value


def read_workbook(content: bytes) -> dict[str, dict[str, Any]]:
    if not content.startswith(b"PK\x03\x04"):
        raise ValueError("File không phải XLSX ZIP hợp lệ.")
    try:
        archive_context = zipfile.ZipFile(io.BytesIO(content))
    except zipfile.BadZipFile as exc:
        raise ValueError("File XLSX không hợp lệ.") from exc
    with archive_context as archive:
        _validate_archive(archive)
        names = set(archive.namelist())
        required = {"[Content_Types].xml", "xl/workbook.xml", "xl/_rels/workbook.xml.rels"}
        if not required.issubset(names):
            raise ValueError("Workbook thiếu cấu trúc XLSX bắt buộc.")
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = _safe_xml(archive.read("xl/sharedStrings.xml"), "sharedStrings")
            for item in root.findall("m:si", NS):
                shared.append("".join(node.text or "" for node in item.iter() if node.tag.endswith("}t")))
                if len(shared) > MAX_XLSX_SHARED_STRINGS:
                    raise ValueError("Workbook có quá nhiều shared strings.")

        workbook = _safe_xml(archive.read("xl/workbook.xml"), "workbook")
        rels = _safe_xml(archive.read("xl/_rels/workbook.xml.rels"), "workbook relationships")
        rel_map = {node.attrib["Id"]: node.attrib["Target"] for node in rels}
        sheets: dict[str, dict[str, Any]] = {}
        sheet_nodes = workbook.find("m:sheets", NS)
        for sheet in sheet_nodes if sheet_nodes is not None else []:
            rel_id = sheet.attrib["{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"]
            target = rel_map[rel_id].lstrip("/")
            path = target if target.startswith("xl/") else f"xl/{target}"
            if path not in names:
                raise ValueError("Workbook tham chiếu worksheet không tồn tại.")
            root = _safe_xml(archive.read(path), path)
            cells: dict[str, Any] = {}
            for cell in root.findall(".//m:c", NS):
                if len(cells) >= MAX_XLSX_CELLS:
                    raise ValueError("Worksheet có quá nhiều ô dữ liệu.")
                ref = cell.attrib.get("r", "")
                kind = cell.attrib.get("t")
                value_node = cell.find("m:v", NS)
                inline = cell.find("m:is", NS)
                raw = value_node.text if value_node is not None else None
                if kind == "s" and raw is not None:
                    index = int(raw)
                    if index < 0 or index >= len(shared):
                        raise ValueError("Workbook có shared string index không hợp lệ.")
                    value: Any = shared[index]
                elif kind == "inlineStr" and inline is not None:
                    value = "".join(node.text or "" for node in inline.iter() if node.tag.endswith("}t"))
                elif kind == "b":
                    value = raw == "1"
                else:
                    value = raw
                    if raw is not None:
                        try:
                            value = float(raw) if "." in raw else int(raw)
                        except ValueError:
                            pass
                if ref:
                    cells[ref] = value
            sheets[sheet.attrib["name"]] = cells
        return sheets


def _normalized(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).replace("Đ", "D").replace("đ", "d")
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def _cell_parts(ref: str) -> tuple[str, int]:
    match = re.fullmatch(r"([A-Z]+)(\d+)", ref.upper())
    return (match.group(1), int(match.group(2))) if match else ("", 0)


def _column_index(letters: str) -> int:
    value = 0
    for char in letters:
        value = value * 26 + ord(char) - 64
    return value


VESSEL_HEADER_ALIASES = {
    "name": ("TEN PHUONG TIEN", "TEN PTTND", "TEN TAU", "TEN SA LAN"),
    "registration_no": ("SO DANG KY", "SO DK"),
    "registry_or_imo": ("SO DANG KIEM", "IMO"),
    "vessel_type": ("LOAI PHUONG TIEN", "LOAI PT", "CONG DUNG"),
    "vessel_class": ("CAP PHUONG TIEN", "CAP PT", "VUNG HOAT DONG"),
    "shell_material": ("VAT LIEU VO",),
    "build_year": ("NAM DONG", "NAM SAN XUAT"),
    "length_m": ("CHIEU DAI", "LMAX"),
    "width_m": ("CHIEU RONG",),
    "side_height_m": ("CHIEU CAO MAN",),
    "draft_m": ("MON NUOC",),
    "deadweight_tons": ("TRONG TAI TOAN PHAN", "DWT"),
    "gross_tonnage": ("DUNG TICH", "GROSS TONNAGE", "GT"),
    "engine_power_cv": ("CONG SUAT MAY", "CONG SUAT"),
    "cargo_capacity_tons": ("KHA NANG KHAI THAC TAN", "SUC CHO HANG"),
    "container_capacity_teu": ("KHA NANG KHAI THAC TEU", "SUC CHO CONTAINER", "TEU"),
    "passenger_capacity": ("SUC CHO KHACH", "SO HANH KHACH"),
    "min_crew": ("SO THUYEN VIEN", "DINH BIEN THUYEN VIEN"),
    "safety_certificate_no": ("SO GCN ATKT", "GCNATKT", "GCN ATKT"),
    "certificate_issue_date": ("NGAY CAP GCN",),
    "certificate_expiry_date": ("NGAY HET HAN GCN", "HET HAN GCN", "GCNATKT BVMT"),
    "tracking_master_name": ("THUYEN TRUONG",),
    "tracking_master_phone": ("SO DIEN THOAI LIEN HE", "DIEN THOAI LIEN HE"),
    "notes": ("GHI CHU",),
    "updated_at": ("NGAY CAP NHAT",),
}

VESSEL_FLOAT_FIELDS = {
    "length_m", "width_m", "side_height_m", "draft_m", "deadweight_tons",
    "gross_tonnage", "engine_power_cv", "cargo_capacity_tons",
    "container_capacity_teu",
}
VESSEL_INTEGER_FIELDS = {"build_year", "passenger_capacity", "min_crew"}


def _field_for_header(value: Any) -> str | None:
    label = _normalized(value)
    candidates: list[tuple[int, str]] = []
    for field, aliases in VESSEL_HEADER_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalized(alias)
            if label == normalized_alias or normalized_alias in label:
                candidates.append((len(normalized_alias), field))
    return max(candidates)[1] if candidates else None


def _numeric_from_excel(value: Any, *, integer: bool = False) -> tuple[Any, str | None]:
    """Return the first listed numeric value and flag ambiguous source text.

    Some operational workbooks store two certified configurations in one cell,
    for example ``2723.79 / 2912.57``. The database schema is scalar, so the
    first (primary) value is imported and the original cell is retained in the
    vessel notes instead of failing or silently discarding the ambiguity.
    """
    if value in (None, ""):
        return None, None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return (int(value) if integer else float(value)), None

    source = str(value).strip()
    tokens = re.findall(r"[-+]?\d[\d.,]*", source)
    if not tokens:
        return None, f"Không đọc được giá trị số: {source}"
    token = tokens[0].rstrip(".,")
    if "," in token and "." in token:
        if token.rfind(",") > token.rfind("."):
            token = token.replace(".", "").replace(",", ".")
        else:
            token = token.replace(",", "")
    elif "," in token:
        token = token.replace(",", ".")
    try:
        parsed = float(token)
    except ValueError:
        return None, f"Không đọc được giá trị số: {source}"
    warning = f"Giữ giá trị đầu tiên từ ô đa giá trị: {source}" if len(tokens) > 1 else None
    return (int(parsed) if integer else parsed), warning


def _numeric_values_from_excel(value: Any, *, integer: bool = False) -> tuple[list[Any], str | None]:
    """Read every certified numeric value from one Excel cell in source order."""
    if value in (None, ""):
        return [], None
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return [int(value) if integer else float(value)], None
    source = str(value).strip()
    tokens = re.findall(r"[-+]?\d[\d.,]*", source)
    values: list[Any] = []
    for token in tokens:
        token = token.rstrip(".,")
        if "," in token and "." in token:
            token = token.replace(".", "").replace(",", ".") if token.rfind(",") > token.rfind(".") else token.replace(",", "")
        elif "," in token:
            token = token.replace(",", ".")
        try:
            parsed = float(token)
            values.append(int(parsed) if integer else parsed)
        except ValueError:
            continue
    return values, None if values else f"Không đọc được giá trị số: {source}"


def _detect_vessel_table(sheets: dict[str, dict[str, Any]]) -> tuple[str, dict[str, Any], int, dict[str, str]]:
    best: tuple[int, str, dict[str, Any], int, dict[str, str]] | None = None
    for sheet_name, sheet in sheets.items():
        by_row: dict[int, dict[str, str]] = {}
        for ref, value in sheet.items():
            column, row_no = _cell_parts(ref)
            field = _field_for_header(value)
            if row_no and field:
                by_row.setdefault(row_no, {})[field] = column
        for row_no, mapping in by_row.items():
            score = len(mapping) + (5 if {"name", "registration_no"}.issubset(mapping) else 0)
            candidate = (score, sheet_name, sheet, row_no, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if not best or not {"name", "registration_no"}.issubset(best[4]):
        raise ValueError("Không tìm thấy bảng phương tiện có cột Tên phương tiện và Số đăng ký.")
    _, sheet_name, sheet, header_row, mapping = best
    return sheet_name, sheet, header_row, mapping


def _value_after_label(sheet: dict[str, Any], aliases: tuple[str, ...]) -> Any:
    normalized_aliases = tuple(_normalized(alias) for alias in aliases)
    for ref, value in sheet.items():
        label = _normalized(value)
        if not any(alias == label or alias in label for alias in normalized_aliases):
            continue
        column, row_no = _cell_parts(ref)
        column_index = _column_index(column)
        for offset in range(1, 4):
            candidate = sheet.get(f"{_column_name(column_index + offset)}{row_no}")
            if candidate not in (None, ""):
                return candidate
    return ""


def vessel_rows(sheets: dict[str, dict[str, Any]]) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    sheet_name, sheet, header_row, mapping = _detect_vessel_table(sheets)
    organization = {
        "name": _value_after_label(sheet, ("Tên doanh nghiệp", "Chủ phương tiện")) or "Khách hàng import",
        "tax_code": _value_after_label(sheet, ("Mã số thuế",)),
        "address": _value_after_label(sheet, ("Địa chỉ",)),
        "contact_name": _value_after_label(sheet, ("Người liên hệ",)),
        "phone": _value_after_label(sheet, ("Điện thoại",)),
    }
    for field in ("name", "tax_code", "address", "contact_name"):
        if organization.get(field) not in (None, ""):
            organization[field] = normalize_import_text(
                organization[field], field=field
            )[0]
    max_row = max((_cell_parts(ref)[1] for ref in sheet), default=header_row)
    rows: list[dict[str, Any]] = []
    blank_run = 0
    for row_no in range(header_row + 1, min(max_row, header_row + 5000) + 1):
        row = {field: sheet.get(f"{column}{row_no}") for field, column in mapping.items()}
        if not any(value not in (None, "") for value in row.values()):
            blank_run += 1
            if blank_run >= 25 and rows:
                break
            continue
        blank_run = 0
        mapping_warnings: list[str] = []
        source_notes: list[str] = []
        raw_operating_values = {
            field: row.get(field)
            for field in ("vessel_class", "deadweight_tons", "cargo_capacity_tons")
        }
        for field in VESSEL_IMPORT_TEXT_FIELDS:
            if field not in row or row[field] in (None, ""):
                continue
            normalized, warning = normalize_import_text(row[field], field=field)
            row[field] = normalized
            if warning:
                mapping_warnings.append(f"{field}: {warning}")
        for field in ("tracking_master_name", "tracking_master_phone"):
            if field in row and row[field] is None:
                row[field] = ""
        for field in VESSEL_FLOAT_FIELDS | VESSEL_INTEGER_FIELDS:
            if field not in row or row[field] in (None, ""):
                continue
            raw_value = row[field]
            parsed, warning = _numeric_from_excel(
                raw_value, integer=field in VESSEL_INTEGER_FIELDS
            )
            row[field] = parsed
            if warning:
                mapping_warnings.append(f"{field}: {warning}")
                source_notes.append(f"{field}={raw_value}")
        areas = [
            normalize_import_text(value, field="vessel_class")[0]
            for value in re.split(r"\s*/\s*", str(raw_operating_values.get("vessel_class") or ""))
            if value.strip()
        ]
        deadweights, deadweight_warning = _numeric_values_from_excel(raw_operating_values.get("deadweight_tons"))
        capacities, capacity_warning = _numeric_values_from_excel(raw_operating_values.get("cargo_capacity_tons"))
        if deadweight_warning:
            mapping_warnings.append(f"deadweight_tons: {deadweight_warning}")
        if capacity_warning:
            mapping_warnings.append(f"cargo_capacity_tons: {capacity_warning}")
        profile_count = max(len(areas), len(deadweights), len(capacities), 0)
        non_empty_counts = [count for count in (len(areas), len(deadweights), len(capacities)) if count]
        if non_empty_counts and len(set(non_empty_counts)) > 1:
            mapping_warnings.append(
                "Vùng hoạt động, trọng tải và khả năng khai thác không có cùng số giá trị; "
                "hệ thống giữ đủ dữ liệu để nhân viên Cảng kiểm tra."
            )
        row["operating_profiles"] = [
            {
                "sequence": index + 1,
                "activity_area": areas[index] if index < len(areas) else (areas[-1] if areas else ""),
                "deadweight_tons": deadweights[index] if index < len(deadweights) else None,
                "cargo_capacity_tons": capacities[index] if index < len(capacities) else None,
            }
            for index in range(profile_count)
        ]
        for field in ("deadweight_tons", "cargo_capacity_tons"):
            source_notes = [note for note in source_notes if not note.startswith(f"{field}=")]
        mapping_warnings = [
            warning for warning in mapping_warnings
            if not (
                warning.startswith("deadweight_tons: Giữ giá trị đầu tiên")
                or warning.startswith("cargo_capacity_tons: Giữ giá trị đầu tiên")
            )
        ]
        if source_notes:
            existing_notes = str(row.get("notes") or "").strip()
            source_note = "Giá trị gốc Excel: " + "; ".join(source_notes)
            row["notes"] = " | ".join(part for part in (existing_notes, source_note) if part)
        if mapping_warnings:
            row["_mapping_warnings"] = mapping_warnings
        row["_source_row"] = row_no
        row["_source_sheet"] = sheet_name
        rows.append(row)
    return organization, rows


CREW_HEADER_ALIASES = {
    "organization_name": ("TEN DOANH NGHIEP", "DOANH NGHIEP", "CHU PHUONG TIEN"),
    "full_name": ("HO VA TEN", "HO TEN", "TEN THUYEN VIEN"),
    "crew_role": ("CHUC DANH", "CHUC VU"),
    "birth_date": ("NGAY SINH", "NAM SINH"),
    "phone": ("SO DIEN THOAI", "DIEN THOAI", "SDT"),
    "identity_no": ("CCCD", "HO CHIEU", "CCCD HO CHIEU"),
    "professional_certificate_type": ("LOAI CHUNG CHI", "CHUNG CHI CHUYEN MON"),
    "professional_certificate_no": ("SO CHUNG CHI", "SO CC"),
    "certificate_issue_date": ("NGAY CAP", "NGAY CAP CHUNG CHI"),
    "certificate_expiry_date": ("NGAY HET HAN", "HAN CHUNG CHI"),
    "notes": ("GHI CHU",),
}


def _field_for_crew_header(value: Any) -> str | None:
    label = _normalized(value)
    candidates: list[tuple[int, str]] = []
    for field, aliases in CREW_HEADER_ALIASES.items():
        for alias in aliases:
            normalized_alias = _normalized(alias)
            if label == normalized_alias or normalized_alias in label:
                candidates.append((len(normalized_alias), field))
    return max(candidates)[1] if candidates else None


def crew_rows(sheets: dict[str, dict[str, Any]]) -> list[dict[str, Any]]:
    """Read a flexible crew table for controlled PORT_STAFF imports."""
    best: tuple[int, str, dict[str, Any], int, dict[str, str]] | None = None
    for sheet_name, sheet in sheets.items():
        by_row: dict[int, dict[str, str]] = {}
        for ref, value in sheet.items():
            column, row_no = _cell_parts(ref)
            field = _field_for_crew_header(value)
            if row_no and field:
                by_row.setdefault(row_no, {})[field] = column
        for row_no, mapping in by_row.items():
            score = len(mapping) + (
                6 if {"organization_name", "full_name", "crew_role"}.issubset(mapping) else 0
            )
            candidate = (score, sheet_name, sheet, row_no, mapping)
            if best is None or candidate[0] > best[0]:
                best = candidate
    if not best or not {"organization_name", "full_name", "crew_role"}.issubset(best[4]):
        raise ValueError(
            "Không tìm thấy bảng thuyền viên có các cột Tên doanh nghiệp, Họ và tên và Chức danh."
        )

    _, sheet_name, sheet, header_row, mapping = best
    max_row = max((_cell_parts(ref)[1] for ref in sheet), default=header_row)
    rows: list[dict[str, Any]] = []
    blank_run = 0
    date_fields = {"birth_date", "certificate_issue_date", "certificate_expiry_date"}
    for row_no in range(header_row + 1, max_row + 1):
        row = {field: sheet.get(f"{column}{row_no}") for field, column in mapping.items()}
        if not any(value not in (None, "") for value in row.values()):
            blank_run += 1
            if blank_run >= 5 and rows:
                break
            continue
        blank_run = 0
        warnings: list[str] = []
        for field, value in list(row.items()):
            if value in (None, ""):
                continue
            if field in date_fields:
                row[field] = excel_date(value)
                continue
            normalized, warning = normalize_import_text(value, field=field)
            row[field] = normalized
            if warning:
                warnings.append(f"{field}: {warning}")
        row["_source_sheet"] = sheet_name
        row["_source_row"] = row_no
        row["_mapping_warnings"] = warnings
        rows.append(row)
    return rows


DECLARATION_CELLS = {
    "company_name": "C6", "declaration_date": "C7", "vessel_name": "C10",
    "registration_no": "C11", "vessel_type": "C12", "vessel_class": "C13",
    "length_m": "C14", "deadweight_tons": "C15", "gross_tonnage": "C16",
    "certificate_expiry_date": "C17", "crew_count": "C18", "passenger_count": "C19",
    "last_port": "C22", "working_port": "C23", "destination_port": "C24",
    "eta": "C25", "etd": "C26", "master_name": "C57", "master_phone": "C58",
    "departure_berth": None, "agent_ptnd_name": None, "is_passenger_call": None,
    "actual_arrival_at": None, "actual_departure_at": None,
}


DECLARATION_LABEL_ALIASES = {
    "company_name": ("Tên doanh nghiệp", "Đại lý"),
    "declaration_date": ("Ngày khai báo",),
    "vessel_name": ("Tên phương tiện", "Tên PTTND"),
    "registration_no": ("Số đăng ký",),
    "vessel_type": ("Loại phương tiện", "Công dụng"),
    "vessel_class": ("Cấp phương tiện", "Cấp PTTND"),
    "length_m": ("Chiều dài lớn nhất", "Chiều dài"),
    "deadweight_tons": ("Trọng tải toàn phần",),
    "gross_tonnage": ("Dung tích",),
    "certificate_expiry_date": ("Ngày hết hạn GCN",),
    "crew_count": ("Số thuyền viên",),
    "passenger_count": ("Số hành khách",),
    "last_port": ("Cảng rời cuối cùng", "Cảng xuất phát"),
    "working_port": ("Cảng cầu bến đến làm hàng", "Cảng đến làm hàng"),
    "departure_berth": ("Cảng cầu bến rời", "Cầu bến rời", "Vị trí rời"),
    "destination_port": ("Cảng đích", "Điểm đến cuối cùng"),
    "agent_ptnd_name": ("Đại lý PTND",),
    "is_passenger_call": ("Lượt tàu khách", "Phương tiện tàu khách"),
    "eta": ("Thời gian dự kiến cập cầu (ETB)", "Thời gian dự kiến đến", "ETB"),
    "etd": ("Thời gian dự kiến rời cầu (ETD)", "Thời gian dự kiến rời", "ETD"),
    "actual_arrival_at": ("Thời gian cập cầu thực tế (ATB)", "Thời gian đến thực tế", "ATB"),
    "actual_departure_at": ("Thời gian rời cầu thực tế (ATD)", "Thời gian rời thực tế", "ATD"),
    "master_name": ("Họ tên thuyền trưởng", "Thuyền trưởng"),
    "master_phone": ("Số điện thoại thuyền trưởng", "SĐT thuyền trưởng"),
}


def _detect_declaration_sheet(sheets: dict[str, dict[str, Any]]) -> tuple[str, dict[str, Any]]:
    for preferred in ("KHAI BÁO", "KHAI BAO"):
        if preferred in sheets:
            return preferred, sheets[preferred]
    best: tuple[int, str, dict[str, Any]] | None = None
    aliases = [alias for values in DECLARATION_LABEL_ALIASES.values() for alias in values]
    for name, sheet in sheets.items():
        score = sum(1 for value in sheet.values() if any(_normalized(alias) in _normalized(value) for alias in aliases))
        if best is None or score > best[0]:
            best = (score, name, sheet)
    if not best or best[0] < 5:
        raise ValueError("Không tìm thấy sheet khai báo có đủ nhãn nhận diện.")
    return best[1], best[2]


def declaration_row(sheets: dict[str, dict[str, Any]]) -> dict[str, Any]:
    sheet_name, sheet = _detect_declaration_sheet(sheets)
    result: dict[str, Any] = {}
    for key, fallback_cell in DECLARATION_CELLS.items():
        result[key] = _value_after_label(sheet, DECLARATION_LABEL_ALIASES[key]) or (
            sheet.get(fallback_cell) if fallback_cell else None
        )
    passenger_call = import_match_key(result.get("is_passenger_call"))
    result["is_passenger_call"] = passenger_call in {"CO", "YES", "TRUE", "1", "TAUKHACH"}
    result["unload"] = _cargo_from_cells(sheet, 30, 31, 32, 33)
    result["load"] = _cargo_from_cells(sheet, 44, 45, 46, 47)
    result["_source_sheet"] = sheet_name
    return result


def _cargo_from_cells(sheet: dict[str, Any], type_row: int, movement_row: int, name_row: int, count_row: int) -> dict[str, Any]:
    return {
        "cargo_type": sheet.get(f"C{type_row}") or "",
        "movement_type": sheet.get(f"C{movement_row}") or "",
        "cargo_name": sheet.get(f"C{name_row}") or "",
        "cont20_full": sheet.get(f"C{count_row}") or 0,
        "cont20_empty": sheet.get(f"C{count_row + 1}") or 0,
        "cont40_full": sheet.get(f"C{count_row + 2}") or 0,
        "cont40_empty": sheet.get(f"C{count_row + 3}") or 0,
        "tons": sheet.get(f"C{count_row + 7}") or 0,
    }


def make_xlsx(title: str, headers: list[str], rows: list[list[Any]]) -> bytes:
    all_rows = [[title], [], headers, *rows]
    sheet_rows = []
    for row_number, values in enumerate(all_rows, 1):
        cells = []
        for column_number, value in enumerate(values, 1):
            if value is None:
                continue
            ref = f"{_column_name(column_number)}{row_number}"
            if isinstance(value, (int, float)):
                cells.append(f'<c r="{ref}"><v>{value}</v></c>')
            else:
                cells.append(f'<c r="{ref}" t="inlineStr"><is><t>{escape(str(value))}</t></is></c>')
        sheet_rows.append(f'<row r="{row_number}">{"".join(cells)}</row>')
    sheet_xml = (
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>'
        '<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">'
        f'<sheetData>{"".join(sheet_rows)}</sheetData></worksheet>'
    )
    output = io.BytesIO()
    with zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", _content_types())
        archive.writestr("_rels/.rels", _root_rels())
        archive.writestr("xl/workbook.xml", _workbook())
        archive.writestr("xl/_rels/workbook.xml.rels", _workbook_rels())
        archive.writestr("xl/worksheets/sheet1.xml", sheet_xml)
    return output.getvalue()


def _report_table_style(
    ws, header_rows: int, column_count: int, data_rows: int, *, header_start: int = 1
) -> None:
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_end = header_start + header_rows - 1
    for row in ws.iter_rows(min_row=header_start, max_row=header_end, min_col=1, max_col=column_count):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10, bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
            cell.fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    for row in ws.iter_rows(
        min_row=header_end + 1,
        max_row=header_end + max(data_rows, 1),
        min_col=1,
        max_col=column_count,
    ):
        for cell in row:
            cell.font = Font(name="Times New Roman", size=10)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_end + 1}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True


def _merged_title_row(ws, row: int, text: str, *, size: int = 11, bold: bool = False) -> None:
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=16)
    cell = ws.cell(row, 1, text)
    cell.font = Font(name="Times New Roman", size=size, bold=bold)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _make_appendix1_xlsx(
    rows: list[list[Any]], report_from=None, report_to=None, reporting_unit: str = ""
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Phụ lục 1"
    _merged_title_row(ws, 1, "Phụ lục 1", size=11, bold=True)
    _merged_title_row(ws, 2, "(Kèm theo biểu mẫu báo cáo hoạt động cảng)", size=10)
    _merged_title_row(ws, 3, "KẾ HOẠCH HOẠT ĐỘNG CỦA PHƯƠNG TIỆN THỦY NỘI ĐỊA", size=14, bold=True)
    period_text = "Ngày " + (report_to.strftime("%d/%m/%Y") if report_to else "……/……/………")
    _merged_title_row(ws, 4, period_text, size=11, bold=True)
    _merged_title_row(ws, 5, f"Tên doanh nghiệp: {reporting_unit or '………………………………'}", size=11)
    _merged_title_row(ws, 6, "Ghi chú: Đối với phương tiện chở container, ghi rõ sức chở theo TEUs.", size=10)
    merges = (
        "A7:A10", "B7:H7", "I7:O7", "P7:P10",
        "B8:B10", "C8:C10", "D8:D10", "E8:E10", "F8:F10",
        "G8:H9", "I8:J8", "K8:L8", "I9:I10", "J9:J10",
        "K9:K10", "L9:L10", "M8:M10", "N8:N10", "O8:O10",
    )
    for cell_range in merges:
        ws.merge_cells(cell_range)
    values = {
        "A7": "TT", "B7": "PHƯƠNG TIỆN", "I7": "HOẠT ĐỘNG",
        "P7": "Tên và số điện thoại thuyền trưởng",
        "B8": "Tên", "C8": "Số đăng ký", "D8": "Cấp phương tiện",
        "E8": "Công dụng", "F8": "Ngày hết hạn GCNATKT&BVMT",
        "G8": "Khả năng khai thác", "G10": "Lượng hàng (tấn/TEU)",
        "H10": "Sức chở (khách)", "I8": "Đến", "I9": "Vị trí (Cảng/cầu)",
        "J9": "Thời gian (ngày, giờ)", "K8": "Rời", "K9": "Vị trí (Cảng/cầu)",
        "L9": "Thời gian (ngày, giờ)", "M8": "Hàng dỡ (loại, số lượng)",
        "N8": "Hàng xếp (loại, số lượng)", "O8": "Số thuyền viên/Hành khách",
    }
    for address, value in values.items():
        ws[address] = value
    for row_number, row in enumerate(rows, start=11):
        for column_number, value in enumerate(row, start=1):
            ws.cell(row_number, column_number, value)
    widths = [6, 18, 14, 16, 22, 16, 20, 12, 18, 18, 18, 18, 24, 24, 16, 24]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(7, 11):
        ws.row_dimensions[row_number].height = 28
    _report_table_style(ws, 4, 16, len(rows), header_start=7)
    footer_row = 11 + max(len(rows), 1) + 1
    ws.merge_cells(start_row=footer_row, start_column=1, end_row=footer_row, end_column=6)
    ws.cell(footer_row, 1, "Người lập báo cáo\n(Ký, ghi rõ họ tên)")
    ws.merge_cells(start_row=footer_row, start_column=11, end_row=footer_row, end_column=16)
    ws.cell(footer_row, 11, "Đại diện doanh nghiệp\n(Ký, ghi rõ họ tên, đóng dấu)")
    for cell in (ws.cell(footer_row, 1), ws.cell(footer_row, 11)):
        cell.font = Font(name="Times New Roman", size=11, italic=True)
        cell.alignment = Alignment(horizontal="center", vertical="top", wrap_text=True)
    ws.row_dimensions[footer_row].height = 48
    ws.print_area = f"A1:P{footer_row}"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _make_appendix2_xlsx(
    rows: list[list[Any]], report_from=None, report_to=None, reporting_unit: str = ""
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Phụ lục 2"
    _merged_title_row(ws, 1, "Phụ lục 2", size=11, bold=True)
    _merged_title_row(ws, 2, "(Kèm theo biểu mẫu báo cáo hoạt động cảng)", size=10)
    _merged_title_row(ws, 3, "BÁO CÁO KHỐI LƯỢNG HÀNG HÓA, LƯỢT TÀU VÀ HÀNH KHÁCH", size=14, bold=True)
    # strftime() không dùng được ở đây: trên Windows, ký tự Unicode ngoài các
    # %-directive đi qua locale C runtime và bị hỏng dấu (vd. "năm" -> "nam").
    # Lấy số bằng strftime, tự ghép chuỗi tiếng Việt bằng f-string thuần Python.
    month_text = f"Tháng {report_to.month:02d} năm {report_to.year}" if report_to else "Tháng ……"
    _merged_title_row(ws, 4, month_text, size=12, bold=True)
    _merged_title_row(ws, 5, f"Đơn vị báo cáo: {reporting_unit or '………………………………'}", size=11)
    merges = (
        "A7:A9", "B7:B9", "C7:F7", "G7:H7", "I7:J7", "K7:L7",
        "M7:N7", "O7:P7", "C8:D8", "E8:F8",
    )
    for cell_range in merges:
        ws.merge_cells(cell_range)
    values = {
        "A7": "STT", "B7": "Chỉ tiêu", "C7": "Container", "G7": "Hàng khô",
        "I7": "Hàng lỏng", "K7": "Hàng XNK", "M7": "Lượt tàu", "O7": "Hành khách",
        "C8": "Thực hiện tháng báo cáo", "E8": "Lũy kế đến tháng báo cáo",
        "G8": "Thực hiện tháng báo cáo", "H8": "Lũy kế đến tháng báo cáo",
        "I8": "Thực hiện tháng báo cáo", "J8": "Lũy kế đến tháng báo cáo",
        "K8": "Thực hiện tháng báo cáo", "L8": "Lũy kế đến tháng báo cáo",
        "M8": "Thực hiện tháng báo cáo", "N8": "Lũy kế đến tháng báo cáo",
        "O8": "Lượt tàu khách", "P8": "Lượt khách",
        "C9": "Tấn", "D9": "TEUs", "E9": "Tấn", "F9": "TEUs",
        "G9": "Tấn", "H9": "Tấn", "I9": "Tấn", "J9": "Tấn",
        "K9": "Tấn", "L9": "Tấn", "M9": "Lượt", "N9": "Lượt",
        "O9": "Lượt", "P9": "Lượt",
    }
    for address, value in values.items():
        ws[address] = value
    for column, value in enumerate(["A", "B", *range(1, 15)], start=1):
        ws.cell(10, column, value)
    for row_number, row in enumerate(rows, start=11):
        for column_number, value in enumerate(row, start=1):
            ws.cell(row_number, column_number, value)
    total_row = 10 + len(rows)
    if len(rows) >= 1:
        ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=2)
    widths = [6, 24, *([12] * 14)]
    for index, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(7, 11):
        ws.row_dimensions[row_number].height = 28
    _report_table_style(ws, 4, 16, len(rows), header_start=7)
    ws.print_area = f"A1:P{max(11, 10 + len(rows))}"
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def _make_appendix3_xlsx(
    rows: list[list[Any]], template_path: Path,
    report_from=None, report_to=None, reporting_unit: str = "",
) -> bytes:
    wb = load_workbook(template_path)
    ws = wb.active
    if report_to:
        ws["A2"] = (
            "(Kèm theo Văn bản số        /CVHHTPHCM-TTTT, ngày        "
            f"tháng {report_to.month} năm {report_to.year} của Cảng vụ Hàng hải Thành phố Hồ Chí Minh)"
        )
    ws["A4"] = f"Đơn vị báo cáo: {reporting_unit or '………………………………'}"
    for cell_range in ("A15:E15", "O15:T15", "A16:E16", "O16:T16"):
        if cell_range in {str(item) for item in ws.merged_cells.ranges}:
            ws.unmerge_cells(cell_range)
    ws.delete_rows(15, 2)
    desired_rows = max(len(rows), 1)
    if desired_rows > 5:
        ws.insert_rows(15, amount=desired_rows - 5)
    elif desired_rows < 5:
        ws.delete_rows(10 + desired_rows, amount=5 - desired_rows)
    source_row = 10
    for target_row in range(10, 10 + desired_rows):
        if target_row != source_row:
            for column_number in range(1, 36):
                source = ws.cell(source_row, column_number)
                target = ws.cell(target_row, column_number)
                target._style = copy(source._style)
                target.number_format = source.number_format
                target.alignment = copy(source.alignment)
                target.border = copy(source.border)
                target.fill = copy(source.fill)
                target.font = copy(source.font)
            ws.row_dimensions[target_row].height = ws.row_dimensions[source_row].height
    for target_row in range(10, 10 + desired_rows):
        for column_number in range(1, 36):
            ws.cell(target_row, column_number).value = None
    for row_number, row in enumerate(rows, start=10):
        for column_number, value in enumerate(row, start=1):
            ws.cell(row_number, column_number, value)
    for address in ("J7", "M7", "P7", "S7", "Z7"):
        ws[address] = "TEUs"
    for address in ("K7", "N7", "Q7", "T7"):
        ws[address] = "TEUs Rỗng"
    for address in ("V7", "X7"):
        ws[address] = "TEUs"
    ws["W6"] = "Quá cảnh\n(bốc dỡ)"
    ws["Y6"] = "Quá cảnh\n(không bốc dỡ)"
    ws.column_dimensions["D"].width = max(ws.column_dimensions["D"].width or 0, 18)
    for target_row in range(10, 10 + desired_rows):
        estimated_lines = 1
        for column_number in range(1, 36):
            cell = ws.cell(target_row, column_number)
            if cell.value is None or not cell.alignment.wrap_text:
                continue
            column_width = ws.column_dimensions[cell.column_letter].width or 8.43
            chars_per_line = max(int(column_width), 1)
            cell_lines = sum(
                max(1, math.ceil(len(line) / chars_per_line))
                for line in str(cell.value).splitlines() or [""]
            )
            estimated_lines = max(estimated_lines, cell_lines)
        ws.row_dimensions[target_row].height = max(
            ws.row_dimensions[target_row].height or 0,
            66,
            min(180, estimated_lines * 18),
        )
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "A10"
    ws.print_area = f"A5:AI{9 + desired_rows}"
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def make_report_xlsx(
    kind: str,
    rows: list[list[Any]],
    *,
    appendix3_template: Path | None = None,
    report_from=None,
    report_to=None,
    reporting_unit: str = "",
) -> bytes:
    """Create report workbooks with the approved appendix table structures."""
    if kind == "appendix1":
        return _make_appendix1_xlsx(rows, report_from, report_to, reporting_unit)
    if kind == "appendix2":
        return _make_appendix2_xlsx(rows, report_from, report_to, reporting_unit)
    if kind == "appendix3" and appendix3_template:
        return _make_appendix3_xlsx(
            rows, appendix3_template, report_from, report_to, reporting_unit,
        )
    raise ValueError(f"Không có cấu trúc bảng cho báo cáo {kind}.")


def _column_name(index: int) -> str:
    result = ""
    while index:
        index, remainder = divmod(index - 1, 26)
        result = chr(65 + remainder) + result
    return result


def _content_types() -> str:
    return '<?xml version="1.0"?><Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types"><Default Extension="rels" ContentType="application/vnd.openxmlformats-package.relationships+xml"/><Default Extension="xml" ContentType="application/xml"/><Override PartName="/xl/workbook.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet.main+xml"/><Override PartName="/xl/worksheets/sheet1.xml" ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml"/></Types>'


def _root_rels() -> str:
    return '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument" Target="xl/workbook.xml"/></Relationships>'


def _workbook() -> str:
    return '<?xml version="1.0"?><workbook xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"><sheets><sheet name="Bao cao" sheetId="1" r:id="rId1"/></sheets></workbook>'


def _workbook_rels() -> str:
    return '<?xml version="1.0"?><Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships"><Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet" Target="worksheets/sheet1.xml"/></Relationships>'


def excel_date(value: Any) -> Any:
    if isinstance(value, (int, float)) and value > 20000:
        return date.fromordinal(date(1899, 12, 30).toordinal() + int(value)).isoformat()
    if isinstance(value, str):
        stripped = value.strip()
        for pattern in ("%d/%m/%Y", "%d-%m-%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(stripped, pattern).date().isoformat()
            except ValueError:
                continue
    return value
