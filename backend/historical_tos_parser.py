"""Memory-bounded, filename-independent parsers for audited historical workbooks.

The parser contract is derived from the read-only R2 audit artifacts in docs/.
It never mutates a workbook, never creates live declarations/master data, reads
hidden rows/columns, and retains cell-level provenance for approved fields.
"""
from __future__ import annotations

import io
import json
import re
import unicodedata
import zipfile
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Iterable

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


MAX_WORKSHEETS = 20
MAX_COLUMNS = 100
MAX_ROWS_PER_SHEET = 200_000
MAX_ZIP_ENTRIES = 5_000
MAX_UNCOMPRESSED_BYTES = 128 * 1024 * 1024
MAX_COMPRESSION_RATIO = 100

BERTH_VERSION = "tos_berth_call_v1"
CARGO_VERSION = "tos_cargo_detail_v1"
PL03_VERSION = "reported_pl03_35col_historical_v1"
TRANSFORM_VERSION = "KBCV-HIST-TOS-1.0"


class HistoricalWorkbookError(ValueError):
    """Safe parser error suitable for a 422 API response."""


@dataclass
class ParsedWorkbook:
    source_kind: str
    mapping_version: str
    appendix_kind: str = ""
    sheet_name: str = ""
    reporting_period: str | None = None
    rows: list[dict[str, Any]] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    receipt: dict[str, Any] = field(default_factory=dict)

    @property
    def counts(self) -> dict[str, int]:
        states = Counter(row.get("validation_status", "PENDING") for row in self.rows)
        return {
            "accepted": states["VALID"],
            "review": states["REVIEW"],
            "rejected": states["REJECTED"],
            "total": len(self.rows),
        }


def normalize_token(value: Any) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).replace("Đ", "D").replace("đ", "d")
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^A-Z0-9]+", " ", text.upper()).strip()


def normalize_vessel_name(value: Any) -> str:
    return normalize_token(value).replace(" ", "")


def normalize_voyage(value: Any) -> str:
    raw = str(value or "").strip()
    if re.fullmatch(r"\d+", raw):
        return str(int(raw))
    return normalize_token(raw).replace(" ", "")


def call_key(vessel_name: Any, year: Any, voyage: Any) -> str:
    return "|".join((
        normalize_vessel_name(vessel_name), str(year or "").strip(), normalize_voyage(voyage)
    ))


def _safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (date, datetime)):
        return value.isoformat()
    return str(value)


def _raw(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M:%S")
    return str(value).strip()


def _validate_zip(content: bytes) -> None:
    try:
        with zipfile.ZipFile(io.BytesIO(content)) as archive:
            infos = archive.infolist()
            if len(infos) > MAX_ZIP_ENTRIES:
                raise HistoricalWorkbookError("Workbook có quá nhiều thành phần nén.")
            total = 0
            for info in infos:
                path = info.filename.replace("\\", "/")
                if path.startswith("/") or "../" in f"/{path}":
                    raise HistoricalWorkbookError("Workbook chứa đường dẫn không an toàn.")
                if info.flag_bits & 0x1:
                    raise HistoricalWorkbookError("Workbook mã hóa không được hỗ trợ.")
                lowered = path.lower()
                if (
                    lowered.endswith("vbaproject.bin")
                    or lowered.startswith("xl/activex/")
                    or lowered.startswith("xl/embeddings/")
                ):
                    raise HistoricalWorkbookError("Workbook chứa nội dung nhúng/chạy mã không được hỗ trợ.")
                total += info.file_size
                if total > MAX_UNCOMPRESSED_BYTES:
                    raise HistoricalWorkbookError("Workbook vượt giới hạn dữ liệu giải nén.")
                if info.compress_size and info.file_size / info.compress_size > MAX_COMPRESSION_RATIO:
                    raise HistoricalWorkbookError("Workbook có tỷ lệ nén không an toàn.")
    except zipfile.BadZipFile as exc:
        raise HistoricalWorkbookError("File không phải workbook XLSX hợp lệ.") from exc


def _header_map(values: Iterable[Any]) -> dict[str, int]:
    result: dict[str, int] = {}
    for index, value in enumerate(values, 1):
        key = normalize_token(value)
        if key and key not in result:
            result[key] = index
    return result


def _find(mapping: dict[str, int], *aliases: str) -> int | None:
    for alias in aliases:
        key = normalize_token(alias)
        if key in mapping:
            return mapping[key]
    return None


def _cell_receipt(sheet: str, row: int, columns: dict[str, int]) -> dict[str, str]:
    return {name: f"{sheet}!{get_column_letter(column)}{row}" for name, column in columns.items()}


def _parse_tos_datetime(value: Any) -> tuple[str, str | None, str | None]:
    raw = _raw(value)
    if not raw:
        return raw, None, "BLANK"
    if isinstance(value, datetime):
        return raw, value.replace(microsecond=0).isoformat(), None
    for pattern in ("%d/%m/%Y %H:%M:%S", "%d/%m/%Y %H:%M"):
        try:
            parsed = datetime.strptime(raw, pattern)
            return raw, parsed.isoformat(), None
        except ValueError:
            continue
    return raw, None, "INVALID"


def _decimal(value: Any) -> tuple[str, float | None, str]:
    raw = _raw(value)
    if raw == "":
        return raw, None, "BLANK"
    if not re.fullmatch(r"[+-]?(?:\d+(?:\.\d+)?|\.\d+)", raw):
        return raw, None, "INVALID"
    try:
        number = Decimal(raw)
    except InvalidOperation:
        return raw, None, "INVALID"
    if number < 0:
        return raw, None, "INVALID"
    return raw, float(number), "ZERO" if number == 0 else "PRESENT"


def _required_columns(headers: dict[str, int], specification: dict[str, tuple[str, ...]]) -> dict[str, int] | None:
    resolved: dict[str, int] = {}
    for field, aliases in specification.items():
        column = _find(headers, *aliases)
        if column is None:
            return None
        resolved[field] = column
    return resolved


BERTH_COLUMNS = {
    "year": ("Năm",), "voyage": ("Chuyến",), "vessel": ("Tên tàu",),
    "berth": ("Mã bến",), "atb": ("ATB",), "atd": ("ATD",),
}
CARGO_COLUMNS = {
    "size": ("Kích cỡ",), "full_empty": ("F/E",),
    "source_call": ("Tên sà lan | Năm | Chuyến",), "weight": ("Trọng lượng",),
    "trade": ("Hàng nội/ ngoại", "Hàng nội/ngoại"), "method": ("Phương án",),
}


def parse_workbook(content: bytes) -> ParsedWorkbook:
    """Detect by structural signatures and parse one approved workbook kind."""
    _validate_zip(content)
    try:
        workbook = load_workbook(
            io.BytesIO(content), read_only=True, data_only=True, keep_links=False,
        )
    except Exception as exc:
        raise HistoricalWorkbookError("Không thể mở workbook XLSX.") from exc
    try:
        if len(workbook.sheetnames) > MAX_WORKSHEETS:
            raise HistoricalWorkbookError("Workbook có quá nhiều sheet.")
        candidates: list[tuple[str, str, dict[str, int], int]] = []
        for sheet in workbook.worksheets:
            if sheet.max_column > MAX_COLUMNS or sheet.max_row > MAX_ROWS_PER_SHEET:
                raise HistoricalWorkbookError(
                    f"Sheet {sheet.title!r} vượt giới hạn {MAX_ROWS_PER_SHEET} hàng/{MAX_COLUMNS} cột."
                )
            first = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
            headers = _header_map(first)
            berth = _required_columns(headers, BERTH_COLUMNS)
            cargo = _required_columns(headers, CARGO_COLUMNS)
            if berth:
                candidates.append(("berth", sheet.title, berth, 1))
            if cargo:
                candidates.append(("cargo", sheet.title, cargo, 1))
            # PL.03 is a multi-row merged header; identity columns are physical
            # A:C and data starts at row 10 in the audited 35-column variants.
            if sheet.max_column == 35:
                early = [
                    normalize_token(cell.value)
                    for row in sheet.iter_rows(min_row=5, max_row=min(9, sheet.max_row), max_col=35)
                    for cell in row if cell.value is not None
                ]
                if all(item in early for item in ("STT", "TEN PTTND", "SO DANG KY")):
                    candidates.append(("pl03", sheet.title, {}, 9))
        if len(candidates) != 1:
            if not candidates:
                raise HistoricalWorkbookError(
                    "Không nhận diện được loại workbook từ cấu trúc header đã duyệt."
                )
            kinds = ", ".join(f"{kind}:{sheet}" for kind, sheet, _, _ in candidates)
            raise HistoricalWorkbookError(f"Workbook có nhiều cấu trúc nguồn không rõ ưu tiên: {kinds}.")
        kind, sheet_name, columns, header_row = candidates[0]
        sheet = workbook[sheet_name]
        if kind == "berth":
            return _parse_berth(sheet, columns)
        if kind == "cargo":
            return _parse_cargo(sheet, columns)
        return _parse_pl03(sheet)
    finally:
        workbook.close()


def _parse_berth(sheet, columns: dict[str, int]) -> ParsedWorkbook:
    parsed = ParsedWorkbook("tos_berth_call", BERTH_VERSION, sheet_name=sheet.title)
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        value = lambda name: values[columns[name] - 1] if columns[name] <= len(values) else None
        if not any(value(name) not in (None, "") for name in columns):
            continue
        vessel, year, voyage = _raw(value("vessel")), _raw(value("year")), _raw(value("voyage"))
        berth = _raw(value("berth"))
        atb_raw, atb_iso, atb_error = _parse_tos_datetime(value("atb"))
        atd_raw, atd_iso, atd_error = _parse_tos_datetime(value("atd"))
        warnings: list[str] = []
        if not vessel or not re.fullmatch(r"\d{4}", year) or not voyage:
            warnings.append("INVALID_CALL_IDENTITY")
        if atb_error:
            warnings.append(f"ATB_{atb_error}")
        if atd_error:
            warnings.append(f"ATD_{atd_error}")
        status = "REJECTED" if "INVALID_CALL_IDENTITY" in warnings else ("REVIEW" if warnings else "VALID")
        parsed.rows.append({
            "source_sheet": sheet.title, "source_row": row_no,
            "vessel_name_raw": vessel, "vessel_name_normalized": normalize_vessel_name(vessel),
            "call_year_raw": year, "voyage_number_raw": voyage,
            "call_key_normalized": call_key(vessel, year, voyage),
            "source_berth_raw": berth, "arrival_berth": berth, "departure_berth": berth,
            "actual_berthing_at_raw": atb_raw, "actual_berthing_at": atb_iso,
            "actual_departure_at_raw": atd_raw, "actual_departure_at": atd_iso,
            "reporting_month": atb_iso[:7] if atb_iso else None,
            "validation_status": status, "ambiguity_status": "NONE", "warnings": warnings,
            "provenance": {"cells": _cell_receipt(sheet.title, row_no, columns), "raw": {
                name: _safe_value(value(name)) for name in columns
            }},
        })
    duplicates = Counter(row["call_key_normalized"] for row in parsed.rows if row["call_key_normalized"])
    for row in parsed.rows:
        if duplicates[row["call_key_normalized"]] > 1:
            row["ambiguity_status"] = "AMBIGUOUS"
            row["validation_status"] = "REVIEW"
            row["warnings"].append("DUPLICATE_CALL_KEY")
    periods = {row["reporting_month"] for row in parsed.rows if row["reporting_month"]}
    parsed.reporting_period = next(iter(periods)) if len(periods) == 1 else None
    if len(periods) > 1:
        parsed.warnings.append("MULTIPLE_REPORTING_PERIODS")
    parsed.receipt = _receipt(parsed, columns)
    return parsed


def _split_source_call(raw: str) -> tuple[str, str, str] | None:
    parts = [part.strip() for part in raw.split("|")]
    if len(parts) != 3 or not parts[0] or not re.fullmatch(r"\d{4}", parts[1]) or not parts[2]:
        return None
    return parts[0], parts[1], parts[2]


def _parse_cargo(sheet, columns: dict[str, int]) -> ParsedWorkbook:
    parsed = ParsedWorkbook("tos_cargo_detail", CARGO_VERSION, sheet_name=sheet.title)
    direction_map = {"TRA RONG": "unload", "HA BAI": "unload", "LAY NGUYEN": "load", "CAP RONG": "load"}
    trade_map = {"HANG NOI": "domestic", "HANG NGOAI": "foreign"}
    for row_no, values in enumerate(sheet.iter_rows(min_row=2, values_only=True), 2):
        value = lambda name: values[columns[name] - 1] if columns[name] <= len(values) else None
        if not any(value(name) not in (None, "") for name in columns):
            continue
        source_raw = _raw(value("source_call"))
        identity = _split_source_call(source_raw)
        size, full_empty = _raw(value("size")), _raw(value("full_empty")).upper()
        trade_raw, method_raw = _raw(value("trade")), _raw(value("method"))
        size_match = re.match(r"^(20|40)", size)
        teu_factor = 1 if size_match and size_match.group(1) == "20" else (2 if size_match else None)
        trade = trade_map.get(normalize_token(trade_raw), "")
        direction = direction_map.get(normalize_token(method_raw), "")
        weight_raw, weight, weight_state = _decimal(value("weight"))
        warnings: list[str] = []
        if identity is None:
            warnings.append("INVALID_CALL_IDENTITY")
        if teu_factor is None:
            warnings.append("UNSUPPORTED_CONTAINER_SIZE")
        if full_empty not in {"F", "E"}:
            warnings.append("UNKNOWN_FULL_EMPTY")
        if not trade:
            warnings.append("UNKNOWN_TRADE_SCOPE")
        if not direction:
            warnings.append("UNKNOWN_MOVEMENT_METHOD")
        if weight_state in {"BLANK", "INVALID"}:
            warnings.append(f"WEIGHT_{weight_state}")
        rejected = identity is None
        status = "REJECTED" if rejected else ("REVIEW" if warnings else "VALID")
        key = call_key(*identity) if identity else ""
        parsed.rows.append({
            "source_sheet": sheet.title, "source_row": row_no,
            "source_call_key_raw": source_raw, "call_key_normalized": key,
            "container_size_code_raw": size, "teu_factor": teu_factor,
            "full_empty_code_raw": full_empty, "trade_scope_raw": trade_raw,
            "trade_scope": trade, "movement_method_raw": method_raw,
            "derived_direction": direction, "weight_raw": weight_raw,
            "weight_tonnes": weight, "weight_state": weight_state,
            "validation_status": status, "warnings": warnings,
            "provenance": {"cells": _cell_receipt(sheet.title, row_no, columns), "raw": {
                name: _safe_value(value(name)) for name in columns
            }},
        })
    parsed.receipt = _receipt(parsed, columns)
    return parsed


PL03_METRICS = {
    9: "export_tons_reported", 10: "export_full_teu_reported", 11: "export_empty_teu_reported",
    12: "import_tons_reported", 13: "import_full_teu_reported", 14: "import_empty_teu_reported",
    15: "domestic_inbound_tons_reported", 16: "domestic_inbound_full_teu_reported",
    17: "domestic_inbound_empty_teu_reported", 18: "domestic_outbound_tons_reported",
    19: "domestic_outbound_full_teu_reported", 20: "domestic_outbound_empty_teu_reported",
    21: "transshipment_tons_reported", 22: "transshipment_teu_reported",
    23: "transit_handled_tons_reported", 24: "transit_handled_teu_reported",
    25: "transit_not_handled_tons_reported", 26: "transit_not_handled_teu_reported",
}


def _parse_pl03(sheet) -> ParsedWorkbook:
    parsed = ParsedWorkbook("reported_pl03", PL03_VERSION, appendix_kind="PL.03", sheet_name=sheet.title)
    for row_no, values in enumerate(sheet.iter_rows(min_row=10, max_col=35, values_only=True), 10):
        ordinal = values[0] if values else None
        try:
            ordinal_int = int(ordinal)
        except (TypeError, ValueError):
            continue
        vessel = _raw(values[1] if len(values) > 1 else None)
        # Historical templates may carry an ordinal in a reserved placeholder
        # row.  Audit contract: blank PL.03/B means no report row.
        if not vessel:
            continue
        warnings: list[str] = []
        metrics = []
        for column, code in PL03_METRICS.items():
            raw, number, state = _decimal(values[column - 1] if len(values) >= column else None)
            metrics.append({
                "metric_code": code, "numeric_value": number, "text_value": raw or None,
                "value_state": state if state in {"BLANK", "ZERO"} else "PRESENT",
                "invalid": state == "INVALID", "source_cell": f"{sheet.title}!{get_column_letter(column)}{row_no}",
            })
            if state == "INVALID":
                warnings.append(f"INVALID_METRIC_{get_column_letter(column)}")
        parsed.rows.append({
            "source_sheet": sheet.title, "source_row": row_no, "appendix_row_no": ordinal_int,
            "vessel_name_raw": vessel, "registration_raw": _raw(values[2] if len(values) > 2 else None),
            "raw_payload": {get_column_letter(index + 1): _safe_value(value) for index, value in enumerate(values)},
            "metrics": metrics, "validation_status": "REVIEW" if warnings else "VALID", "warnings": warnings,
            "provenance": {"row": f"{sheet.title}!A{row_no}:AI{row_no}", "legacy_time_cells": {
                "AG": f"{sheet.title}!AG{row_no}", "AH": f"{sheet.title}!AH{row_no}"
            }, "time_authority": "LEGACY_REPORTED_ONLY_TOS_ATB_ATD_AUTHORITATIVE"},
        })
    parsed.receipt = _receipt(parsed, {"ordinal": 1, "vessel": 2, "registration": 3})
    return parsed


def _receipt(parsed: ParsedWorkbook, columns: dict[str, int]) -> dict[str, Any]:
    return {
        "detectionStrategy": "SHEET_HEADER_STRUCTURE_SIGNATURE",
        "sourceKind": parsed.source_kind,
        "mappingVersion": parsed.mapping_version,
        "sheet": parsed.sheet_name,
        "columns": {name: get_column_letter(index) for name, index in columns.items()},
        "counts": parsed.counts,
        "warnings": parsed.warnings,
        "filenameUsedForDetection": False,
        "hiddenRowsAndColumnsIncluded": True,
        "transformVersion": TRANSFORM_VERSION,
    }


def json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=_safe_value)
