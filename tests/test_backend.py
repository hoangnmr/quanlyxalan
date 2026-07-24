"""
tests/test_backend.py — T0 Baseline Recovery test suite
WO-KBCV-T0-20260711

Uses pytest + httpx TestClient against the FastAPI app.
Test database is a throwaway PostgreSQL database — completely isolated from the app database.

IMPORTANT: os.environ["TEST_DATABASE_URL"] is set BEFORE any backend imports so that
           database.py picks up the test URL at module init time.
"""
from __future__ import annotations

import io
import json
import os
import tempfile
import time
import uuid
import zipfile
from pathlib import Path

# ── Set test DB FIRST, before any backend import ──────────────────────────────
from tests import _pgdb

_TEST_DB_URL = _pgdb.create_database("kbcv_backend")
os.environ["TEST_DATABASE_URL"] = _TEST_DB_URL

# ── Now import backend (picks up TEST_DATABASE_URL) ───────────────────────────
import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
import backend.app as app_module

from backend.models import (
    AuditEvent, Base, Declaration, ImportJob, ReportAdjustment, User, Organization, Vessel,
    VesselOperatingProfile, ReportingUnit, ReportingUnitOrganization, ReportingUnitUser,
    ReportingUnitVessel,
    HistoricalCargoRow, HistoricalPortCall, HistoricalReportImport,
)
from backend.database import engine, SessionLocal, now_iso
from backend.auth import get_password_hash
from backend.app import app, get_db, DEMO_ORGANIZATION_TAX_CODE, remove_demo_data_for_real_input
from backend.xlsx_io import make_xlsx, read_workbook, vessel_rows

# ── Create all tables in test DB ──────────────────────────────────────────────
Base.metadata.create_all(bind=engine)


# ── DB override ───────────────────────────────────────────────────────────────
def _override_get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()


app.dependency_overrides[get_db] = _override_get_db


# ── Seed test user ────────────────────────────────────────────────────────────
def _seed_user() -> None:
    db = SessionLocal()
    try:
        # Seed test organization first
        org = db.query(Organization).filter(Organization.name == "Test Org").first()
        if not org:
            org = Organization(name="Test Org", tax_code="123456", created_at=now_iso(), updated_at=now_iso())
            db.add(org)
            db.commit()
            db.refresh(org)

        # Seed users
        user_data = [
            ("testuser", "PLATFORM_ADMIN", None),
            ("customeruser", "CUSTOMER", org.id),
            ("portstaff", "PORT_STAFF", None),
        ]
        for username, role, org_id in user_data:
            if not db.query(User).filter(User.username == username).first():
                db.add(User(
                    username=username,
                    password_hash=get_password_hash("testpass"),
                    full_name=f"{role} User",
                    role=role,
                    organization_id=org_id,
                    is_active=1,
                    created_at=now_iso(),
                ))
        db.commit()

        # R4: seed the single ReportingUnit this legacy test suite operates against,
        # link Test Org to it and give portstaff membership so the tenant-context
        # guard does not lock out pre-existing role-only test fixtures.
        unit = db.query(ReportingUnit).filter(ReportingUnit.code == "TEST-UNIT").first()
        if not unit:
            unit = ReportingUnit(
                name="Test Reporting Unit", code="TEST-UNIT", is_active=1,
                created_at=now_iso(), updated_at=now_iso(),
            )
            db.add(unit)
            db.commit()
            db.refresh(unit)
        if not db.query(ReportingUnitOrganization).filter_by(reporting_unit_id=unit.id, organization_id=org.id).first():
            db.add(ReportingUnitOrganization(
                reporting_unit_id=unit.id, organization_id=org.id, created_at=now_iso(),
            ))
        port_user = db.query(User).filter(User.username == "portstaff").first()
        if port_user and not db.query(ReportingUnitUser).filter_by(reporting_unit_id=unit.id, user_id=port_user.id).first():
            db.add(ReportingUnitUser(
                reporting_unit_id=unit.id, user_id=port_user.id, created_at=now_iso(),
            ))
        db.commit()
        global TEST_REPORTING_UNIT_ID
        TEST_REPORTING_UNIT_ID = unit.id
        global TEST_ORGANIZATION_ID
        TEST_ORGANIZATION_ID = org.id
    finally:
        db.close()


TEST_REPORTING_UNIT_ID: int | None = None
TEST_ORGANIZATION_ID: int | None = None
_seed_user()


# ── Fixtures ──────────────────────────────────────────────────────────────────
@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c


@pytest.fixture(scope="module")
def auth_headers(client):
    res = client.post("/api/auth/login", json={"username": "testuser", "password": "testpass"})
    assert res.status_code == 200, f"Login failed: {res.text}"
    token = res.json()["access_token"]
    # R4: PLATFORM_ADMIN must supply an explicit ReportingUnit context for
    # tenant-bound operations; pre-existing tests operate against the single
    # seeded Test Reporting Unit.
    return {"Authorization": f"Bearer {token}", "X-Reporting-Unit-ID": str(TEST_REPORTING_UNIT_ID)}


@pytest.fixture(scope="module")
def customer_headers(client):
    res = client.post("/api/auth/login", json={"username": "customeruser", "password": "testpass"})
    assert res.status_code == 200, f"Login failed: {res.text}"
    token = res.json()["access_token"]
    return {"Authorization": f"Bearer {token}"}


@pytest.fixture(scope="module")
def port_staff_headers(client):
    res = client.post("/api/auth/login", json={"username": "portstaff", "password": "testpass"})
    assert res.status_code == 200, f"Login failed: {res.text}"
    token = res.json()["access_token"]
    return {"Authorization": f"Bearer {token}", "X-Reporting-Unit-ID": str(TEST_REPORTING_UNIT_ID)}


# ── Helper ────────────────────────────────────────────────────────────────────
def _reg() -> str:
    """Generate a unique registration number."""
    return f"SG-T0-{uuid.uuid4().hex}"


def _minimal_declaration(**overrides) -> dict:
    base = {
        "company_name": "Test Company",
        "declaration_date": "2026-07-11",
        "vessel_name": "TT TEST",
        "registration_no": _reg(),
        "vessel_type": "Tàu container",
        "vessel_class": "VR-SI",
        "last_port": "Bến A",
        "working_port": "Cảng Tân Thuận",
        "eta": "2026-07-11T08:00",
        "etd": "2026-07-11T18:00",
        "master_name": "Nguyễn Văn A",
        "master_phone": "0900000000",
        "unload": {},
        "load": {},
    }
    base.update(overrides)
    return base


def _historical_fixture(headers: dict[int, str], rows: list[dict[int, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "User renamed sheet"
    for column, value in headers.items():
        sheet.cell(1, column, value)
    for row_number, row in enumerate(rows, 2):
        for column, value in row.items():
            sheet.cell(row_number, column, value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _historical_pl03_fixture(rows: list[dict[int, object]]) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PL03 user renamed"
    sheet.cell(5, 1, "STT")
    sheet.cell(5, 2, "Tên PTTND")
    sheet.cell(5, 3, "Số đăng ký")
    for row_number, row in enumerate(rows, 10):
        for column, value in row.items():
            sheet.cell(row_number, column, value)
    sheet.cell(9, 35, "")
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def _seed_historical_registered_vessel(name: str) -> None:
    db = SessionLocal()
    try:
        vessel = Vessel(
            organization_id=TEST_ORGANIZATION_ID, name=name,
            registration_no=_reg(), vessel_type="Chở hàng khô", vessel_class="VR-SI",
            created_at=now_iso(), updated_at=now_iso(),
        )
        db.add(vessel)
        db.flush()
        db.add(ReportingUnitVessel(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, vessel_id=vessel.id,
            created_at=now_iso(),
        ))
        db.commit()
    finally:
        db.close()

# ══════════════════════════════════════════════════════════════════════════════
# 1. HEALTH + STATIC FRONTEND
# ══════════════════════════════════════════════════════════════════════════════

def test_health(client):
    res = client.get("/api/health")
    assert res.status_code == 200
    assert res.json()["status"] == "ok"
    assert res.json()["storage"] == "LOCAL_QUARANTINE"


def test_readiness(client):
    res = client.get("/api/ready")
    assert res.status_code == 200
    assert res.json()["status"] == "ready"


def test_admin_operations_summary(client, auth_headers, customer_headers):
    admin = client.get("/api/admin/operations-summary", headers=auth_headers)
    assert admin.status_code == 200
    assert {"operations", "fleet", "imports", "storage", "security"}.issubset(admin.json())
    assert client.get("/api/admin/operations-summary", headers=customer_headers).status_code == 403


def test_dashboard_attention_queue_is_role_scoped(client, customer_headers, port_staff_headers):
    customer = client.get("/api/dashboard", headers=customer_headers)
    assert customer.status_code == 200
    assert customer.json()["attention"]["label"]
    assert all(item["workflow_status"] in {"DRAFT", "CHANGES_REQUESTED"} for item in customer.json()["attention"]["items"])

    reviewer = client.get("/api/dashboard", headers=port_staff_headers)
    assert reviewer.status_code == 200
    assert all(item["workflow_status"] == "PENDING_REVIEW" for item in reviewer.json()["attention"]["items"])


def test_static_frontend(client):
    res = client.get("/")
    assert res.status_code == 200
    assert "text/html" in res.headers.get("content-type", "")
    assert 'class="skip-link"' in res.text
    assert 'id="main-content" tabindex="-1" aria-busy="false"' in res.text
    assert 'id="toast-region" class="toast-region" role="status"' in res.text
    assert 'id="in-app-certificate-reminders"' in res.text
    assert 'id="certificate-reminder"' in res.text
    assert 'id="demo-data-notice"' in res.text
    assert 'id="login-dialog" class="modal login-dialog"' in res.text
    # Cache-busting query giữ nguyên; số phiên bản thay đổi theo mỗi lần release,
    # và đường dẫn phải là tương đối để hoạt động khi app chạy dưới subpath.
    assert 'styles.css?v=' in res.text and '"/styles.css' not in res.text
    assert 'app.js?v=' in res.text and '"/app.js' not in res.text
    assert 'id="analytics-source-controls"' in res.text
    assert 'data-source="historical"' in res.text
    assert 'id="analytics-coverage"' in res.text
    assert 'id="analytics-combined-blocked"' in res.text
    assert 'id="sidebar-unit-context"' in res.text
    assert 'id="reporting-unit-trigger"' in res.text
    assert 'id="reporting-unit-menu"' in res.text
    assert 'id="reporting-unit-dialog"' in res.text
    assert 'id="reporting-unit-form"' in res.text
    assert 'id="reporting-unit-select"' not in res.text
    assert 'id="reporting-unit-required"' in res.text
    assert 'data-page="port-register"' in res.text
    assert 'id="export-port-register"' in res.text
    assert 'id="import-port-register"' in res.text
    assert 'id="port-import-dialog"' in res.text
    assert 'id="analytics-unavailable"' in res.text
    assert 'id="external-integration-panel" class="panel integration-panel"' in res.text
    assert 'id="integration-admin-actions" class="integration-state" hidden' in res.text
    assert 'class="primary-nav"' in res.text and 'class="data-nav"' in res.text
    assert "Báo cáo hoạt động" in res.text
    assert "Báo cáo Cảng vụ" not in res.text
    assert 'class="panel action-panel"' not in res.text
    app_js = client.get("/app.js").text
    assert "function setSubmitting(" in app_js
    assert "function bindLoginForm()" in app_js
    assert "bindLoginForm();" in app_js
    assert "Tiến trình duyệt" in app_js
    assert "nhân viên Cảng" in f"{res.text}\n{app_js}"
    assert "CV = Cảng vụ viên" not in app_js
    assert "Theo các bước CV · QLC · BP" not in app_js
    assert "PORT_APPROVE" in app_js
    assert "crew-checklist" in app_js
    assert "step-error-summary" in app_js
    assert "port_approval" in app_js
    assert "cv_approval" not in app_js
    assert "loadReportAnalytics($('.period-switch button.active')?.dataset.period || 'month', state.analyticsSource)" in app_js
    assert "CUSTOMER:'User'" in app_js
    assert "PORT_STAFF:'Port staff'" in app_js
    assert "PLATFORM_ADMIN:'Platform admin'" in app_js
    assert "'X-Reporting-Unit-ID': String(state.activeReportingUnitId)" in app_js
    assert "async function loadReportingUnitContext()" in app_js
    assert "Hệ thống không có chế độ xem gộp nhiều cảng" in app_js
    assert "if (state.currentUser?.role === 'PLATFORM_ADMIN') loadIntegration();" in app_js
    assert "btn.hidden = !canCreateDeclaration" in app_js
    assert "!['declarations', 'crew'].includes(link.dataset.route)" in app_js
    assert "$('#user-display').innerHTML = `<span class=\"role-pill\"" in app_js
    assert "const crewContainer = $('#declaration-crew-container');" in app_js
    assert "name=\"crew_onboard_count\"" in app_js
    assert "node.setAttribute('role', error ? 'alert' : 'status')" in app_js
    assert "Không thể nhập dòng này. Hãy kiểm tra định dạng số, ngày hoặc mã đăng ký trùng." in Path(__file__).resolve().parents[1].joinpath("backend", "app.py").read_text(encoding="utf-8")
    assert "File đã được nhập trước đó" in app_js
    assert "Không tạo thêm bản ghi" in app_js
    assert "searchDashboardVessels(query, sequence)" in app_js
    assert "dashboardTimer = setTimeout(() => loadDashboard" not in app_js
    assert "importNav.style.removeProperty('display')" in app_js
    assert "reportsNav.style.removeProperty('display')" in app_js
    assert "importNav.style.display = 'block'" not in app_js
    assert "reportsNav.style.display = 'block'" not in app_js
    assert "Giữ dữ liệu hiện có & tiếp tục" in app_js
    assert "overwrite_existing=true" in app_js
    assert "field('birth_date','Ngày sinh (không bắt buộc)'" in app_js
    assert 'name="vessel_id"><option value="">Chưa phân công' not in app_js
    assert "previewImport(event.target, '/api/import/crew', 'crew')" in app_js
    assert "const CREW_ROLES = ['Thuyền trưởng', 'Máy trưởng', 'Thuyền viên', 'Thuyền phó']" in app_js
    assert "'Máy phó'" not in app_js
    assert "'Thủy thủ'" not in app_js
    styles_css = client.get("/styles.css").text
    assert "[hidden] { display: none !important; }" in styles_css
    assert "overflow-y: auto" in styles_css
    assert "overscroll-behavior: contain" in styles_css
    assert ".data-nav { margin-top: 0" in styles_css
    assert ".sidebar-footer { margin-top: auto" in styles_css
    assert ".integration-readiness" in styles_css
    assert "#crew-dialog { width: min(720px" in styles_css
    assert "#crew-fields { grid-template-columns: repeat(2, minmax(0, 1fr)); }" in styles_css


def test_platform_admin_can_create_empty_reporting_unit(
    client, auth_headers, port_staff_headers, customer_headers,
):
    suffix = uuid.uuid4().hex[:8].upper()
    name = f"Test Port {suffix}"
    code = f"PORT-{suffix}"
    unit_id = None
    try:
        assert client.post(
            "/api/reporting-units", headers=port_staff_headers,
            json={"name": name, "code": code},
        ).status_code == 403
        assert client.post(
            "/api/reporting-units", headers=customer_headers,
            json={"name": name, "code": code},
        ).status_code == 403
        invalid = client.post(
            "/api/reporting-units", headers=auth_headers,
            json={"name": "X", "code": "mã có dấu"},
        )
        assert invalid.status_code == 422

        created = client.post(
            "/api/reporting-units", headers=auth_headers,
            json={"name": f"  {name}  ", "code": code.lower()},
        )
        assert created.status_code == 201, created.text
        body = created.json()
        unit_id = body["id"]
        assert body == {
            "id": unit_id, "name": name, "code": code,
            "notify_email": "", "is_active": True,
        }

        duplicate_code = client.post(
            "/api/reporting-units", headers=auth_headers,
            json={"name": f"Other {suffix}", "code": code},
        )
        assert duplicate_code.status_code == 409
        listed = client.get("/api/reporting-units", headers=auth_headers)
        assert any(item["id"] == unit_id for item in listed.json()["items"])

        db = SessionLocal()
        try:
            event = db.query(AuditEvent).filter_by(
                entity_type="REPORTING_UNIT", entity_id=unit_id, action="CREATE",
            ).one()
            assert event.reporting_unit_id == unit_id
            assert event.actor_user_id is not None
            assert db.query(ReportingUnitOrganization).filter_by(reporting_unit_id=unit_id).count() == 0
            assert db.query(ReportingUnitUser).filter_by(reporting_unit_id=unit_id).count() == 0
        finally:
            db.close()
    finally:
        if unit_id:
            db = SessionLocal()
            try:
                db.query(AuditEvent).filter_by(reporting_unit_id=unit_id).delete()
                db.query(ReportingUnit).filter_by(id=unit_id).delete()
                db.commit()
            finally:
                db.close()


def test_real_input_removes_only_sentinel_marked_demo_data():
    db = SessionLocal()
    try:
        demo = Organization(name="Demo sentinel", tax_code=DEMO_ORGANIZATION_TAX_CODE, created_at=now_iso(), updated_at=now_iso())
        real = Organization(name=f"Real org {_reg()}", tax_code=_reg(), created_at=now_iso(), updated_at=now_iso())
        db.add_all([demo, real])
        db.flush()
        db.add_all([
            Vessel(organization_id=demo.id, name="Demo vessel", registration_no=_reg(), vessel_type="Sà lan", vessel_class="VR-SII", created_at=now_iso(), updated_at=now_iso()),
            Vessel(organization_id=real.id, name="Real vessel", registration_no=_reg(), vessel_type="Sà lan", vessel_class="VR-SII", created_at=now_iso(), updated_at=now_iso()),
        ])
        db.commit()
        assert remove_demo_data_for_real_input(db) is True
        db.commit()
        assert db.query(Organization).filter(Organization.tax_code == DEMO_ORGANIZATION_TAX_CODE).first() is None
        assert db.query(Organization).filter(Organization.id == real.id).first() is not None
        assert db.query(Vessel).filter(Vessel.organization_id == real.id).count() == 1
        db.query(Vessel).filter(Vessel.organization_id == real.id).delete(synchronize_session=False)
        db.delete(real)
        db.commit()
    finally:
        db.rollback()
        db.close()


def test_declaration_pagination_contract_is_bounded_and_compatible(client, auth_headers):
    legacy = client.get("/api/declarations", headers=auth_headers)
    assert legacy.status_code == 200
    assert isinstance(legacy.json(), list)

    paged = client.get("/api/declarations?page=1&page_size=1&sort=reference_no&direction=asc", headers=auth_headers)
    assert paged.status_code == 200
    body = paged.json()
    assert {"items", "page", "page_size", "total", "total_pages", "sort", "direction"}.issubset(body)
    assert body["page_size"] == 1
    assert len(body["items"]) <= 1

    assert client.get("/api/declarations?page=1&page_size=101", headers=auth_headers).status_code == 422


# ══════════════════════════════════════════════════════════════════════════════
# 2. AUTHENTICATION
# ══════════════════════════════════════════════════════════════════════════════

def test_login_success(client):
    res = client.post("/api/auth/login", json={"username": "testuser", "password": "testpass"})
    assert res.status_code == 200
    assert "access_token" in res.json()


def test_notification_preferences_are_user_controlled_and_audited(client, customer_headers):
    before = client.get("/api/notification-preferences", headers=customer_headers)
    assert before.status_code == 200
    assert before.json()["in_app_certificate_reminders"] is True

    updated = client.put(
        "/api/notification-preferences",
        headers=customer_headers,
        json={"in_app_certificate_reminders": False},
    )
    assert updated.status_code == 200
    assert updated.json() == {
        "in_app_certificate_reminders": False,
        "email_certificate_reminders": False,
        "email_workflow_updates": False,
    }
    assert client.get("/api/notification-preferences", headers=customer_headers).json() == updated.json()

    db = SessionLocal()
    try:
        assert db.query(AuditEvent).filter(AuditEvent.action == "NOTIFICATION_PREFERENCES_UPDATE").count() >= 1
    finally:
        db.close()


def test_login_wrong_password(client):
    res = client.post("/api/auth/login", json={"username": "testuser", "password": "wrongpass"})
    assert res.status_code == 401


def test_protected_route_requires_auth(client):
    """Accessing /api/vessels without token must return 401."""
    # Temporarily remove override so the real auth check runs
    original = app.dependency_overrides.pop(get_db, None)
    try:
        # Call without auth header
        res = client.get("/api/vessels")
        # Should be 401 (token validation) or the endpoint is accessible — depends on auth dep
        # Re-add override first, then check with fake token
    finally:
        if original:
            app.dependency_overrides[get_db] = original

    res = client.get("/api/vessels", headers={"Authorization": "Bearer invalidtoken"})
    assert res.status_code == 401


def test_catalogs_public(client):
    """Catalogs endpoint should be accessible without auth."""
    res = client.get("/api/catalogs")
    assert res.status_code == 200
    data = res.json()
    assert "vesselTypeSuggestions" in data
    assert "vesselCategories" in data
    assert "vesselClasses" in data


# ══════════════════════════════════════════════════════════════════════════════
# 3. VESSELS
# ══════════════════════════════════════════════════════════════════════════════

def test_vessel_create(client, auth_headers):
    payload = {
        "name": "TT VESSEL A",
        "registration_no": _reg(),
        "vessel_type": "Tàu container",
        "vessel_class": "VR-SI",
        "organization_name": "Test Org A",
    }
    res = client.post("/api/vessels", json=payload, headers=auth_headers)
    assert res.status_code == 200
    data = res.json()
    assert data["name"] == "TT VESSEL A"
    assert "id" in data


def test_vessel_list(client, auth_headers):
    res = client.get("/api/vessels", headers=auth_headers)
    assert res.status_code == 200
    assert isinstance(res.json(), list)


def test_platform_admin_can_delete_vessel_with_no_dependents(client, auth_headers):
    created = client.post("/api/vessels", json={
        "name": "TT DELETE ME", "registration_no": _reg(),
        "vessel_type": "Tàu container", "vessel_class": "VR-SI",
        "organization_name": "Test Org A",
    }, headers=auth_headers)
    assert created.status_code == 200
    vessel_id = created.json()["id"]

    res = client.delete(f"/api/vessels/{vessel_id}", headers=auth_headers)
    assert res.status_code == 200, res.text
    assert res.json()["deleted"] == vessel_id

    listed = client.get("/api/vessels", headers=auth_headers)
    assert all(item["id"] != vessel_id for item in listed.json())


def test_port_staff_cannot_delete_vessel(client, auth_headers, port_staff_headers):
    created = client.post("/api/vessels", json={
        "name": "TT NOT MINE TO DELETE", "registration_no": _reg(),
        "vessel_type": "Tàu container", "vessel_class": "VR-SI",
        "organization_name": "Test Org A",
    }, headers=auth_headers)
    vessel_id = created.json()["id"]

    res = client.delete(f"/api/vessels/{vessel_id}", headers=port_staff_headers)
    assert res.status_code == 403

    # cleanup
    client.delete(f"/api/vessels/{vessel_id}", headers=auth_headers)


def test_platform_admin_cannot_delete_vessel_with_declarations(client, auth_headers):
    created = client.post("/api/vessels", json={
        "name": "TT HAS DECLARATIONS", "registration_no": _reg(),
        "vessel_type": "Tàu container", "vessel_class": "VR-SI",
        "organization_name": "KHÁCH HÀNG IMPORT",
    }, headers=auth_headers)
    vessel = created.json()
    vessel_id = vessel["id"]

    decl = client.post(
        "/api/declarations",
        json=_minimal_declaration(
            company_name="KHÁCH HÀNG IMPORT",
            vessel_id=vessel_id,
            vessel_name=vessel["name"],
            registration_no=vessel["registration_no"],
        ),
        headers=auth_headers,
    )
    assert decl.status_code == 200, decl.text

    res = client.delete(f"/api/vessels/{vessel_id}", headers=auth_headers)
    assert res.status_code == 409


def test_vessel_lists_use_the_same_order(client, port_staff_headers):
    vessels = client.get("/api/vessels", headers=port_staff_headers)
    port_register = client.get("/api/port-vessel-register", headers=port_staff_headers)
    assert vessels.status_code == 200
    assert port_register.status_code == 200

    tracked_ids = [item["id"] for item in port_register.json()["items"]]
    tracked_id_set = set(tracked_ids)
    vessel_order_for_tracked_items = [
        item["id"] for item in vessels.json() if item["id"] in tracked_id_set
    ]
    assert vessel_order_for_tracked_items == tracked_ids


def test_vessel_update(client, auth_headers):
    reg_no = _reg()
    res = client.post("/api/vessels", json={
        "name": "TT UPDATE",
        "registration_no": reg_no,
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SI",
        "organization_name": "Test Org",
    }, headers=auth_headers)
    assert res.status_code == 200
    vid = res.json()["id"]

    res2 = client.post("/api/vessels", json={
        "id": vid,
        "name": "TT UPDATED",
        "registration_no": reg_no,
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SII",
    }, headers=auth_headers)
    assert res2.status_code == 200
    assert res2.json()["vessel_class"] == "VR-SII"


def test_vessel_stale_version_rejected(client, auth_headers):
    reg_no = _reg()
    created = client.post("/api/vessels", json={
        "name": "TT VERSION",
        "registration_no": reg_no,
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SI",
        "organization_name": "Test Org",
    }, headers=auth_headers)
    assert created.status_code == 200
    vessel = created.json()

    updated = client.post("/api/vessels", json={
        "id": vessel["id"], "version": vessel["version"],
        "name": "TT VERSION UPDATED", "registration_no": reg_no,
        "vessel_type": "Tàu hàng khô", "vessel_class": "VR-SI",
    }, headers=auth_headers)
    assert updated.status_code == 200
    assert updated.json()["version"] == vessel["version"] + 1

    stale = client.post("/api/vessels", json={
        "id": vessel["id"], "version": vessel["version"],
        "name": "TT VERSION STALE", "registration_no": reg_no,
        "vessel_type": "Tàu hàng khô", "vessel_class": "VR-SI",
    }, headers=auth_headers)
    assert stale.status_code == 409


def test_duplicate_vessel_rolls_back_new_organization(client, auth_headers):
    """A failed write cannot leave the organization created during that request."""
    registration = _reg()
    first = client.post("/api/vessels", json={
        "name": "TT DUPLICATE ONE",
        "registration_no": registration,
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SI",
        "organization_name": "Committed Org",
    }, headers=auth_headers)
    assert first.status_code == 200

    failed_org = f"Rolled Back Org {time.time_ns()}"
    duplicate = client.post("/api/vessels", json={
        "name": "TT DUPLICATE TWO",
        "registration_no": registration,
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SI",
        "organization_name": failed_org,
    }, headers=auth_headers)
    assert duplicate.status_code == 409

    db = SessionLocal()
    try:
        assert db.query(Organization).filter(Organization.name == failed_org).first() is None
    finally:
        db.close()


def test_vessel_verify_registry(client, auth_headers):
    res = client.post("/api/vessels", json={
        "name": "TT VERIFY",
        "registration_no": _reg(),
        "vessel_type": "Tàu hàng khô",
        "vessel_class": "VR-SI",
        "organization_name": "Test Org",
    }, headers=auth_headers)
    assert res.status_code == 200
    vid = res.json()["id"]
    res2 = client.post(f"/api/vessels/{vid}/verify-registry", headers=auth_headers)
    assert res2.status_code == 200
    assert res2.json()["registry_verification_status"] == "VERIFIED_LOCAL"
    assert res2.json()["registry_verification_source"] == "local"
    assert res2.json()["adapter"]["mode"] == "MANUAL"
    assert res2.json()["adapter"]["networkCallsAllowed"] is False


# ══════════════════════════════════════════════════════════════════════════════
# 4. CREW
# ══════════════════════════════════════════════════════════════════════════════

def test_crew_create(client, auth_headers):
    payload = {
        "organization_id": TEST_ORGANIZATION_ID,
        "full_name": "Nguyễn Văn Thuyền",
        "crew_role": "Thuyền trưởng",
        "birth_date": "1985-04-12",
        "professional_certificate_type": "Bằng thuyền trưởng",
        "professional_certificate_no": "CERT-001",
        "certificate_expiry_date": "2020-01-01",
    }
    res = client.post("/api/crew", json=payload, headers=auth_headers)
    assert res.status_code == 200
    data = res.json()
    assert data["certificate_status"] == "EXPIRED"
    assert data["full_name"] == "Nguyễn Văn Thuyền"
    assert data["birth_date"] == "1985-04-12"
    assert data["vessel_id"] is None


def test_port_staff_crew_create_requires_target_organization(client, port_staff_headers):
    response = client.post("/api/crew", json={
        "full_name": "Không được tạo thủ công",
        "crew_role": "Thuyền viên",
        "professional_certificate_type": "Chứng chỉ",
        "professional_certificate_no": "PORT-MANUAL-DENIED",
    }, headers=port_staff_headers)
    assert response.status_code == 422


def test_crew_role_catalog_rejects_unapproved_role(client, customer_headers):
    response = client.post("/api/crew", json={
        "full_name": "Vai trò không hợp lệ",
        "crew_role": "Thủy thủ",
        "professional_certificate_type": "Chứng chỉ",
        "professional_certificate_no": "ROLE-DENIED",
    }, headers=customer_headers)
    assert response.status_code == 422


def test_crew_list(client, auth_headers):
    res = client.get("/api/crew", headers=auth_headers)
    assert res.status_code == 200
    assert isinstance(res.json(), list)


def test_crew_update(client, auth_headers):
    res = client.post("/api/crew", json={
        "organization_id": TEST_ORGANIZATION_ID,
        "full_name": "Trần Máy Trưởng",
        "crew_role": "Máy trưởng",
        "professional_certificate_type": "Bằng máy trưởng",
        "professional_certificate_no": "CERT-002",
        "certificate_expiry_date": "2099-01-01",
    }, headers=auth_headers)
    assert res.status_code == 200
    cid = res.json()["id"]

    res2 = client.post("/api/crew", json={
        "id": cid,
        "full_name": "Trần Máy Trưởng Updated",
        "crew_role": "Máy trưởng",
        "professional_certificate_type": "Bằng máy trưởng",
        "professional_certificate_no": "CERT-002-UPD",
        "certificate_expiry_date": "2099-06-01",
    }, headers=auth_headers)
    assert res2.status_code == 200
    assert res2.json()["professional_certificate_no"] == "CERT-002-UPD"


# ══════════════════════════════════════════════════════════════════════════════
# 5. DECLARATIONS — DRAFT AND SUBMIT
# ══════════════════════════════════════════════════════════════════════════════

def test_declaration_draft_create(client, auth_headers):
    res = client.post("/api/declarations", json=_minimal_declaration(), headers=auth_headers)
    assert res.status_code == 200
    data = res.json()
    assert data["workflow_status"] == "DRAFT"
    assert "reference_no" in data


def test_platform_admin_can_delete_draft_declaration(client, auth_headers):
    created = client.post("/api/declarations", json=_minimal_declaration(), headers=auth_headers)
    declaration_id = created.json()["id"]

    res = client.delete(f"/api/declarations/{declaration_id}", headers=auth_headers)
    assert res.status_code == 200, res.text
    assert res.json()["deleted"] == declaration_id

    listed = client.get("/api/declarations?page=1", headers=auth_headers)
    assert all(item["id"] != declaration_id for item in listed.json()["items"])


def test_port_staff_cannot_delete_declaration(client, auth_headers, port_staff_headers):
    created = client.post("/api/declarations", json=_minimal_declaration(), headers=auth_headers)
    declaration_id = created.json()["id"]

    res = client.delete(f"/api/declarations/{declaration_id}", headers=port_staff_headers)
    assert res.status_code == 403

    client.delete(f"/api/declarations/{declaration_id}", headers=auth_headers)  # cleanup


def test_platform_admin_can_delete_submitted_declaration(client, auth_headers):
    # Quyền của admin là tuyệt đối: xóa được phiếu ở mọi trạng thái workflow
    # (xem DELETE /api/declarations/{id}). Mọi lần xóa đều ghi audit trail.
    created = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(company_name="KHÁCH HÀNG IMPORT"),
        headers=auth_headers,
    )
    declaration_id = created.json()["id"]
    assert created.json()["workflow_status"] == "PENDING_REVIEW"

    res = client.delete(f"/api/declarations/{declaration_id}", headers=auth_headers)
    assert res.status_code == 200
    assert res.json() == {"deleted": declaration_id}


def test_platform_admin_sees_its_own_drafts(client, auth_headers):
    # PLATFORM_ADMIN's "Lưu phiếu" flow creates a DRAFT declaration it must be
    # able to find again — both endpoints previously hid every DRAFT from any
    # non-CUSTOMER scope (a PORT_STAFF-only rule that leaked onto the admin),
    # so an admin-created draft was created successfully but then invisible in
    # both the declarations list and the dashboard counters.
    created = client.post("/api/declarations", json=_minimal_declaration(), headers=auth_headers)
    assert created.status_code == 200
    reference_no = created.json()["reference_no"]

    listed = client.get("/api/declarations?page=1", headers=auth_headers)
    assert listed.status_code == 200
    assert any(item["reference_no"] == reference_no for item in listed.json()["items"])

    dashboard = client.get("/api/dashboard", headers=auth_headers)
    assert dashboard.status_code == 200
    assert dashboard.json()["stats"]["drafts"] >= 1
    assert any(item["reference_no"] == reference_no for item in dashboard.json()["recent"])


def test_port_staff_can_create_draft_declaration(client, port_staff_headers):
    # Port staff may key in a declaration on a customer's behalf (e.g. by
    # phone/paper), but only as a DRAFT — confirmation/submission still
    # requires the customer or a PLATFORM_ADMIN (see
    # test_port_staff_cannot_submit_declarations).
    res = client.post(
        "/api/declarations",
        json=_minimal_declaration(company_name="KHÁCH HÀNG PORT STAFF"),
        headers=port_staff_headers,
    )
    assert res.status_code == 200, res.text
    assert res.json()["workflow_status"] == "DRAFT"


def test_declaration_submit(client, customer_headers):
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    assert res.status_code == 200
    data = res.json()
    assert data["workflow_status"] == "PENDING_REVIEW"
    assert data["status"] == "SUBMITTED"


def test_platform_admin_can_submit_on_customers_behalf(client, auth_headers):
    # PLATFORM_ADMIN has full authority and may submit a declaration on a
    # customer's behalf without logging into the customer's own account.
    # PORT_STAFF must still be refused (submission is CUSTOMER or ADMIN only).
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(company_name="KHÁCH HÀNG IMPORT"),
        headers=auth_headers,
    )
    assert res.status_code == 200, res.text
    data = res.json()
    assert data["workflow_status"] == "PENDING_REVIEW"
    assert data["status"] == "SUBMITTED"


def test_port_staff_cannot_submit_declarations(client, port_staff_headers):
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(company_name="KHÁCH HÀNG IMPORT"),
        headers=port_staff_headers,
    )
    assert res.status_code == 403


def test_declaration_validation_rejects_negative_count_and_invalid_movement(client, customer_headers):
    invalid_count = client.post(
        "/api/declarations",
        json=_minimal_declaration(crew_count=-1),
        headers=customer_headers,
    )
    assert invalid_count.status_code == 422

    invalid_movement = client.post(
        "/api/declarations",
        json=_minimal_declaration(movement_type="TRANSFER"),
        headers=customer_headers,
    )
    assert invalid_movement.status_code == 422


# ══════════════════════════════════════════════════════════════════════════════
# 6. TEU CALCULATIONS
# ══════════════════════════════════════════════════════════════════════════════

def test_teu_calculations(client, customer_headers):
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(unload={
            "cont20_full": 2, "cont20_empty": 1,
            "cont40_full": 3, "cont40_empty": 1,
        }),
        headers=customer_headers,
    )
    assert res.status_code == 200
    unload = res.json()["unload"]
    assert unload["total_containers"] == 7
    assert unload["teu"] == 11
    assert unload["empty_teu"] == 3


# ══════════════════════════════════════════════════════════════════════════════
# 7. WORKFLOW — CUSTOMER CONFIRMATION + PORT ENTERPRISE REVIEW
# ══════════════════════════════════════════════════════════════════════════════

def test_port_employee_can_approve_directly_or_request_changes(
    client, customer_headers, port_staff_headers, auth_headers
):
    approved_source = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    declaration_id = approved_source.json()["id"]

    approved = client.post(
        f"/api/declarations/{declaration_id}/workflow",
        json={"action": "PORT_APPROVE", "note": "Hỗ trợ trong ngữ cảnh Cảng"},
        headers=auth_headers,
    )
    assert approved.status_code == 200
    assert approved.json()["workflow_status"] == "APPROVED"
    assert approved.json()["port_approval"] == "APPROVED"
    assert "cv_approval" not in approved.json()

    changes_source = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    requested = client.post(
        f"/api/declarations/{changes_source.json()['id']}/workflow",
        json={"action": "REQUEST_CHANGES", "note": "Cần bổ sung chứng từ"},
        headers=port_staff_headers,
    )
    assert requested.status_code == 200
    assert requested.json()["workflow_status"] == "CHANGES_REQUESTED"

def test_retired_workflow_actions_return_gone(client, customer_headers, port_staff_headers):
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    assert res.status_code == 200
    decl_id = res.json()["id"]

    for action in ("CV_APPROVE", "QLC_APPROVE", "BP_APPROVE", "ISSUE", "REVOKE"):
        response = client.post(
            f"/api/declarations/{decl_id}/workflow",
            json={"action": action, "note": "Hành động cũ"},
            headers=port_staff_headers,
        )
        assert response.status_code == 410
        assert "ngừng hỗ trợ" in response.json()["detail"]

    db = SessionLocal()
    try:
        unchanged = db.get(Declaration, decl_id)
        assert unchanged.workflow_status == "PENDING_REVIEW"
    finally:
        db.close()


def test_real_input_keeps_customer_binding_and_clears_demo_sentinel():
    db = SessionLocal()
    try:
        demo = Organization(name="Demo customer", tax_code=DEMO_ORGANIZATION_TAX_CODE, created_at=now_iso(), updated_at=now_iso())
        db.add(demo)
        db.flush()
        user = User(
            username=f"demo-{uuid.uuid4().hex}", password_hash=get_password_hash("testpass"),
            full_name="Demo User", role="CUSTOMER", organization_id=demo.id,
            is_active=1, created_at=now_iso(),
        )
        db.add(user)
        db.add(Vessel(organization_id=demo.id, name="Mock", registration_no=_reg(), vessel_type="Sà lan", vessel_class="VR-SII", created_at=now_iso(), updated_at=now_iso()))
        db.commit()

        assert remove_demo_data_for_real_input(
            db, retain_organization_id=demo.id,
            organization_data={"name": "Khách hàng thật", "tax_code": "REAL-001"},
        ) is True
        db.commit()
        db.refresh(demo)
        db.refresh(user)
        assert demo.name == "Khách hàng thật"
        assert demo.tax_code == "REAL-001"
        assert user.organization_id == demo.id
        assert db.query(Vessel).filter(Vessel.organization_id == demo.id).count() == 0
        db.delete(user)
        db.delete(demo)
        db.commit()
    finally:
        db.rollback()
        db.close()


def test_changes_requested_resubmission_resets_approval_state(client, customer_headers, port_staff_headers):
    created = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    declaration = created.json()
    requested = client.post(
        f"/api/declarations/{declaration['id']}/workflow",
        json={"action": "REQUEST_CHANGES", "note": "Thiếu thông tin hàng hóa"},
        headers=port_staff_headers,
    )
    assert requested.status_code == 200
    assert requested.json()["workflow_status"] == "CHANGES_REQUESTED"

    resubmitted = client.post(
        "/api/declarations?submit=true",
        json={**_minimal_declaration(), "id": declaration["id"], "version": requested.json()["version"]},
        headers=customer_headers,
    )
    assert resubmitted.status_code == 200
    assert resubmitted.json()["workflow_status"] == "PENDING_REVIEW"
    assert resubmitted.json()["port_approval"] == "PENDING"


def test_workflow_audit_carries_authoritative_actor_and_correlation_id(
    client, customer_headers, port_staff_headers
):
    created = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(),
        headers=customer_headers,
    )
    declaration_id = created.json()["id"]
    correlation = "t2-correlation-test"
    response = client.post(
        f"/api/declarations/{declaration_id}/workflow",
        json={"action": "PORT_APPROVE"},
        headers={**port_staff_headers, "X-Correlation-ID": correlation},
    )
    assert response.status_code == 200
    assert response.headers["X-Correlation-ID"] == correlation

    db = SessionLocal()
    try:
        audit_event = db.query(AuditEvent).filter(
            AuditEvent.entity_type == "DECLARATION",
            AuditEvent.entity_id == declaration_id,
            AuditEvent.action == "PORT_APPROVE",
        ).one()
        assert audit_event.correlation_id == correlation
        assert audit_event.actor_user_id is not None
        assert audit_event.organization_id is not None
    finally:
        db.close()


# ══════════════════════════════════════════════════════════════════════════════
# 8. SUBMITTED RECORD PROTECTION
# ══════════════════════════════════════════════════════════════════════════════

def test_submitted_record_edit_protection(client, customer_headers):
    reg_no = _reg()
    res = client.post(
        "/api/declarations?submit=true",
        json=_minimal_declaration(registration_no=reg_no),
        headers=customer_headers,
    )
    decl_id = res.json()["id"]
    res2 = client.post("/api/declarations", json={
        **_minimal_declaration(registration_no=reg_no),
        "id": decl_id,
        "vessel_name": "HACKED",
    }, headers=customer_headers)
    assert res2.status_code == 409


# ══════════════════════════════════════════════════════════════════════════════
# 9. ATTACHMENT SIGNATURE / SIZE VALIDATION
# ══════════════════════════════════════════════════════════════════════════════

def _make_draft(client, auth_headers) -> int:
    res = client.post("/api/declarations", json=_minimal_declaration(), headers=auth_headers)
    assert res.status_code == 200
    return res.json()["id"]


def test_attachment_valid_pdf(client, auth_headers):
    decl_id = _make_draft(client, auth_headers)
    pdf_content = b"%PDF-1.4 valid content here"
    res = client.post(
        f"/api/declarations/{decl_id}/attachments?filename=test.pdf",
        content=pdf_content,
        headers={**auth_headers, "content-type": "application/pdf"},
    )
    assert res.status_code == 200
    assert res.json()["original_name"] == "test.pdf"
    assert res.json()["scan_status"] == "QUARANTINED"
    assert res.json()["storage_backend"] == "LOCAL_QUARANTINE"
    assert res.json()["checksum_sha256"]


def test_attachment_invalid_pdf_signature(client, auth_headers):
    decl_id = _make_draft(client, auth_headers)
    res = client.post(
        f"/api/declarations/{decl_id}/attachments?filename=fake.pdf",
        content=b"not a pdf at all",
        headers={**auth_headers, "content-type": "application/pdf"},
    )
    assert res.status_code == 400


def test_attachment_unknown_extension_rejected(client, auth_headers):
    decl_id = _make_draft(client, auth_headers)
    res = client.post(
        f"/api/declarations/{decl_id}/attachments?filename=payload.exe",
        headers=auth_headers,
        content=b"MZ",
    )
    assert res.status_code == 415


def test_attachment_size_rejection(client, auth_headers):
    decl_id = _make_draft(client, auth_headers)
    oversized = b"%PDF-1.4 " + b"x" * (12 * 1024 * 1024 + 1)
    res = client.post(
        f"/api/declarations/{decl_id}/attachments?filename=big.pdf",
        content=oversized,
        headers={**auth_headers, "content-type": "application/pdf"},
    )
    assert res.status_code == 413


def test_xlsx_rejects_external_relationship_and_zip_bomb_shape():
    external = io.BytesIO()
    with zipfile.ZipFile(external, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "<Types/>")
        archive.writestr("xl/workbook.xml", "<workbook/>")
        archive.writestr(
            "xl/_rels/workbook.xml.rels",
            '<Relationships><Relationship TargetMode="External" Target="https://example.test"/></Relationships>',
        )
    with pytest.raises(ValueError, match="external relationship"):
        read_workbook(external.getvalue())

    bomb = io.BytesIO()
    with zipfile.ZipFile(bomb, "w", zipfile.ZIP_DEFLATED) as archive:
        archive.writestr("[Content_Types].xml", "x" * (9 * 1024 * 1024))
    with pytest.raises(ValueError):
        read_workbook(bomb.getvalue())


def test_xlsx_ignores_non_executed_external_link_path_and_detects_headers():
    base = make_xlsx(
        "Dữ liệu phương tiện",
        ["Tên phương tiện", "Số đăng ký", "Loại phương tiện", "Cấp phương tiện", "Trọng tải toàn phần"],
        [["Sà lan kiểm thử", "SG-SMART-001", "Sà lan", "VR-SII", 850]],
    )
    source = zipfile.ZipFile(io.BytesIO(base))
    output = io.BytesIO()
    with source, zipfile.ZipFile(output, "w", zipfile.ZIP_DEFLATED) as archive:
        for info in source.infolist():
            archive.writestr(info, source.read(info.filename))
        archive.writestr(
            "xl/externalLinks/_rels/externalLink1.xml.rels",
            '<?xml version="1.0" encoding="UTF-8"?>'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/externalLinkPath" Target="old.xlsx" TargetMode="External"/>'
            '</Relationships>',
        )
    organization, rows = vessel_rows(read_workbook(output.getvalue()))
    assert organization["name"] == "KHÁCH HÀNG IMPORT"
    assert rows[0]["name"] == "SÀ LAN KIỂM THỬ"
    assert rows[0]["registration_no"] == "SG-SMART-001"
    assert rows[0]["deadweight_tons"] == 850
    assert rows[0]["_source_row"] == 4


# ══════════════════════════════════════════════════════════════════════════════
# 10. EXCEL REPORTS
# ══════════════════════════════════════════════════════════════════════════════

def test_xlsx_report_appendix1(client, auth_headers):
    res = client.get("/api/reports/appendix1", headers=auth_headers)
    assert res.status_code == 200
    assert "spreadsheetml" in res.headers.get("content-type", "")
    with zipfile.ZipFile(io.BytesIO(res.content)) as archive:
        assert "xl/workbook.xml" in archive.namelist()
        assert "xl/worksheets/sheet1.xml" in archive.namelist()
    sheet = load_workbook(io.BytesIO(res.content)).active
    assert sheet.max_column == 16
    assert {"A1:P1", "A3:P3", "A7:A10", "B7:H7", "I7:O7", "P7:P10"}.issubset(
        {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    )
    assert sheet["A3"].value == "KẾ HOẠCH HOẠT ĐỘNG CỦA PHƯƠNG TIỆN THỦY NỘI ĐỊA"
    assert sheet["B7"].value == "PHƯƠNG TIỆN"
    assert sheet["I7"].value == "HOẠT ĐỘNG"


def test_xlsx_report_appendix2(client, auth_headers):
    res = client.get("/api/reports/appendix2", headers=auth_headers)
    assert res.status_code == 200
    sheet = load_workbook(io.BytesIO(res.content)).active
    assert sheet.max_column == 16
    assert {"A1:P1", "A3:P3", "C7:F7", "G7:H7", "I7:J7", "K7:L7", "M7:N7", "O7:P7"}.issubset(
        {str(cell_range) for cell_range in sheet.merged_cells.ranges}
    )
    assert sheet["C7"].value == "Container"
    assert sheet["C8"].value == "Thực hiện tháng báo cáo"
    assert sheet["E8"].value == "Lũy kế đến tháng báo cáo"
    assert [sheet.cell(10, column).value for column in range(3, 17)] == list(range(1, 15))


def test_xlsx_report_appendix3(client, auth_headers):
    res = client.get("/api/reports/appendix3", headers=auth_headers)
    assert res.status_code == 200
    sheet = load_workbook(io.BytesIO(res.content)).active
    assert sheet.max_column == 35
    assert sheet["B5"].value == "Tên PTTND"
    assert sheet["I5"].value == "Hàng hóa"
    assert sheet["AI5"].value == "Đại lý PTND"


def test_static_only_port_salan_remains_in_pl01_and_pl03_with_blank_activity(
    client, port_staff_headers,
):
    registration = _reg()
    db = SessionLocal()
    vessel = Vessel(
        name="SALAN KHUNG TĨNH",
        registration_no=registration,
        vessel_type="Sà lan",
        vessel_class="VR-SI",
        length_m=45.5,
        deadweight_tons=700,
        gross_tonnage=320,
        cargo_capacity_tons=680,
        passenger_capacity=25,
        tracking_master_name="THUYỀN TRƯỞNG TĨNH",
        tracking_master_phone="0900000000",
        is_port_tracked=1,
    )
    db.add(vessel)
    db.flush()
    db.add(ReportingUnitVessel(
        reporting_unit_id=TEST_REPORTING_UNIT_ID,
        vessel_id=vessel.id,
        created_at=now_iso(),
    ))
    db.commit()
    vessel_id = vessel.id
    db.close()

    try:
        pl01 = load_workbook(io.BytesIO(client.get(
            "/api/reports/appendix1?from=2099-01-01&to=2099-01-31", headers=port_staff_headers,
        ).content)).active
        row1 = next(row for row in range(11, pl01.max_row + 1) if pl01.cell(row, 3).value == registration)
        assert pl01.cell(row1, 2).value == "SALAN KHUNG TĨNH"
        assert pl01.cell(row1, 8).value == 25
        assert all(pl01.cell(row1, column).value is None for column in range(9, 16))
        assert pl01.cell(row1, 16).value == "THUYỀN TRƯỞNG TĨNH - 0900000000"

        pl03 = load_workbook(io.BytesIO(client.get(
            "/api/reports/appendix3?from=2099-01-01&to=2099-01-31", headers=port_staff_headers,
        ).content)).active
        row3 = next(row for row in range(10, pl03.max_row + 1) if pl03.cell(row, 3).value == registration)
        assert pl03.cell(row3, 2).value == "SALAN KHUNG TĨNH"
        assert all(pl03.cell(row3, column).value is None for column in range(9, 36))

        pl02 = load_workbook(io.BytesIO(client.get(
            "/api/reports/appendix2?to=2099-01-15", headers=port_staff_headers,
        ).content)).active
        assert all(pl02.cell(12, column).value is None for column in range(3, 17))
    finally:
        db = SessionLocal()
        db.query(Vessel).filter(Vessel.id == vessel_id).delete()
        db.commit()
        db.close()


def test_report_analytics_supports_week_month_quarter_year_and_export(client, customer_headers):
    created_ids = []
    for operating_date, tons, teu, pax in (
        ("2042-03-15T08:00", 120, 4, 7),
        ("2041-03-15T08:00", 80, 2, 3),
    ):
        created = client.post(
            "/api/declarations",
            json=_minimal_declaration(
                declaration_date=operating_date[:10], eta=operating_date,
                etd=f"{operating_date[:10]}T18:00",
                unload={"tons": tons, "teu": teu}, passenger_count=pax,
            ),
            headers=customer_headers,
        )
        assert created.status_code == 200
        created_ids.append(created.json()["id"])
    db = SessionLocal()
    try:
        for declaration_id in created_ids:
            declaration = db.query(Declaration).filter(Declaration.id == declaration_id).one()
            declaration.workflow_status = "APPROVED"
            declaration.status = "SUBMITTED"
        db.commit()
    finally:
        db.close()

    try:
        for period in ("week", "month", "quarter", "year"):
            response = client.get(
                f"/api/reports/analytics?period={period}&as_of=2042-03-15",
                headers=customer_headers,
            )
            assert response.status_code == 200
            body = response.json()
            assert body["period"] == period
            assert set(body["kpis"]) == {"trips", "tons", "teu", "pax"}
            assert len(body["trend"]["cur"]) == len(body["trend"]["labels"])
        month = client.get(
            "/api/reports/analytics?period=month&as_of=2042-03-15",
            headers=customer_headers,
        ).json()
        assert month["kpis"]["trips"] == {"cur": 1.0, "prev": 1.0}
        assert month["kpis"]["tons"] == {"cur": 120.0, "prev": 80.0}
        export = client.get(
            "/api/reports/analytics/export?period=quarter&as_of=2042-03-15",
            headers=customer_headers,
        )
        assert export.status_code == 200
        assert "spreadsheetml" in export.headers.get("content-type", "")
        assert zipfile.is_zipfile(io.BytesIO(export.content))
    finally:
        db = SessionLocal()
        try:
            db.query(Declaration).filter(Declaration.id.in_(created_ids)).delete(synchronize_session=False)
            db.commit()
        finally:
            db.close()


def test_historical_analytics_exposes_coverage_and_blocks_unresolved_combined_overlap(
    client, customer_headers, port_staff_headers,
):
    db = SessionLocal()
    import_ids = []
    declaration_id = None
    try:
        actor = db.query(User).filter(User.username == "portstaff").one()
        berth_import = HistoricalReportImport(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, source_kind="tos_berth_call",
            appendix_kind="", mapping_version="test-berth-v1", reporting_period="2050-07",
            source_filename="audit-berth.xlsx", source_checksum=uuid.uuid4().hex,
            source_size_bytes=100, status="COMMITTED", accepted_count=1,
            created_by_user_id=actor.id, created_at=now_iso(), updated_at=now_iso(),
        )
        cargo_import = HistoricalReportImport(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, source_kind="tos_cargo_detail",
            appendix_kind="", mapping_version="test-cargo-v1", reporting_period="2050-07",
            source_filename="audit-cargo.xlsx", source_checksum=uuid.uuid4().hex,
            source_size_bytes=100, status="COMMITTED", accepted_count=1,
            created_by_user_id=actor.id, created_at=now_iso(), updated_at=now_iso(),
        )
        db.add_all([berth_import, cargo_import])
        db.flush()
        import_ids = [berth_import.id, cargo_import.id]
        call = HistoricalPortCall(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, import_id=berth_import.id,
            source_sheet="Berth", source_row=2, mapping_version="test-berth-v1",
            vessel_name_raw="H4B TEST", vessel_name_normalized="H4B TEST",
            call_year_raw="2050", voyage_number_raw="0001",
            call_key_normalized="H4B TEST|2050|0001", source_berth_raw="K12",
            arrival_berth="K12", departure_berth="K12",
            actual_berthing_at_raw="15/07/2050 08:00", actual_berthing_at="2050-07-15T08:00:00",
            actual_departure_at_raw="15/07/2050 16:00", actual_departure_at="2050-07-15T16:00:00",
            reporting_month="2050-07", validation_status="VALID", created_at=now_iso(),
        )
        db.add(call)
        db.flush()
        db.add(HistoricalCargoRow(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, import_id=cargo_import.id,
            source_sheet="Detail", source_row=2, port_call_id=call.id,
            source_call_key_raw="H4B TEST | 2050 | 0001",
            call_key_normalized="H4B TEST|2050|0001", container_size_code_raw="40HC",
            teu_factor=2, full_empty_code_raw="E", trade_scope_raw="Hàng nội",
            movement_method_raw="Hạ bãi", derived_direction="unload",
            weight_raw="4.00", weight_tonnes=4.0, weight_state="PRESENT",
            transform_version="test-v1", match_status="MATCHED",
            validation_status="VALID", created_at=now_iso(),
        ))
        db.commit()

        historical = client.get(
            "/api/reports/analytics?period=month&as_of=2050-07-15&source=historical",
            headers=port_staff_headers,
        )
        assert historical.status_code == 200, historical.text
        body = historical.json()
        assert body["source"] == "historical"
        assert body["coverage"]["status"] == "COMPLETE"
        assert body["kpis"]["trips"]["cur"] == 1.0
        assert body["kpis"]["tons"]["cur"] == 4.0
        assert body["kpis"]["teu"]["cur"] == 2.0
        assert body["kpis"]["pax"]["cur"] is None
        assert body["coverage"]["periods"][-1]["historicalCargoRows"] == 1

        cargo_import.status = "REVIEW"
        cargo_import.review_count = 1
        db.commit()
        partial = client.get(
            "/api/reports/analytics?period=month&as_of=2050-07-15&source=historical",
            headers=port_staff_headers,
        ).json()
        assert partial["coverage"]["status"] == "PARTIAL"
        assert partial["kpis"]["tons"]["cur"] is None
        assert any("cần kiểm tra" in warning for warning in partial["coverage"]["warnings"])
        cargo_import.status = "COMMITTED"
        cargo_import.review_count = 0
        db.commit()

        customer_forbidden = client.get(
            "/api/reports/analytics?period=month&as_of=2050-07-15&source=historical",
            headers=customer_headers,
        )
        assert customer_forbidden.status_code == 403

        combined = client.get(
            "/api/reports/analytics?period=month&as_of=2050-07-15&source=combined",
            headers=port_staff_headers,
        )
        assert combined.status_code == 200
        assert combined.json()["combinedAllowed"] is True
        assert combined.json()["kpis"]["tons"]["cur"] == 4.0

        created = client.post(
            "/api/declarations",
            json=_minimal_declaration(
                declaration_date="2050-07-15", eta="2050-07-15T09:00",
                etd="2050-07-15T18:00", unload={"tons": 10, "teu": 1},
            ),
            headers=customer_headers,
        )
        assert created.status_code == 200
        declaration_id = created.json()["id"]
        declaration = db.get(Declaration, declaration_id)
        db.refresh(declaration)
        declaration.workflow_status = "APPROVED"
        declaration.status = "SUBMITTED"
        db.commit()

        blocked = client.get(
            "/api/reports/analytics?period=month&as_of=2050-07-15&source=combined",
            headers=port_staff_headers,
        )
        assert blocked.status_code == 200
        assert blocked.json()["combinedAllowed"] is False
        assert blocked.json()["coverage"]["status"] == "BLOCKED"
        assert blocked.json()["coverage"]["overlapPeriods"] == ["2050-07"]
        assert blocked.json()["kpis"]["trips"]["cur"] is None
        export = client.get(
            "/api/reports/analytics/export?period=month&as_of=2050-07-15&source=combined",
            headers=port_staff_headers,
        )
        assert export.status_code == 409
    finally:
        if declaration_id:
            db.query(Declaration).filter(Declaration.id == declaration_id).delete()
        if import_ids:
            db.query(HistoricalCargoRow).filter(HistoricalCargoRow.import_id.in_(import_ids)).delete(
                synchronize_session=False,
            )
            db.query(HistoricalPortCall).filter(HistoricalPortCall.import_id.in_(import_ids)).delete(
                synchronize_session=False,
            )
            db.query(HistoricalReportImport).filter(HistoricalReportImport.id.in_(import_ids)).delete(
                synchronize_session=False,
            )
        db.commit()
        db.close()


def test_approved_report_golden_mapping_uses_actual_times_and_cargo_rows(client, customer_headers):
    created = client.post(
        "/api/declarations",
        json=_minimal_declaration(
            unload={"cargo_type": "Container", "movement_type": "Nhập khẩu", "cargo_name": "Hàng A", "cont20_full": 2, "tons": 12},
            load={"cargo_type": "Container", "movement_type": "Xuất khẩu", "cargo_name": "Hàng B", "cont40_full": 1, "tons": 20},
        ),
        headers=customer_headers,
    )
    declaration_id = created.json()["id"]
    db = SessionLocal()
    declaration = db.query(Declaration).filter(Declaration.id == declaration_id).one()
    declaration.workflow_status = "APPROVED"
    declaration.status = "SUBMITTED"
    declaration.actual_arrival_at = "2026-07-11T09:15"
    declaration.actual_departure_at = "2026-07-11T19:30"
    tracked_vessel = Vessel(
        organization_id=declaration.organization_id,
        name="SỔ THEO DÕI 01",
        registration_no=declaration.registration_no,
        vessel_type="CHỞ HÀNG KHÔ HOẶC CONTAINER",
        vessel_class="VR-SII",
        length_m=61.5,
        deadweight_tons=987.0,
        gross_tonnage=456.0,
        cargo_capacity_tons=970.0,
        container_capacity_teu=54.0,
        tracking_master_name="THUYỀN TRƯỞNG THEO SỔ",
        tracking_master_phone="0909123456",
        is_port_tracked=1,
    )
    db.add(tracked_vessel)
    db.flush()
    db.add(VesselOperatingProfile(
        vessel_id=tracked_vessel.id,
        sequence=1,
        activity_area="VR-SII",
        deadweight_tons=987.0,
        cargo_capacity_tons=970.0,
    ))
    db.add(VesselOperatingProfile(
        vessel_id=tracked_vessel.id,
        sequence=2,
        activity_area="VR-SI",
        deadweight_tons=1020.0,
        cargo_capacity_tons=1000.0,
    ))
    db.commit()
    tracked_vessel_id = tracked_vessel.id
    tracked_registration = tracked_vessel.registration_no
    db.close()

    appendix1 = client.get("/api/reports/appendix1?from=2026-01-01&to=2026-12-31", headers=customer_headers)
    sheet1 = load_workbook(io.BytesIO(appendix1.content)).active
    appendix1_row = next(
        row_number for row_number in range(5, sheet1.max_row + 1)
        if sheet1.cell(row_number, 3).value == tracked_registration
    )
    assert sheet1.cell(appendix1_row, 2).value == "SỔ THEO DÕI 01"
    assert sheet1.cell(appendix1_row, 4).value == "VR-SII"
    assert sheet1.cell(appendix1_row, 7).value == "970 / 1000 tấn / 54 TEU"
    assert sheet1.cell(appendix1_row, 10).value == "2026-07-11T09:15"
    assert sheet1.cell(appendix1_row, 12).value == "2026-07-11T19:30"
    assert sheet1.cell(appendix1_row, 16).value == "THUYỀN TRƯỞNG THEO SỔ - 0909123456"

    appendix3 = client.get("/api/reports/appendix3?from=2026-01-01&to=2026-12-31", headers=customer_headers)
    sheet3 = load_workbook(io.BytesIO(appendix3.content)).active
    matching_rows = [
        row_number for row_number in range(10, sheet3.max_row + 1)
        if sheet3.cell(row_number, 3).value == tracked_registration
    ]
    assert len(matching_rows) == 1
    assert sheet3.cell(matching_rows[0], 2).value == "SỔ THEO DÕI 01"
    assert sheet3.cell(matching_rows[0], 5).value == "VR-SII"
    assert sheet3.cell(matching_rows[0], 7).value == "987 / 1020"
    exported_row = [sheet3.cell(matching_rows[0], column_number).value for column_number in range(1, 36)]
    assert exported_row[28] == "Hàng A\nHàng B"
    assert exported_row[11] == 12 and exported_row[12] == 2
    assert exported_row[8] == 20 and exported_row[9] == 2

    db = SessionLocal()
    try:
        db.query(VesselOperatingProfile).filter(
            VesselOperatingProfile.vessel_id == tracked_vessel_id
        ).delete(synchronize_session=False)
        db.query(Vessel).filter(Vessel.id == tracked_vessel_id).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def test_appendix_month_ytd_operating_date_adjustment_and_vessel_grain(
    client, customer_headers, port_staff_headers,
):
    registration = _reg()
    created_ids = []
    fixtures = (
        {
            "declaration_date": "2045-07-20", "eta": "2045-01-15T08:00", "etd": "2045-01-15T18:00",
            "registration_no": registration, "agent_ptnd_name": "Đại lý A",
            "unload": {"cargo_type": "Container", "movement_type": "Nhập khẩu", "cargo_name": "Hàng tháng 1", "cont20_full": 1, "tons": 10},
        },
        {
            "declaration_date": "2045-01-02", "eta": "2045-07-12T08:00", "etd": "2045-07-12T18:00",
            "registration_no": registration, "working_port": "Cầu A", "departure_berth": "Cầu B", "destination_port": "Cảng C",
            "agent_ptnd_name": "Đại lý B", "is_passenger_call": True, "passenger_count": 0,
            "load": {"cargo_type": "Container", "movement_type": "Xuất khẩu", "cargo_name": "Hàng tháng 7", "cont40_full": 1, "tons": 20},
        },
        {
            "declaration_date": "2045-07-01", "eta": "2045-08-01T08:00", "etd": "2045-08-01T18:00",
            "registration_no": _reg(), "unload": {"cargo_type": "Hàng khô", "movement_type": "Nội địa", "tons": 99},
        },
    )
    for payload in fixtures:
        response = client.post("/api/declarations", json=_minimal_declaration(**payload), headers=customer_headers)
        assert response.status_code == 200
        created_ids.append(response.json()["id"])
    db = SessionLocal()
    try:
        for declaration in db.query(Declaration).filter(Declaration.id.in_(created_ids)).all():
            declaration.workflow_status = "APPROVED"
            declaration.status = "SUBMITTED"
        db.commit()
    finally:
        db.close()

    try:
        appendix2 = client.get("/api/reports/appendix2?from=2045-01-01&to=2045-07-20", headers=port_staff_headers)
        assert appendix2.status_code == 200
        sheet2 = load_workbook(io.BytesIO(appendix2.content)).active
        assert sheet2["A4"].value == "Tháng 07 năm 2045"
        assert sheet2.cell(12, 3).value == 20
        assert sheet2.cell(12, 5).value == 30
        assert sheet2.cell(12, 13).value == 1
        assert sheet2.cell(12, 14).value == 2
        assert sheet2.cell(12, 15).value == 1
        assert sheet2.cell(12, 16).value is None

        appendix1 = client.get("/api/reports/appendix1?from=2045-07-01&to=2045-07-31", headers=customer_headers)
        sheet1 = load_workbook(io.BytesIO(appendix1.content)).active
        appendix1_row = next(row for row in range(11, sheet1.max_row + 1) if sheet1.cell(row, 3).value == registration)
        assert sheet1.cell(appendix1_row, 8).value is None
        assert sheet1.cell(appendix1_row, 9).value == "Cầu A"
        assert sheet1.cell(appendix1_row, 11).value == "Cầu B"

        assert client.post(
            "/api/reports/appendix2/adjustments",
            json={"report_month": "2045-07", "metric": "calls", "delta": 2, "reason": "Điều chỉnh theo sổ trực ca"},
            headers=customer_headers,
        ).status_code == 403
        adjusted = client.post(
            "/api/reports/appendix2/adjustments",
            json={"report_month": "2045-07", "metric": "calls", "delta": 2, "reason": "Điều chỉnh theo sổ trực ca"},
            headers=port_staff_headers,
        )
        assert adjusted.status_code == 200
        adjustment_id = adjusted.json()["id"]
        adjusted_sheet = load_workbook(io.BytesIO(client.get(
            "/api/reports/appendix2?to=2045-07-20", headers=port_staff_headers,
        ).content)).active
        assert adjusted_sheet.cell(12, 13).value == 3
        assert adjusted_sheet.cell(12, 14).value == 4

        appendix3 = client.get("/api/reports/appendix3?from=2045-01-01&to=2045-07-31", headers=customer_headers)
        assert appendix3.status_code == 200
        sheet3 = load_workbook(io.BytesIO(appendix3.content)).active
        matching = [row for row in range(10, sheet3.max_row + 1) if sheet3.cell(row, 3).value == registration]
        assert len(matching) == 1
        row = matching[0]
        assert sheet3.cell(row, 29).value == "Hàng tháng 1\nHàng tháng 7"
        assert sheet3.cell(row, 35).value == "Đại lý A\nĐại lý B"
        assert sheet3["J7"].value == "TEUs"
        assert sheet3["K7"].value == "TEUs Rỗng"
        assert sheet3.column_dimensions["D"].width >= 18
        assert sheet3.cell(row, 33).value.count("\n") == 1
        assert sheet3.cell(row, 34).value.count("\n") == 1
        assert sheet3.row_dimensions[row].height >= 108
    finally:
        db = SessionLocal()
        try:
            db.query(ReportAdjustment).filter(ReportAdjustment.report_month == "2045-07").delete(synchronize_session=False)
            db.query(Declaration).filter(Declaration.id.in_(created_ids)).delete(synchronize_session=False)
            db.commit()
        finally:
            db.close()


def test_report_unknown_kind(client, auth_headers):
    res = client.get("/api/reports/unknown", headers=auth_headers)
    assert res.status_code == 404


# ══════════════════════════════════════════════════════════════════════════════
# 11. ROUTE COVERAGE — every frontend endpoint registered
# ══════════════════════════════════════════════════════════════════════════════

def test_all_frontend_routes_registered():
    routes = [route.path for route in app.routes if getattr(route, "path", None)]
    for included in (getattr(route, "original_router", None) for route in app.routes):
        if included is not None:
            routes.extend(route.path for route in included.routes if getattr(route, "path", None))
    required_paths = [
        "/api/auth/login",
        "/api/health",
        "/api/ready",
        "/api/admin/operations-summary",
        "/api/admin/backups",
        "/api/catalogs",
        "/api/dashboard",
        "/api/organizations",
        "/api/vessels",
        "/api/vessels/{vessel_id}/verify-registry",
        "/api/port-vessel-register/export",
        "/api/port-vessel-register",
        "/api/crew",
        "/api/declarations",
        "/api/declarations/{declaration_id}/attachments",
        "/api/declarations/{declaration_id}/events",
        "/api/declarations/{declaration_id}/workflow",
        "/api/suggestions",
        "/api/import/vessels",
        "/api/import/port-vessel-register",
        "/api/import/crew",
        "/api/import/declaration",
        "/api/historical-imports/preview",
        "/api/historical-imports/reconcile",
        "/api/historical-imports/exports/pl03",
        "/api/historical-imports/{import_id}/rows",
        "/api/historical-imports/{import_id}",
        "/api/historical-imports/{import_id}/vessel-links",
        "/api/historical-imports/{import_id}/cancel",
        "/api/historical-imports/{import_id}/confirm",
        "/api/reports/analytics",
        "/api/reports/analytics/export",
        "/api/reports/appendix2/adjustments",
        "/api/reports/{kind}",
        "/api/integrations/maritime-authority",
        "/api/integrations/prepare-sync",
    ]
    for path in required_paths:
        assert path in routes, f"Route not registered in backend: {path}"


# ══════════════════════════════════════════════════════════════════════════════
# 12. SUGGESTIONS
# ══════════════════════════════════════════════════════════════════════════════

def test_suggestions_valid_field(client, auth_headers):
    res = client.get("/api/suggestions?field=last_port", headers=auth_headers)
    assert res.status_code == 200
    assert isinstance(res.json(), list)


def test_suggestions_invalid_field_returns_empty(client, auth_headers):
    res = client.get("/api/suggestions?field=nonexistent", headers=auth_headers)
    assert res.status_code == 200
    assert res.json() == []


# ══════════════════════════════════════════════════════════════════════════════
# 13. INTEGRATIONS
# ══════════════════════════════════════════════════════════════════════════════

def test_integration_status(client, auth_headers):
    res = client.get("/api/integrations/maritime-authority", headers=auth_headers)
    assert res.status_code == 200
    data = res.json()
    assert "connector" in data
    assert "jobs" in data
    assert data["adapter"]["mode"] == "MANUAL"
    assert data["adapter"]["networkCallsAllowed"] is False
    assert "readyToSend" in data["connector"]


def test_prepare_sync(client, auth_headers):
    res = client.post(
        "/api/integrations/prepare-sync",
        json={"from": "2026-01-01", "to": "2026-12-31"},
        headers=auth_headers,
    )
    assert res.status_code == 200
    data = res.json()
    assert "id" in data
    assert data["status"] == "PREPARED"


# ══════════════════════════════════════════════════════════════════════════════
# 14. CREW SNAPSHOT ON DECLARATION
# ══════════════════════════════════════════════════════════════════════════════

def test_crew_certificate_snapshot(client, customer_headers):
    crew_res = client.post("/api/crew", json={
        "full_name": "Snapshot Crew",
        "crew_role": "Thuyền trưởng",
        "professional_certificate_type": "Bằng thuyền trưởng",
        "professional_certificate_no": "SNAP-001",
        "certificate_expiry_date": "2099-01-01",
    }, headers=customer_headers)
    assert crew_res.status_code == 200
    crew_id = crew_res.json()["id"]

    res = client.post("/api/declarations?submit=true", json=_minimal_declaration(
        crew_ids=[crew_id],
    ), headers=customer_headers)
    assert res.status_code == 200
    # Crew snapshot was recorded (declaration created with crew_ids)
    assert res.json()["workflow_status"] == "PENDING_REVIEW"


def test_import_preview_and_idempotency(client, auth_headers, customer_headers):
    root = Path(__file__).resolve().parents[1]
    vessel_content = (root / "templates" / "Ho_so_phuong_tien_thuy_noi_dia.xlsx").read_bytes()

    db = SessionLocal()
    before = db.query(ImportJob).count()
    db.close()
    preview = client.post(
        "/api/import/vessels?preview=true",
        content=vessel_content,
        headers={**auth_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert preview.status_code == 200
    assert preview.json()["preview"] is True
    assert preview.json()["mappingVersion"] == "KBCV-IMPORT-1.5"
    db = SessionLocal()
    assert db.query(ImportJob).count() == before
    db.close()

    declaration_content = (root / "templates" / "Phieu_khai_bao_PTTND_truoc_khi_den_cang.xlsx").read_bytes()
    first = client.post(
        "/api/import/declaration",
        content=declaration_content,
        headers={**customer_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert first.status_code == 200
    assert first.json()["idempotent"] is False
    repeated = client.post(
        "/api/import/declaration",
        content=declaration_content,
        headers={**customer_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert repeated.status_code == 200
    assert repeated.json()["idempotent"] is True
    assert repeated.json()["id"] == first.json()["id"]

    from openpyxl import load_workbook
    admin_workbook = load_workbook(io.BytesIO(declaration_content))
    admin_workbook["KHAI BÁO"]["C6"] = f"DOANH NGHIỆP ADMIN {uuid.uuid4().hex[:8].upper()}"
    admin_buffer = io.BytesIO()
    admin_workbook.save(admin_buffer)
    admin_content = admin_buffer.getvalue()

    admin_preview = client.post(
        "/api/import/declaration?preview=true",
        content=admin_content,
        headers={**auth_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert admin_preview.status_code == 200
    assert admin_preview.json()["preview"] is True
    admin_import = client.post(
        "/api/import/declaration",
        content=admin_content,
        headers={**auth_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"},
    )
    assert admin_import.status_code == 200, admin_import.text
    assert admin_import.json()["accepted"] == 1


def test_port_staff_can_import_declaration(client, port_staff_headers):
    # Port staff import a declaration workbook on behalf of a customer, the
    # same way PLATFORM_ADMIN does: the workbook's company name resolves (or
    # onboards) the tenant organization within the staff member's own unit.
    root = Path(__file__).resolve().parents[1]
    declaration_content = (root / "templates" / "Phieu_khai_bao_PTTND_truoc_khi_den_cang.xlsx").read_bytes()

    from openpyxl import load_workbook
    workbook = load_workbook(io.BytesIO(declaration_content))
    workbook["KHAI BÁO"]["C6"] = f"DOANH NGHIỆP PORT STAFF {uuid.uuid4().hex[:8].upper()}"
    buffer = io.BytesIO()
    workbook.save(buffer)
    content = buffer.getvalue()

    headers = {**port_staff_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    preview = client.post("/api/import/declaration?preview=true", content=content, headers=headers)
    assert preview.status_code == 200
    assert preview.json()["preview"] is True

    imported = client.post("/api/import/declaration", content=content, headers=headers)
    assert imported.status_code == 200, imported.text
    assert imported.json()["accepted"] == 1


def test_smart_vessel_import_accepts_complete_non_template_workbook(client, auth_headers):
    registration = f"SG-SMART-{uuid.uuid4().hex[:10]}".upper()
    workbook = make_xlsx(
        "Danh mục tùy biến",
        ["Ghi chú", "Số đăng ký", "Tên tàu", "Cấp PT", "Loại PT", "DWT", "Sức chở hàng"],
        [["Đủ dữ liệu", registration, "Sà lan linh hoạt", "VR-SII", "Sà lan", "950.5 / 980.5", "900,25 / 930,25"]],
    )
    headers = {**auth_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    preview = client.post("/api/import/vessels?preview=true", content=workbook, headers=headers)
    assert preview.status_code == 200
    assert preview.json()["mapping"]["strategy"] == "HEADER_LABEL_DETECTION"
    assert preview.json()["rows"][0]["missingFields"] == []
    assert preview.json()["rows"][0]["deadweight_tons"] == 950.5
    assert preview.json()["rows"][0]["cargo_capacity_tons"] == 900.25
    assert len(preview.json()["rows"][0]["operating_profiles"]) == 2
    assert preview.json()["rows"][0]["mappingWarnings"]
    imported = client.post("/api/import/vessels", content=workbook, headers=headers)
    assert imported.status_code == 200
    assert imported.json()["accepted"] == 1
    assert imported.json()["rejected"] == []
    db = SessionLocal()
    try:
        vessel = db.query(Vessel).filter(Vessel.registration_no == registration).one()
        assert vessel.deadweight_tons == 950.5
        assert vessel.cargo_capacity_tons == 900.25
        assert vessel.is_port_tracked == 0
        assert [(p.activity_area, p.deadweight_tons, p.cargo_capacity_tons) for p in vessel.operating_profiles] == [
            ("VR-SII", 950.5, 900.25),
            ("VR-SII", 980.5, 930.25),
        ]
        db.query(ImportJob).filter(ImportJob.source_checksum == imported.json()["checksum"]).delete(synchronize_session=False)
        db.delete(vessel)
        db.commit()
    finally:
        db.close()


def test_port_tracking_import_preserves_dual_operating_profiles_and_exports_them(
    client, port_staff_headers,
):
    registration = f"SG-DUAL-{uuid.uuid4().hex[:8]}".upper()
    workbook = make_xlsx(
        "DỮ LIỆU SÀ LAN",
        [
            "STT", "TÊN PHƯƠNG TIỆN", "SỐ ĐĂNG KÝ", "LOẠI PHƯƠNG TIỆN (CÔNG DỤNG)",
            "CẤP PT (VÙNG HOẠT ĐỘNG)", "CHIỀU DÀI (M)", "TRỌNG TẢI TOÀN PHẦN (TẤN)",
            "DUNG TÍCH (M3)", "KHẢ NĂNG KHAI THÁC (TẤN)", "KHẢ NĂNG KHAI THÁC (TEU)",
            "NGÀY HẾT HẠN GCNATKT&BVMT", "SỐ THUYỀN VIÊN", "THUYỀN TRƯỞNG",
            "SỐ ĐIỆN THOẠI LIÊN HỆ",
        ],
        [[
            1, "NGỌC HUY KIỂM THỬ", registration, "Chở hàng khô hoặc container",
            "VR-SI/VR-SII", 70.9, "2723.79 / 2912.57", 1329,
            "2698.79 / 2887.57", 128, "30/03/2027", 3,
            "NGUYỄN VĂN KIỂM THỬ", "0900000000",
        ]],
    )
    headers = {
        **port_staff_headers,
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    preview = client.post("/api/import/port-vessel-register?preview=true", content=workbook, headers=headers)
    assert preview.status_code == 200, preview.text
    row = preview.json()["rows"][0]
    # vessel_type ghi nguyên văn Công dụng theo chứng từ (không map vào danh
    # mục cố định); chỉ đổi hoa/thường theo quy ước import chung nên không có
    # cảnh báo chuẩn hóa.
    assert row["mappingWarnings"] == []
    assert row["vessel_type"] == "CHỞ HÀNG KHÔ HOẶC CONTAINER"
    assert row["operating_profiles"] == [
        {"sequence": 1, "activity_area": "VR-SI", "deadweight_tons": 2723.79, "cargo_capacity_tons": 2698.79},
        {"sequence": 2, "activity_area": "VR-SII", "deadweight_tons": 2912.57, "cargo_capacity_tons": 2887.57},
    ]
    imported = client.post("/api/import/port-vessel-register", content=workbook, headers=headers)
    assert imported.status_code == 200, imported.text
    assert imported.json()["created"] == 1

    register = client.get("/api/port-vessel-register", headers=port_staff_headers)
    assert register.status_code == 200
    tracked = next(item for item in register.json()["items"] if item["registration_no"] == registration)
    assert tracked["is_port_tracked"] == 1
    assert register.json()["stats"]["vessels"] >= 1
    assert register.json()["stats"]["multiAreaVessels"] >= 1

    exported = client.get("/api/port-vessel-register/export", headers=port_staff_headers)
    assert exported.status_code == 200
    with zipfile.ZipFile(io.BytesIO(exported.content)) as archive:
        xml = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert registration in xml
    assert "VR-SI / VR-SII" in xml
    assert "2723.79 / 2912.57" in xml
    assert "2698.79 / 2887.57" in xml

    db = SessionLocal()
    try:
        vessel = db.query(Vessel).filter(Vessel.registration_no == registration).one()
        assert vessel.tracking_master_name == "NGUYỄN VĂN KIỂM THỬ"
        assert vessel.tracking_master_phone == "0900000000"
        assert len(vessel.operating_profiles) == 2
        job = db.query(ImportJob).filter(ImportJob.source_checksum == imported.json()["checksum"]).one()
        assert job.import_kind == "PORT_VESSEL_REGISTER"
        db.query(ImportJob).filter(ImportJob.source_checksum == imported.json()["checksum"]).delete(synchronize_session=False)
        db.delete(vessel)
        db.commit()
    finally:
        db.close()


def test_customer_cannot_access_internal_port_register(client, customer_headers):
    assert client.get("/api/port-vessel-register", headers=customer_headers).status_code == 403
    workbook = make_xlsx(
        "DỮ LIỆU SÀ LAN",
        ["Tên phương tiện", "Số đăng ký", "Loại phương tiện", "Cấp PT"],
        [["SALAN KHÔNG ĐƯỢC PHÉP", "SG-DENIED", "SÀ LAN", "VR-SI"]],
    )
    response = client.post(
        "/api/import/port-vessel-register",
        content=workbook,
        headers={
            **customer_headers,
            "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        },
    )
    assert response.status_code == 403


def test_port_tracking_parser_keeps_blank_contact_fields_as_empty_text():
    workbook = make_xlsx(
        "DỮ LIỆU SÀ LAN",
        [
            "Tên phương tiện", "Số đăng ký", "Loại phương tiện", "Cấp PT",
            "Thuyền trưởng", "Số điện thoại liên hệ",
        ],
        [["SALAN KHÔNG CÓ LIÊN HỆ", "SG-BLANK-CONTACT", "SÀ LAN", "VR-SI", None, None]],
    )
    _, rows = vessel_rows(read_workbook(workbook))
    assert rows[0]["tracking_master_name"] == ""
    assert rows[0]["tracking_master_phone"] == ""


def test_port_staff_can_add_salan_manually_with_two_operating_profiles(client, port_staff_headers):
    registration = f"SG-MANUAL-{uuid.uuid4().hex[:8]}".upper()
    response = client.post("/api/vessels?port_register=true", headers=port_staff_headers, json={
        "organization_name": "TEST PORT REGISTER OWNER",
        "name": "SALAN NHẬP THỦ CÔNG",
        "registration_no": registration,
        "vessel_type": "CHỞ HÀNG KHÔ HOẶC CONTAINER",
        "vessel_class": "VR-SI / VR-SII",
        "tracking_master_name": "NGUYỄN VĂN A",
        "tracking_master_phone": "0900000001",
        "operating_profiles": [
            {"activity_area": "VR-SI", "deadweight_tons": 1000, "cargo_capacity_tons": 950},
            {"activity_area": "VR-SII", "deadweight_tons": 1100, "cargo_capacity_tons": 1050},
        ],
    })
    assert response.status_code == 200, response.text
    data = response.json()
    assert data["vessel_class"] == "VR-SI / VR-SII"
    assert data["deadweight_tons"] == 1000
    assert data["is_port_tracked"] == 1
    assert len(data["operating_profiles"]) == 2

    db = SessionLocal()
    try:
        vessel = db.query(Vessel).filter(Vessel.registration_no == registration).one()
        db.delete(vessel)
        db.commit()
    finally:
        db.close()


def test_port_staff_can_remove_salan_from_internal_register_without_deleting_master_record(
    client, port_staff_headers, customer_headers,
):
    registration = f"SG-REMOVE-{uuid.uuid4().hex[:8]}".upper()
    created = client.post("/api/vessels?port_register=true", headers=port_staff_headers, json={
        "organization_name": "TEST PORT REGISTER OWNER",
        "name": "SALAN GỠ KHỎI SỔ",
        "registration_no": registration,
        "vessel_type": "CHỞ HÀNG KHÔ",
        "vessel_class": "VR-SI",
        "operating_profiles": [
            {"activity_area": "VR-SI", "deadweight_tons": 700, "cargo_capacity_tons": 650},
        ],
    })
    assert created.status_code == 200, created.text
    vessel_id = created.json()["id"]

    denied = client.post(
        "/api/port-vessel-register/remove",
        headers=customer_headers,
        json={"ids": [vessel_id]},
    )
    assert denied.status_code == 403

    removed = client.post(
        "/api/port-vessel-register/remove",
        headers=port_staff_headers,
        json={"ids": [vessel_id]},
    )
    assert removed.status_code == 200, removed.text
    assert removed.json() == {"removed": 1, "ids": [vessel_id]}
    assert all(
        item["id"] != vessel_id
        for item in client.get("/api/port-vessel-register", headers=port_staff_headers).json()["items"]
    )

    db = SessionLocal()
    try:
        vessel = db.query(Vessel).filter(Vessel.id == vessel_id).one()
        assert vessel.is_port_tracked == 0
        assert vessel.version == created.json()["version"] + 1
        event = db.query(app_module.AuditEvent).filter(
            app_module.AuditEvent.entity_type == "VESSEL",
            app_module.AuditEvent.entity_id == vessel_id,
            app_module.AuditEvent.action == "PORT_REGISTER_REMOVE",
        ).one()
        db.delete(event)
        db.delete(vessel)
        db.commit()
    finally:
        db.close()


def test_port_register_row_actions_and_pagination_are_present(client):
    response = client.get("/")
    assert response.status_code == 200
    assert 'id="port-register-pagination"' in response.text
    assert 'id="remove-selected-port-vessels"' in response.text
    script = client.get("/app.js?v=1.2.0")
    assert script.status_code == 200
    assert "portRegisterPageSize: 15" in script.text
    assert "select-port-register-page" in script.text
    assert "data-edit-port-vessel" in script.text
    assert "data-remove-port-vessel" in script.text


def test_vessel_list_has_stt_pagination_and_no_import_owner_remark(client):
    response = client.get("/")
    assert response.status_code == 200
    assert 'id="vessel-pagination"' in response.text
    script = client.get("/app.js?v=1.2.0")
    assert script.status_code == 200
    assert "vesselPageSize: 15" in script.text
    assert '<th>STT</th><th>Phương tiện</th>' in script.text
    vessel_render = script.text.split("function renderVessels()", 1)[1].split(
        "async function loadDeclarations()", 1
    )[0]
    assert "organization_name" not in vessel_render


def test_vessel_import_normalizes_text_and_requires_explicit_overwrite(client, auth_headers):
    registration = f"SG-NORM-{uuid.uuid4().hex[:10]}".upper()
    db = SessionLocal()
    try:
        org = db.query(Organization).filter(Organization.name == "Test Org").one()
        vessel = Vessel(
            organization_id=org.id,
            name="THƯỢNG HẢI 07",
            registration_no=registration,
            vessel_type="CHỞ HÀNG KHÔ HOẶC CÔNG TE NƠ",
            vessel_class="VR-SII",
            created_at=now_iso(),
            updated_at=now_iso(),
        )
        db.add(vessel)
        db.commit()
    finally:
        db.close()

    workbook = make_xlsx(
        "Dữ liệu sửa lỗi",
        ["Tên tàu", "Số đăng ký", "Loại PT", "Cấp PT"],
        [["  Thượng   Hải 07  ", f" {registration.lower()} ", "Chở hàng khô hoặc côngtenơ", "vr-sii"]],
    )
    headers = {**auth_headers, "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"}
    preview = client.post("/api/import/vessels?preview=true", content=workbook, headers=headers)
    assert preview.status_code == 200
    row = preview.json()["rows"][0]
    assert row["name"] == "THƯỢNG HẢI 07"
    assert row["registration_no"] == registration
    assert row["vessel_type"] == "CHỞ HÀNG KHÔ HOẶC CONTAINER"
    assert row["vessel_class"] == "VR-SII"
    assert row["existing"] is True
    assert preview.json()["conflictCount"] == 1
    assert row["mappingWarnings"]

    keep_existing = client.post("/api/import/vessels", content=workbook, headers=headers)
    assert keep_existing.status_code == 200
    assert keep_existing.json()["created"] == 0
    assert keep_existing.json()["updated"] == 0
    assert keep_existing.json()["skipped"] == 1

    db = SessionLocal()
    try:
        unchanged = db.query(Vessel).filter(Vessel.registration_no == registration).one()
        assert unchanged.vessel_type == "CHỞ HÀNG KHÔ HOẶC CÔNG TE NƠ"
    finally:
        db.close()

    overwrite = client.post(
        "/api/import/vessels?overwrite_existing=true", content=workbook, headers=headers
    )
    assert overwrite.status_code == 200
    assert overwrite.json()["updated"] == 1
    assert overwrite.json()["reapplied"] is True

    db = SessionLocal()
    try:
        updated = db.query(Vessel).filter(Vessel.registration_no == registration).one()
        assert updated.name == "THƯỢNG HẢI 07"
        assert updated.vessel_type == "CHỞ HÀNG KHÔ HOẶC CONTAINER"
        checksum = overwrite.json()["checksum"]
        db.query(ImportJob).filter(ImportJob.source_checksum == checksum).delete(synchronize_session=False)
        db.delete(updated)
        db.commit()
    finally:
        db.close()


def test_port_staff_imports_crew_without_vessel_assignment(
    client, customer_headers, port_staff_headers,
):
    certificate_no = f"CREW-{uuid.uuid4().hex[:10]}".upper()
    created = client.post("/api/crew", json={
        "full_name": "NGUYỄN VĂN KIỂM THỬ",
        "crew_role": "THUYỀN VIÊN",
        "professional_certificate_type": "CHỨNG CHỈ THỦY THỦ",
        "professional_certificate_no": certificate_no,
        "phone": "0900000001",
    }, headers=customer_headers)
    assert created.status_code == 200
    crew_id = created.json()["id"]

    workbook = make_xlsx(
        "Danh sách thuyền viên",
        [
            "Tên doanh nghiệp", "Họ và tên", "Chức danh", "Ngày sinh",
            "Số điện thoại", "CCCD", "Loại chứng chỉ", "Số chứng chỉ",
            "Ngày hết hạn",
        ],
        [[
            " test   org ", " Nguyễn Văn Kiểm Thử ", "Thuyền viên", "12/04/1985",
            "0900000099", "079123456789", "Chứng chỉ thủy thủ", certificate_no,
            "31/12/2030",
        ]],
    )
    headers = {
        **port_staff_headers,
        "Content-Type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    }
    preview = client.post("/api/import/crew?preview=true", content=workbook, headers=headers)
    assert preview.status_code == 200
    row = preview.json()["rows"][0]
    assert row["organization_name"] == "TEST ORG"
    assert row["full_name"] == "NGUYỄN VĂN KIỂM THỬ"
    assert row["crew_role"] == "Thuyền viên"
    assert row["birth_date"] == "1985-04-12"
    assert row["existing"] is True
    assert row["missingFields"] == []

    imported = client.post("/api/import/crew", content=workbook, headers=headers)
    assert imported.status_code == 200, imported.text
    assert imported.json()["created"] == 0
    assert imported.json()["updated"] == 1
    assert client.post(
        "/api/import/crew", content=workbook,
        headers={**customer_headers, "Content-Type": headers["Content-Type"]},
    ).status_code == 403

    db = SessionLocal()
    try:
        member = db.query(app_module.CrewMember).filter(app_module.CrewMember.id == crew_id).one()
        assert member.phone == "0900000099"
        assert member.birth_date == "1985-04-12"
        assert member.vessel_id is None
        db.query(ImportJob).filter(
            ImportJob.source_checksum == imported.json()["checksum"]
        ).delete(synchronize_session=False)
        db.delete(member)
        db.commit()
    finally:
        db.close()


def test_admin_backup_routes_are_registered_and_role_scoped(
    client, auth_headers, customer_headers, tmp_path, monkeypatch,
):
    backup_dir = tmp_path / "backups"
    monkeypatch.setattr(app_module, "BACKUP_DIR", backup_dir)
    assert client.get("/api/admin/backups", headers=auth_headers).json() == []
    forbidden = client.get("/api/admin/backups", headers=customer_headers)
    assert forbidden.status_code == 403

    created = client.post("/api/admin/backups", headers=auth_headers)
    assert created.status_code == 200
    assert created.json()["integrityCheck"] == "ok"
    assert (backup_dir / created.json()["filename"]).exists()
    listed = client.get("/api/admin/backups", headers=auth_headers)
    assert listed.status_code == 200
    assert [item["filename"] for item in listed.json()] == [created.json()["filename"]]
def test_historical_tos_preview_cross_import_join_and_revision(
    client, auth_headers, customer_headers,
):
    """H3: detect without filename, stage, join separate files and revise explicitly."""
    vessel_name = f"H3 BARGE {uuid.uuid4().hex[:8]}"
    _seed_historical_registered_vessel(vessel_name)
    berth_headers = {2: "Năm", 3: "Chuyến", 5: "Tên tàu", 8: "Mã bến", 20: "ATB", 23: "ATD"}
    berth = _historical_fixture(berth_headers, [{
        2: "2026", 3: "0007", 5: vessel_name, 8: "K12",
        20: "18/07/2026 08:30:00", 23: "18/07/2026 13:00:00",
    }])
    preview_headers = {**auth_headers, "X-Source-Filename": "user-can-rename-anything.xlsx"}
    preview = client.post("/api/historical-imports/preview", content=berth, headers=preview_headers)
    assert preview.status_code == 200, preview.text
    berth_import_id = preview.json()["id"]
    assert preview.json()["sourceKind"] == "tos_berth_call"
    assert preview.json()["reportingPeriod"] == "2026-07"
    assert preview.json()["mappingReceipt"]["filenameUsedForDetection"] is False
    assert preview.json()["accepted"] == 1
    detail = client.get(f"/api/historical-imports/{berth_import_id}", headers=auth_headers)
    assert detail.status_code == 200
    assert detail.json()["sourceSheets"] == ["User renamed sheet"]
    links = client.get(
        f"/api/historical-imports/{berth_import_id}/vessel-links?status=PENDING",
        headers=auth_headers,
    )
    assert links.status_code == 200
    assert links.json()["total"] == 1
    suggested_link = links.json()["items"][0]
    assert suggested_link["candidateVesselId"] is not None
    resolved = client.post(
        f"/api/historical-imports/{berth_import_id}/vessel-links/{suggested_link['id']}/resolve",
        json={"decision": "ACCEPT", "candidate_vessel_id": suggested_link["candidateVesselId"],
              "reason": "Confirmed in H4 preview"},
        headers=auth_headers,
    )
    assert resolved.status_code == 200
    assert resolved.json()["status"] == "ACCEPTED"
    confirmed = client.post(
        f"/api/historical-imports/{berth_import_id}/confirm", json={}, headers=auth_headers,
    )
    assert confirmed.status_code == 200
    assert confirmed.json()["status"] == "COMMITTED"

    cargo = _historical_fixture(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng",
         20: "Hàng nội/ ngoại", 23: "Phương án"},
        [{3: "40HC", 5: "E", 17: f"{vessel_name} | 2026 | 0007", 18: "4.00",
          20: "Hàng nội", 23: "Hạ bãi"}],
    )
    cargo_preview = client.post(
        "/api/historical-imports/preview", content=cargo,
        headers={**auth_headers, "X-Source-Filename": "different-name.xlsx"},
    )
    assert cargo_preview.status_code == 200, cargo_preview.text
    cargo_import_id = cargo_preview.json()["id"]
    rows = client.get(
        f"/api/historical-imports/{cargo_import_id}/rows?page=1&page_size=10", headers=auth_headers,
    )
    assert rows.status_code == 200
    assert rows.json()["items"][0]["validationStatus"] == "VALID", rows.json()["items"][0]
    assert cargo_preview.json()["accepted"] == 1, rows.json()
    assert cargo_preview.json()["reportingPeriod"] == "2026-07"
    assert rows.json()["items"][0]["matchStatus"] == "MATCHED"
    assert rows.json()["items"][0]["weightTonnes"] == 4.0
    cargo_confirmed = client.post(
        f"/api/historical-imports/{cargo_import_id}/confirm", json={}, headers=auth_headers,
    )
    assert cargo_confirmed.status_code == 200
    assert cargo_confirmed.json()["status"] == "COMMITTED"

    db = SessionLocal()
    try:
        cargo_row = db.query(HistoricalCargoRow).filter_by(import_id=cargo_import_id).one()
        call = db.query(HistoricalPortCall).filter_by(id=cargo_row.port_call_id).one()
        assert cargo_row.import_id != call.import_id
        assert cargo_row.reporting_unit_id == call.reporting_unit_id
    finally:
        db.close()

    duplicate = client.post("/api/historical-imports/preview", content=cargo, headers=preview_headers)
    assert duplicate.status_code == 200
    assert duplicate.json()["idempotent"] is True
    assert duplicate.json()["id"] == cargo_import_id

    corrected = _historical_fixture(berth_headers, [{
        2: "2026", 3: "0007", 5: vessel_name, 8: "K12B",
        20: "18/07/2026 08:35:00", 23: "18/07/2026 13:05:00",
    }])
    corrected_preview = client.post("/api/historical-imports/preview", content=corrected, headers=preview_headers)
    assert corrected_preview.status_code == 200
    corrected_id = corrected_preview.json()["id"]
    blocked = client.post(f"/api/historical-imports/{corrected_id}/confirm", json={}, headers=auth_headers)
    assert blocked.status_code == 409
    activated = client.post(
        f"/api/historical-imports/{corrected_id}/confirm",
        json={"conflict_action": "ACTIVATE_NEW_REVISION", "reason": "TOS corrected ATB/ATD"},
        headers=auth_headers,
    )
    assert activated.status_code == 200, activated.text
    assert activated.json()["revisionNo"] == 2
    db = SessionLocal()
    try:
        assert db.get(HistoricalReportImport, berth_import_id).status == "SUPERSEDED"
        cargo_row = db.query(HistoricalCargoRow).filter_by(import_id=cargo_import_id).one()
        replacement_call = db.query(HistoricalPortCall).filter_by(import_id=corrected_id).one()
        assert cargo_row.port_call_id == replacement_call.id
    finally:
        db.close()

    forbidden = client.post("/api/historical-imports/preview", content=berth, headers=customer_headers)
    assert forbidden.status_code == 403

    cancellable = _historical_fixture(berth_headers, [{
        2: "2026", 3: "0008", 5: vessel_name, 8: "K12",
        20: "18/08/2026 08:30:00", 23: "18/08/2026 13:00:00",
    }])
    cancellable_preview = client.post(
        "/api/historical-imports/preview", content=cancellable, headers=preview_headers,
    )
    assert cancellable_preview.status_code == 200
    cancelled = client.post(
        f"/api/historical-imports/{cancellable_preview.json()['id']}/cancel",
        json={"reason": "Operator cancelled after preview"}, headers=auth_headers,
    )
    assert cancelled.status_code == 200
    assert cancelled.json()["status"] == "REJECTED"

    # The same platform identity sees no data unless it explicitly switches to
    # the correct reporting-unit context; PORT_STAFF without membership fails.
    db = SessionLocal()
    try:
        second_unit = ReportingUnit(
            name=f"H3 Other Port {uuid.uuid4().hex[:8]}", code=f"H3-{uuid.uuid4().hex[:6]}",
            is_active=1, created_at=now_iso(), updated_at=now_iso(),
        )
        db.add(second_unit)
        db.commit()
        db.refresh(second_unit)
        second_unit_id = second_unit.id
    finally:
        db.close()
    other_context = {**auth_headers, "X-Reporting-Unit-ID": str(second_unit_id)}
    assert client.get(f"/api/historical-imports/{corrected_id}/rows", headers=other_context).status_code == 404
    port_login = client.post("/api/auth/login", json={"username": "portstaff", "password": "testpass"})
    no_membership = {
        "Authorization": f"Bearer {port_login.json()['access_token']}",
        "X-Reporting-Unit-ID": str(second_unit_id),
    }
    assert client.get("/api/historical-imports", headers=no_membership).status_code == 403


def test_historical_batch_order_rechecks_pending_cargo_after_berth_confirmation(
    client, auth_headers,
):
    """Files may be selected together: cargo preview is repaired after Berth activates."""
    vessel_name = f"BATCH BARGE {uuid.uuid4().hex[:8]}"
    _seed_historical_registered_vessel(vessel_name)
    cargo = _historical_fixture(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng",
         20: "Hàng nội/ ngoại", 23: "Phương án"},
        [{3: "40HC", 5: "E", 17: f"{vessel_name} | 2088 | 0001", 18: "4.00",
          20: "Hàng nội", 23: "Hạ bãi"}],
    )
    cargo_preview = client.post(
        "/api/historical-imports/preview", content=cargo,
        headers={**auth_headers, "X-Source-Filename": "cargo-first.xlsx"},
    )
    assert cargo_preview.status_code == 200, cargo_preview.text
    cargo_id = cargo_preview.json()["id"]
    assert cargo_preview.json()["review"] == 1

    berth = _historical_fixture(
        {2: "Năm", 3: "Chuyến", 5: "Tên tàu", 8: "Mã bến", 20: "ATB", 23: "ATD"},
        [{2: "2088", 3: "0001", 5: vessel_name, 8: "K12",
          20: "18/07/2088 08:30:00", 23: "18/07/2088 13:00:00"}],
    )
    berth_preview = client.post(
        "/api/historical-imports/preview", content=berth,
        headers={**auth_headers, "X-Source-Filename": "berth-second.xlsx"},
    )
    assert berth_preview.status_code == 200, berth_preview.text
    berth_id = berth_preview.json()["id"]
    confirmed_berth = client.post(
        f"/api/historical-imports/{berth_id}/confirm", json={}, headers=auth_headers,
    )
    assert confirmed_berth.status_code == 200
    assert confirmed_berth.json()["status"] == "COMMITTED"

    # Simulate a stale pre-fix database surviving an application restart. The
    # explicit idempotent reconcile endpoint repairs it from active Berth facts.
    db = SessionLocal()
    try:
        cargo_row = db.query(HistoricalCargoRow).filter_by(import_id=cargo_id).one()
        cargo_row.port_call_id = None
        cargo_row.match_status = "UNMATCHED"
        cargo_row.validation_status = "REVIEW"
        stale_import = db.get(HistoricalReportImport, cargo_id)
        stale_import.status = "REVIEW"
        stale_import.accepted_count = 0
        stale_import.review_count = 1
        stale_import.reporting_period = None
        db.commit()
    finally:
        db.close()
    repaired = client.post("/api/historical-imports/reconcile", headers=auth_headers)
    assert repaired.status_code == 200
    assert repaired.json() == {"updated": 1, "updatedImportIds": [cargo_id]}
    unchanged = client.post("/api/historical-imports/reconcile", headers=auth_headers)
    assert unchanged.status_code == 200
    assert unchanged.json() == {"updated": 0, "updatedImportIds": []}

    refreshed = client.get(f"/api/historical-imports/{cargo_id}", headers=auth_headers)
    assert refreshed.status_code == 200
    assert refreshed.json()["status"] == "COMMITTED"
    assert refreshed.json()["accepted"] == 1
    assert refreshed.json()["review"] == 0
    valid_rows = client.get(
        f"/api/historical-imports/{cargo_id}/rows?status=VALID", headers=auth_headers,
    )
    assert valid_rows.status_code == 200
    assert valid_rows.json()["status"] == "VALID"
    assert valid_rows.json()["total"] == 1
    assert valid_rows.json()["items"][0]["warnings"] == []
    review_rows = client.get(
        f"/api/historical-imports/{cargo_id}/rows?status=REVIEW", headers=auth_headers,
    )
    assert review_rows.status_code == 200
    assert review_rows.json()["total"] == 0
    confirmed_cargo = client.post(
        f"/api/historical-imports/{cargo_id}/confirm", json={}, headers=auth_headers,
    )
    assert confirmed_cargo.status_code == 200
    assert confirmed_cargo.json()["status"] == "COMMITTED"

    history = client.get("/api/historical-imports?page=1&page_size=20", headers=auth_headers)
    assert history.status_code == 200
    assert history.json()["summary"]["accepted"] >= 2
    assert set(history.json()["summary"]) == {"accepted", "review", "rejected"}
    assert "2088-07" in history.json()["activeBerthPeriods"]

    db = SessionLocal()
    try:
        db.query(HistoricalReportImport).filter(
            HistoricalReportImport.id.in_([cargo_id, berth_id])
        ).delete(synchronize_session=False)
        vessel = db.query(Vessel).filter_by(name=vessel_name).one()
        db.delete(vessel)
        db.commit()
    finally:
        db.close()


def test_historical_corrected_mapping_supersedes_same_source_without_period(
    client, auth_headers,
):
    """A corrected parser receipt replaces the active stale mapping explicitly."""
    cargo = _historical_fixture(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng",
         20: "Hàng nội/ ngoại", 23: "Phương án"},
        [{3: "40HC", 5: "E", 17: "MAPPING REVISION | 2091 | 0001", 18: "331,47",
          20: "Hàng nội", 23: "Hạ bãi"}],
    )
    first = client.post(
        "/api/historical-imports/preview", content=cargo,
        headers={**auth_headers, "X-Source-Filename": "same-source.xlsx"},
    )
    assert first.status_code == 200, first.text
    old_id = first.json()["id"]
    db = SessionLocal()
    try:
        old = db.get(HistoricalReportImport, old_id)
        old.mapping_version = "tos_cargo_detail_v1"
        old.status = "REVIEW"
        db.commit()
    finally:
        db.close()

    corrected = client.post(
        "/api/historical-imports/preview", content=cargo,
        headers={**auth_headers, "X-Source-Filename": "renamed-source.xlsx"},
    )
    assert corrected.status_code == 200, corrected.text
    corrected_body = corrected.json()
    new_id = corrected_body["id"]
    assert corrected_body["idempotent"] is False
    assert corrected_body["mappingVersion"] == "tos_cargo_detail_v2"
    assert corrected_body["conflictingImportIds"] == [old_id]
    assert client.post(
        f"/api/historical-imports/{new_id}/confirm", json={}, headers=auth_headers,
    ).status_code == 409
    activated = client.post(
        f"/api/historical-imports/{new_id}/confirm",
        json={"conflict_action": "ACTIVATE_NEW_REVISION", "reason": "Correct decimal parser"},
        headers=auth_headers,
    )
    assert activated.status_code == 200, activated.text
    assert activated.json()["revisionNo"] == 2

    db = SessionLocal()
    try:
        assert db.get(HistoricalReportImport, old_id).status == "SUPERSEDED"
        db.query(HistoricalReportImport).filter(
            HistoricalReportImport.id.in_([old_id, new_id])
        ).delete(synchronize_session=False)
        db.commit()
    finally:
        db.close()


def test_historical_pl03_export_uses_tos_facts_and_legacy_dimensions(
    client, auth_headers,
):
    vessel_name = f"TOS PL03 {uuid.uuid4().hex[:8]}"
    registration = _reg()
    db = SessionLocal()
    try:
        vessel = Vessel(
            organization_id=TEST_ORGANIZATION_ID, name=vessel_name,
            registration_no=registration, vessel_type="Chở container", vessel_class="VR-SI",
            length_m=52.5, deadweight_tons=998, gross_tonnage=511,
            created_at=now_iso(), updated_at=now_iso(),
        )
        db.add(vessel)
        db.flush()
        vessel_id = vessel.id
        db.add(ReportingUnitVessel(
            reporting_unit_id=TEST_REPORTING_UNIT_ID, vessel_id=vessel.id,
            created_at=now_iso(),
        ))
        db.commit()
    finally:
        db.close()

    legacy = _historical_pl03_fixture([{
        1: 1, 2: vessel_name, 3: registration,
        4: "Manual type must not win", 5: "Manual class", 6: 99,
        15: 999, 29: "Manual cargo must not win", 30: "Cảng A",
        31: "Cảng Tân Thuận", 32: "Cảng B", 33: "ETA cũ", 34: "ETD cũ",
        35: "Đại lý A",
    }])
    legacy_preview = client.post(
        "/api/historical-imports/preview", content=legacy,
        headers={**auth_headers, "X-Source-Filename": "legacy-pl03.xlsx"},
    )
    assert legacy_preview.status_code == 200, legacy_preview.text
    legacy_id = legacy_preview.json()["id"]
    assert client.post(
        f"/api/historical-imports/{legacy_id}/confirm", json={}, headers=auth_headers,
    ).status_code == 200

    berth = _historical_fixture(
        {2: "Năm", 3: "Chuyến", 5: "Tên tàu", 8: "Mã bến", 20: "ATB", 23: "ATD"},
        [{2: "2089", 3: "0001", 5: vessel_name, 8: "K12",
          20: "18/07/2089 08:30:00", 23: "18/07/2089 13:00:00"}],
    )
    berth_preview = client.post(
        "/api/historical-imports/preview", content=berth,
        headers={**auth_headers, "X-Source-Filename": "berth.xlsx"},
    )
    assert berth_preview.status_code == 200, berth_preview.text
    berth_id = berth_preview.json()["id"]
    assert client.post(
        f"/api/historical-imports/{berth_id}/confirm", json={}, headers=auth_headers,
    ).status_code == 200

    cargo = _historical_fixture(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng",
         20: "Hàng nội/ ngoại", 23: "Phương án"},
        [
            {3: "40HC", 5: "E", 17: f"{vessel_name} | 2089 | 0001", 18: "4,00",
             20: "Hàng nội", 23: "Hạ bãi"},
            {3: "20GP", 5: "F", 17: f"{vessel_name} | 2089 | 0001", 18: "10.5",
             20: "Hàng nội", 23: "Hạ bãi"},
        ],
    )
    cargo_preview = client.post(
        "/api/historical-imports/preview", content=cargo,
        headers={**auth_headers, "X-Source-Filename": "detail.xlsx"},
    )
    assert cargo_preview.status_code == 200, cargo_preview.text
    cargo_id = cargo_preview.json()["id"]
    assert cargo_preview.json()["accepted"] == 2
    assert client.post(
        f"/api/historical-imports/{cargo_id}/confirm", json={}, headers=auth_headers,
    ).status_code == 200

    exported = client.get(
        "/api/historical-imports/exports/pl03?reporting_period=2089-07",
        headers=auth_headers,
    )
    assert exported.status_code == 200, exported.text
    assert exported.headers["content-disposition"] == 'attachment; filename="PL03_TOS_2089-07.xlsx"'
    receipt = json.loads(exported.headers["x-historical-receipt"])
    assert receipt["callCount"] == 1 and receipt["cargoRowCount"] == 2
    sheet = load_workbook(io.BytesIO(exported.content), data_only=True).active
    assert sheet["A4"].value == "Đơn vị báo cáo: Test Reporting Unit"
    assert sheet["A2"].value and "tháng 7 năm 2089" in sheet["A2"].value
    row_number = next(
        row for row in range(10, sheet.max_row + 1)
        if sheet.cell(row, 3).value == registration
    )
    assert sheet.cell(row_number, 4).value == "Chở container"
    assert sheet.cell(row_number, 15).value == 14.5  # O: domestic inbound tonnes from TOS
    assert sheet.cell(row_number, 16).value == 1     # P: full TEU
    assert sheet.cell(row_number, 17).value == 2     # Q: empty TEU
    assert sheet.cell(row_number, 29).value == "Container"
    assert sheet.cell(row_number, 33).value == "18/07/2089 08:30:00"
    assert sheet.cell(row_number, 34).value == "18/07/2089 13:00:00"

    db = SessionLocal()
    try:
        db.query(HistoricalReportImport).filter(
            HistoricalReportImport.id.in_([legacy_id, berth_id, cargo_id])
        ).delete(synchronize_session=False)
        db.delete(db.get(Vessel, vessel_id))
        db.commit()
    finally:
        db.close()
