"""
tests/test_rbac.py — T1 Role-Based Access Control and Tenant Isolation Test Suite
WO-KBCV-T1-20260711
"""
from __future__ import annotations

import os
import tempfile
import time
import zipfile
from datetime import timedelta
from io import BytesIO
from pathlib import Path

# ── Set test DB FIRST, before any backend import ──────────────────────────────
_tmp_dir = tempfile.mkdtemp()
_test_db_path = Path(_tmp_dir) / "test_rbac.db"
os.environ["TEST_DATABASE_URL"] = f"sqlite:///{_test_db_path}"

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from alembic import command
from alembic.config import Config
from sqlalchemy import create_engine, inspect, text
from sqlalchemy.orm import Session

from backend.models import (
    Base, User, Organization, Vessel, CrewMember, Declaration, AuditEvent,
    ReportingUnit, ReportingUnitOrganization, ReportingUnitUser, ReportingUnitVessel,
)
from backend.database import engine, SessionLocal, now_iso
from backend.auth import create_access_token, get_password_hash
from backend.app import app, get_db

# ── Create all tables in test DB ──────────────────────────────────────────────
Base.metadata.create_all(bind=engine)

def _override_get_db():
    db = SessionLocal()
    try:
        yield db
    finally:
        db.close()

app.dependency_overrides[get_db] = _override_get_db

# Seed data
db = SessionLocal()
try:
    # 1. Seed Organizations
    org_a = Organization(name="Organization A", tax_code="TAXA", created_at=now_iso(), updated_at=now_iso())
    org_b = Organization(name="Organization B", tax_code="TAXB", created_at=now_iso(), updated_at=now_iso())
    db.add_all([org_a, org_b])
    db.commit()
    db.refresh(org_a)
    db.refresh(org_b)

    # 2. Seed Users
    users = [
        User(username="admin", password_hash=get_password_hash("adminpass"), full_name="Admin User", role="PLATFORM_ADMIN", is_active=1),
        User(username="cust_a", password_hash=get_password_hash("custpass"), full_name="Customer A", role="CUSTOMER", organization_id=org_a.id, is_active=1),
        User(username="cust_b", password_hash=get_password_hash("custpass"), full_name="Customer B", role="CUSTOMER", organization_id=org_b.id, is_active=1),
        User(username="cust_no_org", password_hash=get_password_hash("custpass"), full_name="No Org Cust", role="CUSTOMER", organization_id=None, is_active=1),
        User(username="disabled_user", password_hash=get_password_hash("pass"), full_name="Disabled", role="CUSTOMER", organization_id=org_a.id, is_active=0),
        User(username="user_port", password_hash=get_password_hash("portpass"), full_name="Port Employee", role="PORT_STAFF", is_active=1),
        User(username="user_port_b", password_hash=get_password_hash("portpass"), full_name="Port B Employee", role="PORT_STAFF", is_active=1),
        User(username="user_port_none", password_hash=get_password_hash("portpass"), full_name="Unassigned Port Employee", role="PORT_STAFF", is_active=1),
    ]
    db.add_all(users)
    db.commit()

    unit_a = ReportingUnit(
        name="Reporting Unit A", code="UNIT-A", is_active=1,
        created_at=now_iso(), updated_at=now_iso(),
    )
    unit_b = ReportingUnit(
        name="Reporting Unit B", code="UNIT-B", is_active=1,
        created_at=now_iso(), updated_at=now_iso(),
    )
    unit_inactive = ReportingUnit(
        name="Inactive Reporting Unit", code="UNIT-INACTIVE", is_active=0,
        created_at=now_iso(), updated_at=now_iso(),
    )
    db.add_all([unit_a, unit_b, unit_inactive])
    db.flush()
    port_user = db.query(User).filter(User.username == "user_port").one()
    port_user_b = db.query(User).filter(User.username == "user_port_b").one()
    db.add_all([
        ReportingUnitOrganization(
            reporting_unit_id=unit_a.id, organization_id=org_a.id, created_at=now_iso(),
        ),
        ReportingUnitOrganization(
            reporting_unit_id=unit_b.id, organization_id=org_b.id, created_at=now_iso(),
        ),
        ReportingUnitUser(
            reporting_unit_id=unit_a.id, user_id=port_user.id, created_at=now_iso(),
        ),
        ReportingUnitUser(
            reporting_unit_id=unit_b.id, user_id=port_user_b.id, created_at=now_iso(),
        ),
    ])
    db.commit()
    TEST_UNIT_A_ID = unit_a.id
    TEST_UNIT_B_ID = unit_b.id
    TEST_UNIT_INACTIVE_ID = unit_inactive.id

    # 3. Seed Vessel for Org A
    vessel_a = Vessel(
        organization_id=org_a.id,
        name="Vessel Org A",
        registration_no="REG-A-123",
        vessel_type="Tàu container",
        vessel_class="VR-SI",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    # Seed Vessel for Org B
    vessel_b = Vessel(
        organization_id=org_b.id,
        name="Vessel Org B",
        registration_no="REG-B-123",
        vessel_type="Tàu container",
        vessel_class="VR-SI",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    db.add_all([vessel_a, vessel_b])
    db.commit()
    db.refresh(vessel_a)
    db.refresh(vessel_b)
    db.add_all([
        ReportingUnitVessel(
            reporting_unit_id=TEST_UNIT_A_ID, vessel_id=vessel_a.id, created_at=now_iso(),
        ),
        ReportingUnitVessel(
            reporting_unit_id=TEST_UNIT_B_ID, vessel_id=vessel_b.id, created_at=now_iso(),
        ),
    ])
    db.commit()

    # 4. Seed Crew for Org A
    crew_a = CrewMember(
        organization_id=org_a.id,
        vessel_id=vessel_a.id,
        full_name="Crew Member A",
        crew_role="Thủy thủ",
        professional_certificate_type="An toàn",
        professional_certificate_no="CERT-A",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    db.add(crew_a)
    db.commit()
    db.refresh(crew_a)

    # 5. Seed Declarations
    # Draft Declaration for Org A
    decl_a_draft = Declaration(
        organization_id=org_a.id,
        vessel_id=vessel_a.id,
        company_name=org_a.name,
        declaration_date="2026-07-11",
        vessel_name=vessel_a.name,
        registration_no=vessel_a.registration_no,
        vessel_type=vessel_a.vessel_type,
        vessel_class=vessel_a.vessel_class,
        last_port="Port A",
        working_port="Port B",
        eta="2026-07-11T12:00",
        etd="2026-07-11T20:00",
        master_name="Captain A",
        master_phone="0901",
        unload_json="{}",
        load_json="{}",
        workflow_status="DRAFT",
        status="DRAFT",
        reference_no="REF-A-DRAFT-1",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    # Submitted Declaration for Org A
    decl_a_sub = Declaration(
        organization_id=org_a.id,
        vessel_id=vessel_a.id,
        company_name=org_a.name,
        declaration_date="2026-07-11",
        vessel_name=vessel_a.name,
        registration_no=vessel_a.registration_no,
        vessel_type=vessel_a.vessel_type,
        vessel_class=vessel_a.vessel_class,
        last_port="Port A",
        working_port="Port B",
        eta="2026-07-11T12:00",
        etd="2026-07-11T20:00",
        master_name="Captain A",
        master_phone="0901",
        unload_json="{}",
        load_json="{}",
        workflow_status="PENDING_REVIEW",
        status="SUBMITTED",
        reference_no="REF-A-SUB-1",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    decl_b_sub = Declaration(
        organization_id=org_b.id,
        vessel_id=vessel_b.id,
        company_name=org_b.name,
        declaration_date="2026-07-11",
        vessel_name=vessel_b.name,
        registration_no=vessel_b.registration_no,
        vessel_type=vessel_b.vessel_type,
        vessel_class=vessel_b.vessel_class,
        last_port="Port B",
        working_port="Port C",
        eta="2026-07-11T12:00",
        etd="2026-07-11T20:00",
        master_name="Captain B",
        master_phone="0902",
        unload_json="{}",
        load_json="{}",
        workflow_status="PENDING_REVIEW",
        status="SUBMITTED",
        reference_no="REF-B-SUB-1",
        created_at=now_iso(),
        updated_at=now_iso(),
    )
    db.add_all([decl_a_draft, decl_a_sub, decl_b_sub])
    db.commit()
finally:
    db.close()


@pytest.fixture(scope="module")
def client():
    with TestClient(app) as c:
        yield c

# Token headers helper
def get_auth_header(client, username, password, *, unit_id="default"):
    res = client.post("/api/auth/login", json={"username": username, "password": password})
    if res.status_code == 200:
        headers = {"Authorization": f"Bearer {res.json()['access_token']}"}
        if unit_id == "default" and username in {"admin", "user_port", "user_port_none"}:
            headers["X-Reporting-Unit-ID"] = str(TEST_UNIT_A_ID)
        elif unit_id == "default" and username == "user_port_b":
            headers["X-Reporting-Unit-ID"] = str(TEST_UNIT_B_ID)
        elif unit_id is not None and unit_id != "default":
            headers["X-Reporting-Unit-ID"] = str(unit_id)
        return headers
    return {}

# ══════════════════════════════════════════════════════════════════════════════
# 1. AUTHENTICATION & SECURITY CONTROLS
# ══════════════════════════════════════════════════════════════════════════════

def test_disabled_user_rejected(client):
    """Disabled users must not be allowed to log in or access endpoints."""
    headers = get_auth_header(client, "disabled_user", "pass")
    assert headers == {}  # Login returns 403 Forbidden directly or fails

    # Verify endpoint rejection for disabled users
    res = client.post("/api/auth/login", json={"username": "disabled_user", "password": "pass"})
    assert res.status_code == 403
    assert "Tài khoản đã bị vô hiệu hóa" in res.json()["detail"]

def test_missing_org_fails_closed(client):
    """Customer without organization must fail closed when accessing resources."""
    headers = get_auth_header(client, "cust_no_org", "custpass")
    assert headers != {}
    res = client.get("/api/vessels", headers=headers)
    assert res.status_code == 403
    assert "chưa được liên kết với tổ chức" in res.json()["detail"]

def test_login_rate_limiting(client):
    """Failed login attempts block IP after 5 attempts."""
    # Ensure fresh state
    from backend.app import _login_attempts
    _login_attempts.clear()

    # Try wrong logins 5 times
    for _ in range(5):
        res = client.post("/api/auth/login", json={"username": "cust_a", "password": "wrongpassword"})
        assert res.status_code == 401

    # Sixth attempt should be rate limited with 429
    res = client.post("/api/auth/login", json={"username": "cust_a", "password": "wrongpassword"})
    assert res.status_code == 429
    assert "tạm khóa" in res.json()["detail"]

    # Reset for other tests
    _login_attempts.clear()

# ══════════════════════════════════════════════════════════════════════════════
# 2. TENANT ISOLATION (CUSTOMER A vs CUSTOMER B)
# ══════════════════════════════════════════════════════════════════════════════

def test_tenant_isolation_vessel(client):
    """Customer A cannot read or write Customer B's vessels."""
    headers_a = get_auth_header(client, "cust_a", "custpass")
    headers_b = get_auth_header(client, "cust_b", "custpass")

    # Get B's vessel ID
    db = SessionLocal()
    vessel_b = db.query(Vessel).filter(Vessel.name == "Vessel Org B").first()
    vessel_b_id = vessel_b.id
    db.close()

    # Customer A tries to update Customer B's vessel -> should fail (403 or 404 depending on verify)
    res = client.post("/api/vessels", headers=headers_a, json={
        "id": vessel_b_id,
        "name": "Hacked name",
        "registration_no": "REG-B-123",
        "vessel_type": "Tàu container",
        "vessel_class": "VR-SI"
    })
    assert res.status_code == 403

    # Customer A tries to verify registry of Customer B's vessel -> 403
    res = client.post(f"/api/vessels/{vessel_b_id}/verify-registry", headers=headers_a)
    assert res.status_code == 403

def test_tenant_isolation_declaration(client):
    """Customer B cannot read or write Customer A's declarations."""
    headers_b = get_auth_header(client, "cust_b", "custpass")

    db = SessionLocal()
    decl_a = db.query(Declaration).filter(Declaration.reference_no == "REF-A-DRAFT-1").first()
    decl_a_id = decl_a.id
    db.close()

    # Get events of A's declaration -> 403
    res = client.get(f"/api/declarations/{decl_a_id}/events", headers=headers_b)
    assert res.status_code == 403

    # Upload attachment to A's declaration -> 403
    res = client.post(f"/api/declarations/{decl_a_id}/attachments?filename=test.jpg", headers=headers_b, content=b"fakecontent")
    assert res.status_code == 403


def test_customer_scope_covers_lists_suggestions_crew_and_reports(client):
    """A customer sees only its own tenant's data across read endpoints."""
    headers_a = get_auth_header(client, "cust_a", "custpass")
    headers_b = get_auth_header(client, "cust_b", "custpass")

    decls_b = client.get("/api/declarations", headers=headers_b)
    assert decls_b.status_code == 200
    assert {item["reference_no"] for item in decls_b.json()} == {"REF-B-SUB-1"}

    suggestions_a = client.get("/api/suggestions?field=master_name", headers=headers_a)
    suggestions_b = client.get("/api/suggestions?field=master_name", headers=headers_b)
    assert "Captain A" in suggestions_a.json()
    assert "Captain B" not in suggestions_a.json()
    assert "Captain B" in suggestions_b.json()
    assert "Captain A" not in suggestions_b.json()

    crew_a = client.get("/api/crew", headers=headers_a)
    crew_b = client.get("/api/crew", headers=headers_b)
    assert crew_a.status_code == crew_b.status_code == 200
    assert [member["full_name"] for member in crew_a.json()] == ["Crew Member A"]
    assert crew_b.json() == []

    report_a = client.get("/api/reports/appendix1", headers=headers_a)
    assert report_a.status_code == 200
    with zipfile.ZipFile(BytesIO(report_a.content)) as archive:
        worksheet = archive.read("xl/worksheets/sheet1.xml").decode("utf-8")
    assert "REF-A-SUB-1" not in worksheet  # pending review is not report-eligible
    assert "REF-B-SUB-1" not in worksheet


def test_customer_cannot_update_other_tenant_crew(client):
    """Sequential crew IDs cannot be used to write another tenant's record."""
    headers_b = get_auth_header(client, "cust_b", "custpass")
    db = SessionLocal()
    crew_a = db.query(CrewMember).filter(CrewMember.full_name == "Crew Member A").first()
    db.close()

    response = client.post("/api/crew", headers=headers_b, json={
        "id": crew_a.id,
        "full_name": "Overwritten",
        "crew_role": "Thuyền viên",
        "professional_certificate_type": "An toàn",
        "professional_certificate_no": "CERT-A",
    })
    assert response.status_code == 403


# ══════════════════════════════════════════════════════════════════════════════
# 2B. REPORTING-UNIT CONTEXT (R4)
# ══════════════════════════════════════════════════════════════════════════════

def test_port_staff_requires_fk_membership_and_unit_list_is_scoped(client):
    unassigned = get_auth_header(client, "user_port_none", "portpass")
    assert client.get("/api/vessels", headers=unassigned).status_code == 403

    staff_a = get_auth_header(client, "user_port", "portpass")
    units = client.get("/api/reporting-units", headers=staff_a)
    assert units.status_code == 200
    assert [item["id"] for item in units.json()["items"]] == [TEST_UNIT_A_ID]
    organizations = client.get("/api/reporting-unit/organizations", headers=staff_a)
    assert organizations.status_code == 200
    assert [item["name"] for item in organizations.json()["items"]] == ["Organization A"]


def test_platform_admin_context_is_explicit_and_validated(client):
    no_context = get_auth_header(client, "admin", "adminpass", unit_id=None)
    assert client.get("/api/vessels", headers=no_context).status_code == 400

    malformed = {**no_context, "X-Reporting-Unit-ID": "not-an-id"}
    assert client.get("/api/vessels", headers=malformed).status_code == 400

    missing = {**no_context, "X-Reporting-Unit-ID": "999999"}
    assert client.get("/api/vessels", headers=missing).status_code == 404

    inactive = {**no_context, "X-Reporting-Unit-ID": str(TEST_UNIT_INACTIVE_ID)}
    assert client.get("/api/vessels", headers=inactive).status_code == 403


def test_live_reads_dashboard_and_register_are_isolated_by_unit(client):
    admin_a = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_A_ID)
    admin_b = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_B_ID)

    vessels_a = client.get("/api/vessels", headers=admin_a)
    vessels_b = client.get("/api/vessels", headers=admin_b)
    assert {item["registration_no"] for item in vessels_a.json()} == {"REG-A-123"}
    assert {item["registration_no"] for item in vessels_b.json()} == {"REG-B-123"}

    dashboard_a = client.get("/api/dashboard?q=REG-B-123", headers=admin_a).json()
    dashboard_b = client.get("/api/dashboard?q=REG-A-123", headers=admin_b).json()
    assert dashboard_a["stats"]["vessels"] == 1 and dashboard_a["matches"] == []
    assert dashboard_b["stats"]["vessels"] == 1 and dashboard_b["matches"] == []

    register_a = client.get("/api/port-vessel-register", headers=admin_a).json()["items"]
    register_b = client.get("/api/port-vessel-register", headers=admin_b).json()["items"]
    assert {item["registration_no"] for item in register_a} == {"REG-A-123"}
    assert {item["registration_no"] for item in register_b} == {"REG-B-123"}


def test_port_register_export_and_remove_cannot_cross_units(client):
    staff_a = get_auth_header(client, "user_port", "portpass")
    exported = client.get("/api/port-vessel-register/export", headers=staff_a)
    assert exported.status_code == 200
    sheet = load_workbook(BytesIO(exported.content), read_only=True).active
    values = {str(cell.value) for row in sheet.iter_rows() for cell in row if cell.value is not None}
    assert "REG-A-123" in values
    assert "REG-B-123" not in values

    db = SessionLocal()
    vessel_b_id = db.query(Vessel.id).filter(Vessel.registration_no == "REG-B-123").scalar()
    db.close()
    denied = client.post(
        "/api/port-vessel-register/remove", headers=staff_a, json={"ids": [vessel_b_id]},
    )
    assert denied.status_code == 404


def test_same_vessel_register_membership_is_independent_between_units(client):
    admin_a = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_A_ID)
    admin_b = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_B_ID)
    db = SessionLocal()
    vessel_a_id = db.query(Vessel.id).filter(Vessel.registration_no == "REG-A-123").scalar()
    db.close()

    added = client.post(
        "/api/port-vessel-register/add", headers=admin_b, json={"ids": [vessel_a_id]},
    )
    assert added.status_code == 200 and added.json()["added"] == 1
    assert {item["registration_no"] for item in client.get(
        "/api/port-vessel-register", headers=admin_b,
    ).json()["items"]} == {"REG-A-123", "REG-B-123"}
    assert {item["registration_no"] for item in client.get(
        "/api/port-vessel-register", headers=admin_a,
    ).json()["items"]} == {"REG-A-123"}

    removed = client.post(
        "/api/port-vessel-register/remove", headers=admin_b, json={"ids": [vessel_a_id]},
    )
    assert removed.status_code == 200 and removed.json()["removed"] == 1
    assert {item["registration_no"] for item in client.get(
        "/api/port-vessel-register", headers=admin_a,
    ).json()["items"]} == {"REG-A-123"}


def test_cross_unit_mutation_and_workflow_are_rejected(client):
    admin_a = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_A_ID)
    db = SessionLocal()
    vessel_b = db.query(Vessel).filter(Vessel.registration_no == "REG-B-123").one()
    declaration_b = db.query(Declaration).filter(Declaration.reference_no == "REF-B-SUB-1").one()
    vessel_b_id, declaration_b_id = vessel_b.id, declaration_b.id
    db.close()

    changed = client.post("/api/vessels", headers=admin_a, json={
        "id": vessel_b_id,
        "name": "Cross-tenant overwrite",
        "registration_no": "REG-B-123",
        "vessel_type": "Tàu container",
        "vessel_class": "VR-SI",
        "organization_name": "Organization B",
    })
    assert changed.status_code == 403

    workflow = client.post(
        f"/api/declarations/{declaration_b_id}/workflow",
        headers=admin_a,
        json={"action": "PORT_APPROVE"},
    )
    assert workflow.status_code == 403

    create_for_foreign_org = client.post("/api/vessels", headers=admin_a, json={
        "name": "Foreign Org Vessel",
        "registration_no": "FOREIGN-UNIT-NEW",
        "vessel_type": "Sà lan",
        "vessel_class": "VR-SI",
        "organization_name": "Organization B",
    })
    assert create_for_foreign_org.status_code == 403


def test_tenant_mutation_audit_records_reporting_unit(client):
    admin_a = get_auth_header(client, "admin", "adminpass", unit_id=TEST_UNIT_A_ID)
    registration = f"AUDIT-{time.time_ns()}"
    created = client.post("/api/vessels?port_register=true", headers=admin_a, json={
        "name": "Tenant Audit Vessel",
        "registration_no": registration,
        "vessel_type": "Sà lan",
        "vessel_class": "VR-SI",
        "organization_name": "Organization A",
    })
    assert created.status_code == 200, created.text
    vessel_id = created.json()["id"]

    db = SessionLocal()
    try:
        events = db.query(AuditEvent).filter(
            AuditEvent.entity_type == "VESSEL",
            AuditEvent.entity_id == vessel_id,
            AuditEvent.action == "CREATE",
            AuditEvent.summary.contains(registration),
        ).all()
        assert events
        assert {event.reporting_unit_id for event in events} == {TEST_UNIT_A_ID}
        db.query(ReportingUnitVessel).filter_by(vessel_id=vessel_id).delete()
        for event in events:
            db.delete(event)
        db.query(Vessel).filter(Vessel.id == vessel_id).delete()
        db.commit()
    finally:
        db.close()


def test_customer_cannot_access_internal_port_register(client):
    customer = get_auth_header(client, "cust_a", "custpass")
    assert client.get("/api/port-vessel-register", headers=customer).status_code == 403

# ══════════════════════════════════════════════════════════════════════════════
# 3. ROLE-BASED ACCESS CONTROL (RBAC MATRIX)
# ══════════════════════════════════════════════════════════════════════════════

def test_workflow_transitions_matrix(client):
    """Port staff and platform support in explicit context may review."""
    headers_port = get_auth_header(client, "user_port", "portpass")
    headers_customer = get_auth_header(client, "cust_a", "custpass")
    headers_admin = get_auth_header(client, "admin", "adminpass")

    db = SessionLocal()
    decl = db.query(Declaration).filter(Declaration.reference_no == "REF-A-SUB-1").first()
    decl_id = decl.id
    db.close()

    res = client.post(f"/api/declarations/{decl_id}/workflow", headers=headers_customer, json={
        "action": "PORT_APPROVE"
    })
    assert res.status_code == 403

    res = client.post(
        f"/api/declarations/{decl_id}/workflow",
        headers=headers_port,
        json={"action": "CV_APPROVE"},
    )
    assert res.status_code == 410

    res = client.post(f"/api/declarations/{decl_id}/workflow", headers=headers_admin, json={
        "action": "PORT_APPROVE"
    })
    assert res.status_code == 200
    assert res.json()["workflow_status"] == "APPROVED"

def test_actor_identity_audit_protection(client):
    """Verify that client-supplied actor fields are ignored and derived from JWT."""
    headers_port = get_auth_header(client, "user_port", "portpass")

    # We will query another declaration for Org A (let's create a submitted declaration)
    db = SessionLocal()
    org_a = db.query(Organization).filter(Organization.name == "Organization A").first()
    vessel_a = db.query(Vessel).filter(Vessel.organization_id == org_a.id).first()
    decl_2 = Declaration(
        organization_id=org_a.id,
        vessel_id=vessel_a.id,
        company_name=org_a.name,
        declaration_date="2026-07-11",
        vessel_name=vessel_a.name,
        registration_no=vessel_a.registration_no,
        vessel_type=vessel_a.vessel_type,
        vessel_class=vessel_a.vessel_class,
        last_port="Port A",
        working_port="Port B",
        eta="2026-07-11T12:00",
        etd="2026-07-11T20:00",
        master_name="Captain A",
        master_phone="0901",
        unload_json="{}",
        load_json="{}",
        workflow_status="PENDING_REVIEW",
        status="SUBMITTED",
        reference_no="REF-A-SUB-AUDIT-TEST",
        created_at=now_iso(),
        updated_at=now_iso()
    )
    db.add(decl_2)
    db.commit()
    decl_id = decl_2.id
    db.close()

    # Perform workflow transition and inject fake role & name
    res = client.post(f"/api/declarations/{decl_id}/workflow", headers=headers_port, json={
        "action": "PORT_APPROVE",
        "actor_role": "ADMIN",  # Fake role
        "actor_name": "Hacker Officer"  # Fake name
    })
    assert res.status_code == 200

    # Retrieve events and verify the actual actor from DB
    res_events = client.get(f"/api/declarations/{decl_id}/events", headers=headers_port)
    assert res_events.status_code == 200
    events = res_events.json()
    port_approve_event = [e for e in events if e["action"] == "PORT_APPROVE"][0]

    assert port_approve_event["actor_role"] == "PORT_STAFF"
    assert port_approve_event["actor_name"] == "Port Employee"


def test_unknown_role_and_expired_token_fail_closed(client):
    """Invalid stored roles and expired tokens never grant endpoint access."""
    db = SessionLocal()
    invalid = User(
        username="invalid_role",
        password_hash=get_password_hash("invalidpass"),
        full_name="Invalid Role",
        role="SUPERUSER",
        is_active=1,
    )
    db.add(invalid)
    db.commit()
    db.close()

    invalid_headers = get_auth_header(client, "invalid_role", "invalidpass")
    assert invalid_headers
    assert client.get("/api/vessels", headers=invalid_headers).status_code == 403

    expired = create_access_token(
        {"sub": "cust_a", "role": "CUSTOMER", "org_id": 1},
        expires_delta=timedelta(seconds=-1),
    )
    expired_response = client.get(
        "/api/auth/me", headers={"Authorization": f"Bearer {expired}"}
    )
    assert expired_response.status_code == 401

# ══════════════════════════════════════════════════════════════════════════════
# 4. FAIL-FAST CONFIGURATION TEST
# ══════════════════════════════════════════════════════════════════════════════

def test_fail_fast_outside_test_mode():
    """Verify that auth module triggers fail-fast check outside local test/db mode."""
    # CI supplies both variables globally. Remove both so this test actually
    # exercises the unsafe production-default branch, then restore them exactly.
    original_db_url = os.environ.pop("TEST_DATABASE_URL", None)
    original_secret = os.environ.pop("SECRET_KEY", None)
    try:
        # Reloading auth inside try-except should fail with SystemExit
        import importlib
        import backend.auth
        with pytest.raises(SystemExit):
            importlib.reload(backend.auth)
    finally:
        # Restore environment variable
        if original_db_url:
            os.environ["TEST_DATABASE_URL"] = original_db_url
        if original_secret:
            os.environ["SECRET_KEY"] = original_secret
        importlib.reload(backend.auth)


def test_alembic_t1_upgrade_and_downgrade_rehearsal(monkeypatch, tmp_path):
    """The T1 migration upgrades a legacy user table and reverses cleanly."""
    legacy_db = tmp_path / "legacy_t0.db"
    url = f"sqlite:///{legacy_db}"
    legacy_engine = create_engine(url)
    with legacy_engine.begin() as connection:
        connection.execute(text("""
            CREATE TABLE organizations (
                id INTEGER PRIMARY KEY,
                name VARCHAR NOT NULL UNIQUE,
                tax_code VARCHAR NOT NULL DEFAULT '',
                address VARCHAR NOT NULL DEFAULT '',
                contact_name VARCHAR NOT NULL DEFAULT '',
                contact_role VARCHAR NOT NULL DEFAULT '',
                phone VARCHAR NOT NULL DEFAULT '',
                email VARCHAR NOT NULL DEFAULT '',
                created_at VARCHAR NOT NULL,
                updated_at VARCHAR NOT NULL
            )
        """))
        connection.execute(text("""
            CREATE TABLE users (
                id INTEGER PRIMARY KEY,
                username VARCHAR NOT NULL UNIQUE,
                password_hash VARCHAR NOT NULL,
                full_name VARCHAR DEFAULT '',
                role VARCHAR DEFAULT 'CUSTOMER',
                created_at VARCHAR
            )
        """))
        connection.execute(text("""
            INSERT INTO users(id, username, password_hash, full_name, role)
            VALUES (1, 'legacy_customer', 'hash', 'Legacy Customer', 'CUSTOMER')
        """))

    import backend.database as database
    monkeypatch.setattr(database, "SQLALCHEMY_DATABASE_URL", url)
    config = Config(str(Path(__file__).resolve().parents[1] / "alembic.ini"))
    command.upgrade(config, "head")

    inspector = inspect(legacy_engine)
    columns = {column["name"] for column in inspector.get_columns("users")}
    assert {"organization_id", "is_active"}.issubset(columns)
    with legacy_engine.connect() as connection:
        assert connection.execute(text("SELECT is_active FROM users WHERE id=1")).scalar_one() == 0

    command.downgrade(config, "base")
    columns_after = {column["name"] for column in inspect(legacy_engine).get_columns("users")}
    assert "organization_id" not in columns_after
    assert "is_active" not in columns_after


def test_alembic_fresh_database_reaches_t2_head(monkeypatch, tmp_path):
    """A fresh local database reaches the current governed schema head."""
    fresh_db = tmp_path / "fresh.db"
    url = f"sqlite:///{fresh_db}"
    import backend.database as database

    monkeypatch.setattr(database, "SQLALCHEMY_DATABASE_URL", url)
    config = Config(str(Path(__file__).resolve().parents[1] / "alembic.ini"))
    command.upgrade(config, "head")

    inspector = inspect(create_engine(url))
    assert {"users", "organizations", "declarations", "audit_events"}.issubset(
        set(inspector.get_table_names())
    )
    declaration_columns = {column["name"] for column in inspector.get_columns("declarations")}
    assert "version" in declaration_columns
    assert "port_approval" in declaration_columns
    assert "cv_approval" not in declaration_columns
    assert "correlation_id" in {column["name"] for column in inspector.get_columns("audit_events")}
    assert "birth_date" in {column["name"] for column in inspector.get_columns("crew_members")}
    vessel_columns = {column["name"] for column in inspector.get_columns("vessels")}
    assert {"is_port_tracked", "port_tracking_updated_at"}.issubset(vessel_columns)
    assert "vessel_operating_profiles" in inspector.get_table_names()
