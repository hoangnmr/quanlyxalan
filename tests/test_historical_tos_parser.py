from __future__ import annotations

import io

import pytest
from openpyxl import Workbook

from backend.historical_tos_parser import HistoricalWorkbookError, parse_workbook


def _xlsx(headers: dict[int, str], rows: list[dict[int, object]], *, title: str = "Anything") -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    for column, value in headers.items():
        sheet.cell(1, column, value)
    for row_number, row in enumerate(rows, 2):
        for column, value in row.items():
            sheet.cell(row_number, column, value)
    output = io.BytesIO()
    workbook.save(output)
    return output.getvalue()


def test_detects_berth_by_headers_and_preserves_leading_zero_voyage():
    content = _xlsx(
        {2: "Năm", 3: "Chuyến", 5: "Tên tàu", 8: "Mã bến", 20: "ATB", 23: "ATD"},
        [{2: "2026", 3: "0007", 5: "SÀ LAN-01", 8: "K12", 20: "18/07/2026 08:30:00", 23: "18/07/2026 13:00:00"}],
        title="Renamed by user",
    )
    parsed = parse_workbook(content)
    assert parsed.source_kind == "tos_berth_call"
    assert parsed.reporting_period == "2026-07"
    assert parsed.rows[0]["voyage_number_raw"] == "0007"
    assert parsed.rows[0]["call_key_normalized"].endswith("|2026|7")
    assert parsed.rows[0]["actual_berthing_at"] == "2026-07-18T08:30:00"


def test_cargo_dimensions_are_independent_and_hidden_rows_are_read():
    content = _xlsx(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng", 20: "Hàng nội/ ngoại", 23: "Phương án"},
        [{3: "40HC", 5: "E", 17: "SÀ LAN-01 | 2026 | 0007", 18: "4.00", 20: "Hàng nội", 23: "Hạ bãi"}],
    )
    parsed = parse_workbook(content)
    row = parsed.rows[0]
    assert row["teu_factor"] == 2
    assert row["full_empty_code_raw"] == "E"
    assert row["weight_tonnes"] == 4.0
    assert row["derived_direction"] == "unload"
    assert row["trade_scope"] == "domestic"


def test_unknown_size_and_invalid_weight_enter_review_not_zero():
    content = _xlsx(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng", 20: "Hàng nội/ ngoại", 23: "Phương án"},
        [{3: "45HC", 5: "F", 17: "A | 2026 | 1", 18: "n/a", 20: "Hàng ngoại", 23: "Lấy Nguyên"}],
    )
    row = parse_workbook(content).rows[0]
    assert row["validation_status"] == "REVIEW"
    assert row["teu_factor"] is None
    assert row["weight_tonnes"] is None
    assert row["weight_state"] == "INVALID"


def test_decimal_parser_accepts_vietnamese_and_mixed_excel_number_formats():
    content = _xlsx(
        {3: "Kích cỡ", 5: "F/E", 17: "Tên sà lan | Năm | Chuyến", 18: "Trọng lượng", 20: "Hàng nội/ ngoại", 23: "Phương án"},
        [
            {3: "40HC", 5: "F", 17: "A | 2026 | 1", 18: "331,47", 20: "Hàng nội", 23: "Hạ bãi"},
            {3: "40HC", 5: "F", 17: "B | 2026 | 2", 18: "1,088.84", 20: "Hàng nội", 23: "Hạ bãi"},
            {3: "40HC", 5: "F", 17: "C | 2026 | 3", 18: "1.088,84", 20: "Hàng nội", 23: "Hạ bãi"},
        ],
    )
    parsed = parse_workbook(content)
    rows = parsed.rows
    assert parsed.mapping_version == "tos_cargo_detail_v2"
    assert [row["weight_tonnes"] for row in rows] == [331.47, 1088.84, 1088.84]
    assert all(row["validation_status"] == "VALID" for row in rows)


def test_pl03_decimal_comma_is_a_valid_tonnage_metric():
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "PL03"
    sheet.cell(5, 1, "STT")
    sheet.cell(5, 2, "Tên PTTND")
    sheet.cell(5, 3, "Số đăng ký")
    sheet.cell(10, 1, 1)
    sheet.cell(10, 2, "PHƯỚC TẠO 12")
    sheet.cell(10, 3, "SG.10249")
    sheet.cell(10, 15, "331,47")
    sheet.cell(10, 16, 80)
    sheet.cell(10, 35, None)
    output = io.BytesIO()
    workbook.save(output)

    parsed = parse_workbook(output.getvalue())
    assert parsed.mapping_version == "reported_pl03_35col_historical_v2"
    row = parsed.rows[0]
    inbound_tons = next(
        metric for metric in row["metrics"]
        if metric["metric_code"] == "domestic_inbound_tons_reported"
    )
    assert inbound_tons["numeric_value"] == 331.47
    assert inbound_tons["invalid"] is False
    assert row["validation_status"] == "VALID"


def test_unknown_workbook_fails_closed():
    with pytest.raises(HistoricalWorkbookError, match="Không nhận diện"):
        parse_workbook(_xlsx({1: "unrelated"}, [{1: "value"}]))
